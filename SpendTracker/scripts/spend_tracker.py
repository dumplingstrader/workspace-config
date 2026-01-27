#!/usr/bin/env python3
"""
Marathon Petroleum - LA Refinery Process Control Spend Tracker
Automates consolidation of ME2K exports and TS Actuals data

Author: Process Control Team (developed with AI assistance)
Version: 2.0
Date: January 2026

USAGE:
    python spend_tracker.py

REQUIREMENTS:
    - Python 3.8+
    - pandas
    - openpyxl

INSTALL DEPENDENCIES:
    pip install pandas openpyxl --break-system-packages
"""

import os
import sys
import shutil
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print("ERROR: Missing required packages.")
    print("Please run: pip install pandas openpyxl --break-system-packages")
    sys.exit(1)


# =============================================================================
# CONFIGURATION
# =============================================================================

# Determine base folder (parent of scripts folder)
SCRIPT_DIR = Path(__file__).resolve().parent
BASE_FOLDER = SCRIPT_DIR.parent

# Folder structure
DATA_FOLDER = BASE_FOLDER / "data"
ME2K_FOLDER = DATA_FOLDER / "me2k"
TS_ACTUALS_FOLDER = DATA_FOLDER / "ts_actuals"
OUTPUT_FOLDER = BASE_FOLDER / "output"
TEMPLATE_FILE = BASE_FOLDER / "SpendTracker_Template.xlsx"

# Internal Order mapping for LA Refinery Process Control
INTERNAL_ORDERS = {
    "9909811": "PC - Software & Contracts",
    "9909813": "PC - Outside Support (Contractors)",
    "9909815": "PC - Hardware & Materials",
    "9909818": "PC - APC Work & Maint",
    "9912360": "PC - Dept Misc Exp & Travel",
    "9913117": "Watson Cogen",
}

# Legacy to Unify order number mapping
LEGACY_TO_UNIFY = {
    "99001110": "9909811",
    "99001108": "9909813",
    "99001106": "9909815",
    "99001103": "9909818",
    "99001107": "9909814",
    "99001109": "9909812",
    "99001105": "9909816",
    "99001104": "9909817",
}

# Watson Cogen specific
WATSON_COGEN_ORDER = "9913117"
WATSON_COGEN_PLANT = "R190"

# File naming patterns
ME2K_PATTERN = "EXPORT_*.xlsx"
TS_ACTUALS_PATTERN = "*ts_actual*.xlsx"

# Cross-charge document types
CROSS_CHARGE_TYPES = [
    'G/L account document',
    'G/L Accrual Posting',
    'CO Posting',
    'Account Maintenance'
]


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def ensure_folders_exist():
    """Create working folders if they do not exist."""
    ME2K_FOLDER.mkdir(parents=True, exist_ok=True)
    TS_ACTUALS_FOLDER.mkdir(parents=True, exist_ok=True)
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    print(f"Base folder: {BASE_FOLDER}")
    print(f"ME2K folder: {ME2K_FOLDER}")
    print(f"TS Actuals folder: {TS_ACTUALS_FOLDER}")
    print(f"Output folder: {OUTPUT_FOLDER}")


def find_latest_file(folder, pattern):
    """Find the most recently modified file matching pattern."""
    files = list(folder.glob(pattern))
    if not files:
        return None
    return max(files, key=lambda f: f.stat().st_mtime)


def safe_string(value):
    """Convert value to ASCII-safe string."""
    if pd.isna(value):
        return ""
    s = str(value)
    return s.encode('ascii', 'replace').decode('ascii')


def parse_date(date_val):
    """Safely parse date values from various formats."""
    if pd.isna(date_val):
        return None
    if isinstance(date_val, datetime):
        return date_val
    try:
        return pd.to_datetime(date_val)
    except:
        return None


def extract_fiscal_year(date_val):
    """Extract fiscal year from date."""
    dt = parse_date(date_val)
    if dt is None:
        return None
    return dt.year


def extract_month(date_val):
    """Extract month from date."""
    dt = parse_date(date_val)
    if dt is None:
        return None
    return dt.month


def extract_quarter(date_val):
    """Extract quarter from date."""
    dt = parse_date(date_val)
    if dt is None:
        return None
    return (dt.month - 1) // 3 + 1


def normalize_io_number(io_num):
    """Normalize Internal Order number (convert legacy to Unify if needed)."""
    if pd.isna(io_num):
        return ""
    io_str = str(io_num).strip()
    # Check if it's a legacy number
    if io_str in LEGACY_TO_UNIFY:
        return LEGACY_TO_UNIFY[io_str]
    return io_str


def get_io_description(io_num):
    """Get description for Internal Order number."""
    io_str = normalize_io_number(io_num)
    return INTERNAL_ORDERS.get(io_str, "")


# =============================================================================
# DATA LOADING FUNCTIONS
# =============================================================================

def load_me2k_export(filepath):
    """Load and process ME2K export file."""
    print(f"\nLoading ME2K export: {filepath.name}")

    try:
        df = pd.read_excel(filepath, header=0)
    except Exception as e:
        print(f"ERROR: Could not read ME2K file: {e}")
        return None

    print(f"  Loaded {len(df)} rows")
    df.columns = df.columns.str.strip()

    # Standardize to common schema
    records = []
    for _, row in df.iterrows():
        # Parse date
        doc_date = parse_date(row.get('Document Date'))

        # Get IO number
        io_raw = row.get('Order', '')
        io_num = normalize_io_number(io_raw)

        record = {
            'Fiscal_Year': extract_fiscal_year(doc_date),
            'Month': extract_month(doc_date),
            'Quarter': extract_quarter(doc_date),
            'Posting_Date': doc_date,
            'IO_Number': io_num,
            'IO_Description': get_io_description(io_num),
            'Vendor': safe_string(row.get('Name of Supplier', row.get('Supplier/Supplying Plant', ''))),
            'Vendor_Number': safe_string(row.get('Supplier/Supplying Plant', '')),
            'PO_Number': safe_string(row.get('Purchasing Document', '')),
            'Description': safe_string(row.get('Short Text', '')),
            'PO_Value': pd.to_numeric(row.get('Net Order Value', 0), errors='coerce') or 0,
            'GR_Value': pd.to_numeric(row.get('Net Order Value', 0), errors='coerce') or 0,
            'Open_Amount': pd.to_numeric(row.get('Still to be invoiced (val.)', 0), errors='coerce') or 0,
            'Is_Cross_Charge': False,
            'Doc_Type': 'PO Commitment',
            'Source': 'ME2K',
            'Classification': '',  # User can fill in
            'Status': '',  # User can fill in (Scheduled/Completed)
            'Comment': '',  # User can add notes
        }
        records.append(record)

    return pd.DataFrame(records)


def load_ts_actuals(filepath):
    """Load and process TS Actuals file."""
    print(f"\nLoading TS Actuals: {filepath.name}")

    try:
        xlsx = pd.ExcelFile(filepath)
        sheet_name = None
        for name in xlsx.sheet_names:
            if 'actual' in name.lower():
                sheet_name = name
                break
        if sheet_name is None:
            sheet_name = xlsx.sheet_names[0]
            if sheet_name.startswith('_com.sap'):
                sheet_name = xlsx.sheet_names[1] if len(xlsx.sheet_names) > 1 else None

        if sheet_name is None:
            print("  ERROR: Could not find data sheet")
            return None

        df = pd.read_excel(filepath, sheet_name=sheet_name, header=0)
    except Exception as e:
        print(f"ERROR: Could not read TS Actuals file: {e}")
        return None

    print(f"  Loaded {len(df)} rows from sheet '{sheet_name}'")
    df.columns = df.columns.str.strip()

    # Standardize to common schema
    records = []
    for _, row in df.iterrows():
        # Parse date
        post_date = parse_date(row.get('Posting Date'))

        # Get IO number
        io_raw = row.get('Order', '')
        io_num = normalize_io_number(io_raw)

        # Check if cross-charge
        doc_type = str(row.get('Acctg Doc Type', ''))
        doc_header = str(row.get('Doc Header Text', ''))
        is_cross_charge = (
            doc_type in CROSS_CHARGE_TYPES or
            any(kw in doc_header.upper() for kw in ['ALLOCATION', 'ALLOC', 'CROSS'])
        )

        record = {
            'Fiscal_Year': row.get('Fiscal Year', extract_fiscal_year(post_date)),
            'Month': row.get('Month', extract_month(post_date)),
            'Quarter': extract_quarter(post_date),
            'Posting_Date': post_date,
            'IO_Number': io_num,
            'IO_Description': row.get('Order Desc', get_io_description(io_num)),
            'Vendor': safe_string(row.get('Vendor Name', '')),
            'Vendor_Number': safe_string(row.get('Vendor', '')),
            'PO_Number': safe_string(row.get('Purchase Order', '')),
            'Description': safe_string(row.get('Line Item Text', row.get('Doc Header Text', ''))),
            'PO_Value': 0,  # TS Actuals are actuals, not PO values
            'GR_Value': pd.to_numeric(row.get('Actuals', 0), errors='coerce') or 0,
            'Open_Amount': 0,
            'Is_Cross_Charge': is_cross_charge,
            'Doc_Type': safe_string(doc_type),
            'Source': 'TS_Actuals',
            'Classification': '',
            'Status': '',
            'Comment': '',
        }
        records.append(record)

    return pd.DataFrame(records)


# =============================================================================
# DATA CONSOLIDATION
# =============================================================================

def consolidate_data(me2k_df, ts_actuals_df):
    """Consolidate ME2K and TS Actuals into unified dataset."""
    frames = []

    if me2k_df is not None and len(me2k_df) > 0:
        frames.append(me2k_df)
        print(f"  ME2K records: {len(me2k_df)}")

    if ts_actuals_df is not None and len(ts_actuals_df) > 0:
        frames.append(ts_actuals_df)
        print(f"  TS Actuals records: {len(ts_actuals_df)}")

    if not frames:
        return None

    consolidated = pd.concat(frames, ignore_index=True)

    # Sort by date descending, then by IO
    consolidated = consolidated.sort_values(
        ['Fiscal_Year', 'Month', 'IO_Number', 'Posting_Date'],
        ascending=[False, False, True, False]
    )

    print(f"  Total consolidated records: {len(consolidated)}")

    # Summary stats
    cross_charges = consolidated[consolidated['Is_Cross_Charge'] == True]
    if len(cross_charges) > 0:
        print(f"  Cross-charges identified: {len(cross_charges)} (${cross_charges['GR_Value'].sum():,.2f})")

    return consolidated


# =============================================================================
# EXCEL OUTPUT FUNCTIONS
# =============================================================================

def create_template_workbook():
    """Create the template workbook with sheets for pivot tables."""
    wb = Workbook()

    # ----- Data Sheet -----
    ws_data = wb.active
    ws_data.title = "Data"

    # Column headers for the Data table
    headers = [
        'Fiscal_Year', 'Month', 'Quarter', 'Posting_Date',
        'IO_Number', 'IO_Description', 'Vendor', 'Vendor_Number',
        'PO_Number', 'Description', 'PO_Value', 'GR_Value', 'Open_Amount',
        'Is_Cross_Charge', 'Doc_Type', 'Source', 'Classification', 'Status', 'Comment'
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    # Add sample row so table can be created
    ws_data.cell(row=2, column=1, value="(Data will be populated by script)")

    # ----- Pivot Sheet (instructions) -----
    ws_pivot = wb.create_sheet("Pivot")
    ws_pivot.cell(row=1, column=1, value="PIVOT TABLE INSTRUCTIONS")
    ws_pivot.cell(row=1, column=1).font = Font(bold=True, size=14)

    instructions = [
        "",
        "To create pivot tables from the Data sheet:",
        "",
        "1. Go to the Data sheet",
        "2. Select any cell in the data table",
        "3. Insert > PivotTable",
        "4. Choose 'New Worksheet' or place on this sheet",
        "",
        "Recommended Pivot Table configurations:",
        "",
        "SPEND BY IO AND YEAR:",
        "  - Rows: IO_Number, IO_Description",
        "  - Columns: Fiscal_Year",
        "  - Values: Sum of GR_Value",
        "  - Filter: Source = 'TS_Actuals'",
        "",
        "MONTHLY TREND:",
        "  - Rows: Fiscal_Year, Month",
        "  - Columns: IO_Number",
        "  - Values: Sum of GR_Value",
        "",
        "CROSS-CHARGES:",
        "  - Filter: Is_Cross_Charge = TRUE",
        "  - Rows: IO_Number, Vendor, Description",
        "  - Values: Sum of GR_Value",
        "",
        "PO COMMITMENTS:",
        "  - Filter: Source = 'ME2K'",
        "  - Rows: IO_Number, Vendor, PO_Number",
        "  - Values: Sum of PO_Value, Sum of Open_Amount",
        "",
        "Add Slicers:",
        "  - PivotTable Analyze > Insert Slicer",
        "  - Select: Fiscal_Year, IO_Number, Source, Is_Cross_Charge, Classification",
    ]

    for row, text in enumerate(instructions, 2):
        ws_pivot.cell(row=row, column=1, value=text)

    ws_pivot.column_dimensions['A'].width = 80

    # ----- Budget Sheet -----
    ws_budget = wb.create_sheet("Budget")
    ws_budget.cell(row=1, column=1, value="BUDGET TRACKING")
    ws_budget.cell(row=1, column=1).font = Font(bold=True, size=14)

    budget_headers = ['IO_Number', 'IO_Description', 'Annual_Budget', 'YTD_Actuals', 'Remaining', 'Variance_%']
    for col, header in enumerate(budget_headers, 1):
        cell = ws_budget.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    # Pre-populate IO numbers
    row = 4
    for io_num, io_desc in INTERNAL_ORDERS.items():
        ws_budget.cell(row=row, column=1, value=io_num)
        ws_budget.cell(row=row, column=2, value=io_desc)
        ws_budget.cell(row=row, column=3, value=0)  # Budget - user fills in
        ws_budget.cell(row=row, column=4, value=f"=SUMIFS(Data!L:L,Data!E:E,A{row},Data!P:P,\"TS_Actuals\")")
        ws_budget.cell(row=row, column=5, value=f"=C{row}-D{row}")
        ws_budget.cell(row=row, column=6, value=f"=IF(C{row}=0,0,(C{row}-D{row})/C{row})")
        ws_budget.cell(row=row, column=6).number_format = '0.0%'
        row += 1

    # Adjust column widths
    ws_budget.column_dimensions['A'].width = 12
    ws_budget.column_dimensions['B'].width = 35
    ws_budget.column_dimensions['C'].width = 15
    ws_budget.column_dimensions['D'].width = 15
    ws_budget.column_dimensions['E'].width = 15
    ws_budget.column_dimensions['F'].width = 12

    # ----- IO Reference Sheet -----
    ws_io = wb.create_sheet("IO_Reference")
    ws_io.cell(row=1, column=1, value="INTERNAL ORDER REFERENCE")
    ws_io.cell(row=1, column=1).font = Font(bold=True, size=14)

    io_headers = ['Unify_Order', 'Legacy_Order', 'Description', 'Budget_Owner']
    for col, header in enumerate(io_headers, 1):
        cell = ws_io.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    # IO mapping data
    io_mapping = [
        ('9909811', '99001110', 'PC - Software & Contracts', 'Process Control'),
        ('9909813', '99001108', 'PC - Outside Support (Contractors)', 'Process Control'),
        ('9909815', '99001106', 'PC - Hardware & Materials', 'Process Control'),
        ('9909818', '99001103', 'PC - APC Work & Maint', 'Process Control'),
        ('9912360', '', 'PC - Dept Misc Exp & Travel', 'Process Control'),
        ('9913117', '', 'Watson Cogen', 'Process Control'),
        ('9909814', '99001107', 'HW ISA Contract', 'Process Control'),
        ('9909812', '99001109', 'APC Optimizer Project', 'Process Control'),
        ('9909816', '99001105', 'DCS Maintenance LAW', 'Process Control'),
        ('9909817', '99001104', 'DCS Maintenance LAC', 'Process Control'),
    ]

    row = 4
    for unify, legacy, desc, owner in io_mapping:
        ws_io.cell(row=row, column=1, value=unify)
        ws_io.cell(row=row, column=2, value=legacy)
        ws_io.cell(row=row, column=3, value=desc)
        ws_io.cell(row=row, column=4, value=owner)
        row += 1

    ws_io.column_dimensions['A'].width = 15
    ws_io.column_dimensions['B'].width = 15
    ws_io.column_dimensions['C'].width = 35
    ws_io.column_dimensions['D'].width = 20

    # ----- Cross Charges Sheet -----
    ws_cc = wb.create_sheet("Cross_Charges")
    ws_cc.cell(row=1, column=1, value="CROSS-CHARGES (Auto-filtered from Data)")
    ws_cc.cell(row=1, column=1).font = Font(bold=True, size=14, color='FF0000')
    ws_cc.cell(row=2, column=1, value="These entries bypass the PO process - review monthly")
    ws_cc.cell(row=4, column=1, value="To view cross-charges:")
    ws_cc.cell(row=5, column=1, value="1. Go to Data sheet")
    ws_cc.cell(row=6, column=1, value="2. Filter Is_Cross_Charge column = TRUE")
    ws_cc.cell(row=7, column=1, value="3. Or create a PivotTable filtered on Is_Cross_Charge")

    return wb


def write_data_to_workbook(wb, data_df):
    """Write consolidated data to the Data sheet in workbook."""
    ws = wb['Data']

    # Clear existing data (keep header row)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    if data_df is None or len(data_df) == 0:
        ws.cell(row=2, column=1, value="No data available")
        return

    # Column order must match headers
    columns = [
        'Fiscal_Year', 'Month', 'Quarter', 'Posting_Date',
        'IO_Number', 'IO_Description', 'Vendor', 'Vendor_Number',
        'PO_Number', 'Description', 'PO_Value', 'GR_Value', 'Open_Amount',
        'Is_Cross_Charge', 'Doc_Type', 'Source', 'Classification', 'Status', 'Comment'
    ]

    # Write data rows
    for row_idx, (_, row) in enumerate(data_df.iterrows(), 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row.get(col_name, '')

            # Handle special types
            if col_name == 'Posting_Date' and pd.notna(value):
                if isinstance(value, str):
                    value = parse_date(value)
            elif col_name == 'Is_Cross_Charge':
                value = bool(value) if pd.notna(value) else False
            elif col_name in ['PO_Value', 'GR_Value', 'Open_Amount']:
                value = float(value) if pd.notna(value) else 0
            elif pd.isna(value):
                value = ''

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Format dates
            if col_name == 'Posting_Date' and value:
                cell.number_format = 'YYYY-MM-DD'
            # Format currency
            elif col_name in ['PO_Value', 'GR_Value', 'Open_Amount']:
                cell.number_format = '#,##0.00'

    # Create Excel Table
    last_row = len(data_df) + 1
    last_col = len(columns)
    table_range = f"A1:{get_column_letter(last_col)}{last_row}"

    # Remove existing table if present
    if 'SpendData' in [t.name for t in ws.tables.values()]:
        del ws.tables['SpendData']

    table = Table(displayName="SpendData", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Auto-adjust column widths
    column_widths = {
        'A': 12, 'B': 8, 'C': 8, 'D': 12,
        'E': 12, 'F': 35, 'G': 30, 'H': 12,
        'I': 15, 'J': 50, 'K': 12, 'L': 12, 'M': 12,
        'N': 14, 'O': 20, 'P': 12, 'Q': 15, 'R': 12, 'S': 30
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


def generate_summary(data_df):
    """Generate and print summary statistics."""
    if data_df is None or len(data_df) == 0:
        print("\nNo data to summarize.")
        return

    print("\n" + "=" * 60)
    print("SUMMARY BY INTERNAL ORDER")
    print("=" * 60)

    # Filter to TS Actuals for YTD spend
    actuals = data_df[data_df['Source'] == 'TS_Actuals']
    me2k = data_df[data_df['Source'] == 'ME2K']

    for io_num, io_desc in INTERNAL_ORDERS.items():
        io_actuals = actuals[actuals['IO_Number'] == io_num]['GR_Value'].sum()
        io_commitments = me2k[me2k['IO_Number'] == io_num]['PO_Value'].sum()
        io_open = me2k[me2k['IO_Number'] == io_num]['Open_Amount'].sum()
        io_cross = actuals[(actuals['IO_Number'] == io_num) & (actuals['Is_Cross_Charge'] == True)]['GR_Value'].sum()

        if io_actuals != 0 or io_commitments != 0:
            print(f"\n{io_num} - {io_desc}")
            print(f"  PO Commitments:    ${io_commitments:>12,.2f}")
            print(f"  Open (not invoiced): ${io_open:>12,.2f}")
            print(f"  Actuals YTD:       ${io_actuals:>12,.2f}")
            if io_cross != 0:
                print(f"  Cross-Charges:     ${io_cross:>12,.2f} ***")


# =============================================================================
# MAIN EXECUTION
# =============================================================================

def main():
    """Main execution function."""
    print("=" * 60)
    print("LA Refinery Process Control - Spend Tracker v2.0")
    print("=" * 60)
    print(f"Run time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # Setup folders
    ensure_folders_exist()

    # Check for template, create if missing
    if not TEMPLATE_FILE.exists():
        print(f"\nCreating template: {TEMPLATE_FILE.name}")
        template_wb = create_template_workbook()
        template_wb.save(TEMPLATE_FILE)
        print("  Template created successfully")

    # Find input files
    print("\n--- Locating Input Files ---")

    me2k_file = find_latest_file(ME2K_FOLDER, ME2K_PATTERN)
    if me2k_file:
        print(f"Found ME2K export: {me2k_file.name}")
    else:
        print(f"WARNING: No ME2K export found matching '{ME2K_PATTERN}' in {ME2K_FOLDER}")

    ts_actuals_file = find_latest_file(TS_ACTUALS_FOLDER, TS_ACTUALS_PATTERN)
    if ts_actuals_file:
        print(f"Found TS Actuals: {ts_actuals_file.name}")
    else:
        print(f"WARNING: No TS Actuals found matching '{TS_ACTUALS_PATTERN}' in {TS_ACTUALS_FOLDER}")

    # Load data
    print("\n--- Loading Data ---")
    me2k_df = load_me2k_export(me2k_file) if me2k_file else None
    ts_actuals_df = load_ts_actuals(ts_actuals_file) if ts_actuals_file else None

    if me2k_df is None and ts_actuals_df is None:
        print("\nERROR: No data files found. Please ensure files are in the correct location.")
        print(f"  - ME2K exports: {ME2K_FOLDER}")
        print(f"  - TS Actuals: {TS_ACTUALS_FOLDER}")
        sys.exit(1)

    # Consolidate data
    print("\n--- Consolidating Data ---")
    consolidated_df = consolidate_data(me2k_df, ts_actuals_df)

    # Generate output
    print("\n--- Generating Output ---")

    # Copy template to output file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f"SpendTracker_{timestamp}.xlsx"
    output_path = OUTPUT_FOLDER / output_filename

    shutil.copy(TEMPLATE_FILE, output_path)
    print(f"Copied template to: {output_filename}")

    # Open and populate with data
    wb = load_workbook(output_path)
    write_data_to_workbook(wb, consolidated_df)
    wb.save(output_path)

    print(f"Data written to: {output_path}")

    # Print summary
    generate_summary(consolidated_df)

    print("\n" + "=" * 60)
    print("Processing complete!")
    print("=" * 60)
    print(f"\nOutput file: {output_path}")
    print("\nNext steps:")
    print("  1. Open the output file in Excel")
    print("  2. Go to 'Data' sheet - your consolidated data is here")
    print("  3. Create PivotTables (see 'Pivot' sheet for instructions)")
    print("  4. Update 'Budget' sheet with your annual budget figures")
    print("  5. Add Classification/Status/Comments as needed")


if __name__ == "__main__":
    main()
