import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import glob

# Automatically find all training files (BC-LAR-ENGPRO*.xlsx)
training_folder = Path("Training")
training_files = sorted(glob.glob(str(training_folder / "BC-LAR-ENGPRO*.xlsx")))
print(f"Found {len(training_files)} training files to process")

# Dictionary to store all training data
all_training_data = []
course_data = {}

print("Processing training files...")
print("=" * 80)

for file in training_files:
    file_path = Path(file)
    if not file_path.exists():
        print(f"⚠️  File not found: {file}")
        continue
    
    course_name = file_path.stem
    print(f"\n📁 Processing: {course_name}")
    
    try:
        # Try to read enrollment sheet first, then assignment sheet
        df = None
        sheet_type = None
        
        try:
            df = pd.read_excel(file_path, sheet_name='View Learning Content Enrollmen')
            sheet_type = 'enrollment'
        except:
            try:
                df = pd.read_excel(file_path, sheet_name='View Learning Content Assignmen')
                sheet_type = 'assignment'
            except Exception as e:
                raise Exception(f"Could not find enrollment or assignment sheet: {e}")
        
        if df is not None:
            # Extract employee name and course name from the enrollment/learner field
            import re
            
            if sheet_type == 'enrollment':
                # Format: "Name - LAR 8901C Course Name (BC_LAR_ENGPRO001)"
                def extract_person(text):
                    if pd.isna(text):
                        return text
                    match = re.match(r'([^-]+)', str(text))
                    return match.group(1).strip() if match else text
                
                def extract_course_name(text):
                    if pd.isna(text):
                        return None
                    match = re.search(r'- (.*?) \(BC_LAR', str(text))
                    return match.group(1).strip() if match else None
                
                df['Employee_Name'] = df['Enrollment'].apply(extract_person)
                
                # Extract course name from first non-null enrollment
                first_enrollment = df['Enrollment'].dropna().iloc[0] if len(df['Enrollment'].dropna()) > 0 else None
                full_course_name = extract_course_name(first_enrollment) if first_enrollment else course_name
                
            else:  # assignment
                # Learner field should just be the name
                df['Employee_Name'] = df['Learner'] if 'Learner' in df.columns else df['Enrollment']
                full_course_name = course_name  # Will need to extract from Learning Assignment field if available
                
                if 'Learning Assignment' in df.columns:
                    first_assignment = df['Learning Assignment'].dropna().iloc[0] if len(df['Learning Assignment'].dropna()) > 0 else None
                    if first_assignment:
                        # Try different patterns for assignment format
                        # Pattern 1: "Course Name (BC_LAR_ENGPRO005) assigned to"
                        match = re.search(r'^(.*?) \(BC_LAR_ENGPRO\d+', str(first_assignment))
                        if not match:
                            # Pattern 2: "- Course Name (BC_LAR"
                            match = re.search(r'- (.*?) \(BC_LAR', str(first_assignment))
                        if match:
                            full_course_name = match.group(1).strip()
            
            # Store the full course name for later use
            if full_course_name:
                course_name = full_course_name
            
            # For files with session numbers (e.g., BC-LAR-ENGPRO008-1), add session to course name
            stem = file_path.stem
            session_match = re.search(r'-(\d+)$', stem)
            if session_match:
                session_num = session_match.group(1)
                course_name = f"{course_name} (Session {session_num})"
            
            # Normalize column names for assignments to match enrollments
            if sheet_type == 'assignment':
                column_mapping = {
                    'Learner': 'Enrollment',
                    'Assignment Status': 'Enrollment Status',
                    'Assigned Date': 'Enrollment Date'
                }
                df = df.rename(columns=column_mapping)
                
                # Add missing columns if needed
                if 'Completion Status' not in df.columns:
                    df['Completion Status'] = df['Enrollment Status'].apply(
                        lambda x: 'Complete' if x == 'Complete' else 'Incomplete'
                    )
                if 'Drop Date' not in df.columns:
                    df['Drop Date'] = None
                if 'Drop Reason' not in df.columns:
                    df['Drop Reason'] = None
            
            print(f"   Found {len(df)} {'enrollment' if sheet_type == 'enrollment' else 'assignment'} records")
            print(f"   Course: {course_name}")
            
            # Store for course-specific sheet
            course_data[course_name] = df
            
            # Add course name to each record and append to master list
            df['Training Course'] = course_name
            all_training_data.append(df)
            
            # Show sample data
            if len(df) > 0:
                if 'Enrollment Status' in df.columns:
                    status_counts = df['Enrollment Status'].value_counts()
                    print(f"   Status summary: {status_counts.to_dict()}")
    
    except Exception as e:
        print(f"   ❌ Error reading file: {e}")

print("\n" + "=" * 80)

if all_training_data:
    # Combine all dataframes
    df_master = pd.concat(all_training_data, ignore_index=True)
    
    print(f"\n✅ Total enrollment records: {len(df_master)}")
    
    # Create the tracking spreadsheet
    output_file = "Training/Training_Attendance_Tracker.xlsx"
    
    # Try to preserve existing Job Duty assignments if the file exists
    existing_job_duties = {}
    df_dcs_existing = None
    if Path(output_file).exists():
        try:
            df_existing = pd.read_excel(output_file, sheet_name='Employee-Course Matrix')
            if 'Job Duty' in df_existing.columns and 'Employee' in df_existing.columns:
                # Create a mapping of Employee -> Job Duty
                for _, row in df_existing.iterrows():
                    if pd.notna(row['Employee']) and pd.notna(row['Job Duty']) and str(row['Job Duty']).strip():
                        existing_job_duties[row['Employee']] = row['Job Duty']
                print(f"   Preserved {len(existing_job_duties)} existing Job Duty assignments")
        except Exception as e:
            print(f"   Note: Could not preserve Job Duty assignments: {e}")
        
        # Try to read existing DCS Training data
        try:
            df_dcs_existing = pd.read_excel(output_file, sheet_name='DCS Training')
            # Remove empty rows
            df_dcs_existing = df_dcs_existing.dropna(how='all')
            print(f"   Preserved {len(df_dcs_existing)} existing DCS Training records")
        except Exception as e:
            print(f"   Note: Could not preserve DCS Training data: {e}")
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write Summary sheet
        summary_data = {
            'Metric': [
                'Total Training Courses',
                'Total Enrollments',
                'Generated Date',
                '',
                'Training Courses:'
            ],
            'Value': [
                len(training_files),
                len(df_master),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                '',
                ''
            ]
        }
        
        # Add course list to summary
        for course in sorted(course_data.keys()):
            count = len(course_data[course])
            summary_data['Metric'].append(f'  • {course}')
            summary_data['Value'].append(f'{count} enrollments')
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Create DCS Training manual input sheet (preserve existing data)
        dcs_columns = ['Employee', 'Job Duty', 'Course Name', 'Training Date', 'Status', 'Instructor', 'Notes']
        if df_dcs_existing is not None and len(df_dcs_existing) > 0:
            df_dcs = df_dcs_existing
        else:
            df_dcs = pd.DataFrame(columns=dcs_columns)
        df_dcs.to_excel(writer, sheet_name='DCS Training', index=False)
        
        # Write Master Data sheet
        df_master.to_excel(writer, sheet_name='Master Data', index=False)
        
        # Create analysis sheets
        
        # 1. Employee Training Matrix (Pivot view)
        if 'Employee_Name' in df_master.columns and 'Training Course' in df_master.columns:
            # Create a more detailed status with date
            # Note: Enrolled = Completed (manual attendance tracking)
            # Use Completion Date if available, otherwise Enrollment Date
            def create_status_with_date(row):
                status = row['Enrollment Status']
                # Prefer Completion Date over Enrollment Date
                date = row.get('Completion Date') if pd.notna(row.get('Completion Date')) else row['Enrollment Date']
                
                if pd.notna(date):
                    date_obj = pd.to_datetime(date)
                    date_str = date_obj.strftime('%b %Y')  # Format as "Mar 2023"
                    if status == 'Enrolled' or status == 'Completed':
                        return f'✓ ({date_str})'
                    elif status == 'Dropped':
                        return '✗ Dropped'
                    else:
                        return f'{status} ({date_str})'
                else:
                    if status == 'Enrolled' or status == 'Completed':
                        return '✓'
                    elif status == 'Dropped':
                        return '✗ Dropped'
                    else:
                        return status
            
            df_master['Status_With_Date'] = df_master.apply(create_status_with_date, axis=1)
            
            # Create a pivot table showing which courses each employee has taken
            df_matrix = df_master.pivot_table(
                index='Employee_Name',
                columns='Training Course',
                values='Status_With_Date',
                aggfunc='first',
                fill_value=''
            )
            
            # Reset index to make employee name a column
            df_matrix.reset_index(inplace=True)
            df_matrix.rename(columns={'Employee_Name': 'Employee'}, inplace=True)
            
            # Add employees from DCS Training who aren't in the LAR courses
            if df_dcs_existing is not None and len(df_dcs_existing) > 0:
                dcs_employees = df_dcs_existing['Employee'].dropna().unique()
                matrix_employees = df_matrix['Employee'].tolist()
                
                # Find employees in DCS Training but not in matrix
                new_employees = [emp for emp in dcs_employees if emp not in matrix_employees]
                
                if new_employees:
                    # Create rows for new employees with empty values for all course columns
                    new_rows = []
                    for emp in new_employees:
                        new_row = {'Employee': emp}
                        # Add empty values for all existing course columns
                        for col in df_matrix.columns:
                            if col != 'Employee':
                                new_row[col] = ''
                        new_rows.append(new_row)
                    
                    # Append new employees to the matrix
                    df_new_employees = pd.DataFrame(new_rows)
                    df_matrix = pd.concat([df_matrix, df_new_employees], ignore_index=True)
            
            # Add Job Duty column as second column, preserving existing assignments
            if existing_job_duties:
                df_matrix.insert(1, 'Job Duty', df_matrix['Employee'].map(existing_job_duties).fillna(''))
            else:
                df_matrix.insert(1, 'Job Duty', '')
            
            # Add DCS Training column from manual input sheet
            if df_dcs_existing is not None and len(df_dcs_existing) > 0:
                # Group DCS training by employee - combine multiple courses into one cell
                dcs_by_employee = {}
                for _, row in df_dcs_existing.iterrows():
                    if pd.notna(row['Employee']) and pd.notna(row['Course Name']):
                        employee = row['Employee']
                        course = row['Course Name']
                        date = row.get('Training Date')
                        
                        # Format: Course Name (Mon YYYY) if date available
                        if pd.notna(date):
                            try:
                                date_obj = pd.to_datetime(date)
                                date_str = date_obj.strftime('%b %Y')
                                course_entry = f"{course} ({date_str})"
                            except:
                                course_entry = course
                        else:
                            course_entry = course
                        
                        # Append to existing courses for this employee
                        if employee in dcs_by_employee:
                            dcs_by_employee[employee] += f"\n{course_entry}"
                        else:
                            dcs_by_employee[employee] = course_entry
                
                # Add DCS Training column after Job Duty
                df_matrix.insert(2, 'DCS Training', df_matrix['Employee'].map(dcs_by_employee).fillna(''))
            
            df_matrix.to_excel(writer, sheet_name='Employee-Course Matrix', index=False)
            
            # Create a detailed training record per employee
            employee_records = []
            for employee in df_master['Employee_Name'].unique():
                emp_data = df_master[df_master['Employee_Name'] == employee].copy()
                courses_taken = emp_data['Training Course'].tolist()
                statuses = emp_data['Enrollment Status'].tolist()
                completion = emp_data['Completion Status'].tolist()
                enroll_dates = emp_data['Enrollment Date'].tolist()
                
                # Count Enrolled and Completed as completed (manual attendance tracking)
                completed_count = ((emp_data['Enrollment Status'] == 'Enrolled') | 
                                 (emp_data['Enrollment Status'] == 'Completed') |
                                 (emp_data['Completion Status'] == 'Complete')).sum()
                
                employee_records.append({
                    'Employee': employee,
                    'Total Courses': len(courses_taken),
                    'Completed': completed_count,
                    'Dropped': (emp_data['Enrollment Status'] == 'Dropped').sum(),
                    'Completion Rate %': (completed_count / len(courses_taken) * 100).round(1),
                    'Courses': ', '.join(courses_taken),
                    'Latest Enrollment': max([d for d in enroll_dates if pd.notna(d)]) if any(pd.notna(d) for d in enroll_dates) else None
                })
            
            df_employee_detail = pd.DataFrame(employee_records)
            df_employee_detail = df_employee_detail.sort_values('Total Courses', ascending=False)
            df_employee_detail.to_excel(writer, sheet_name='Employee Training Detail', index=False)
        
        # 3. Completion Timeline (if dates available)
        if 'Completion Date' in df_master.columns:
            df_timeline = df_master[df_master['Completion Date'].notna()].copy()
            if len(df_timeline) > 0:
                df_timeline['Completion Date'] = pd.to_datetime(df_timeline['Completion Date'])
                df_timeline = df_timeline.sort_values('Completion Date')
                df_timeline.to_excel(writer, sheet_name='Completion Timeline', index=False)
        
        # 5. Training Gap Analysis - Who hasn't taken which courses
        if 'Employee_Name' in df_master.columns:
            all_employees = df_master['Employee_Name'].unique()
            all_courses = df_master['Training Course'].unique()
            
            gap_records = []
            for employee in all_employees:
                emp_courses = df_master[df_master['Employee_Name'] == employee]['Training Course'].tolist()
                missing_courses = [c for c in all_courses if c not in emp_courses]
                
                if missing_courses:
                    gap_records.append({
                        'Employee': employee,
                        'Courses Taken': len(emp_courses),
                        'Courses Missing': len(missing_courses),
                        'Missing Course List': ', '.join(missing_courses)
                    })
            
            if gap_records:
                df_gaps = pd.DataFrame(gap_records)
                df_gaps = df_gaps.sort_values('Courses Missing', ascending=False)
                df_gaps.to_excel(writer, sheet_name='Training Gaps', index=False)
        
        # Write individual course sheets at the END
        for course_name, df_course in sorted(course_data.items()):
            # Limit sheet name length (Excel has 31 char limit) and remove invalid characters
            sheet_name = course_name[:31]
            # Remove invalid Excel sheet name characters: : \ / ? * [ ]
            sheet_name = sheet_name.replace(':', '-').replace('\\', '-').replace('/', '-')
            sheet_name = sheet_name.replace('?', '').replace('*', '').replace('[', '').replace(']', '')
            df_course.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Now format the workbook
    wb = openpyxl.load_workbook(output_file)
    
    # Reorder sheets: Move Employee-Course Matrix to the front
    if 'Employee-Course Matrix' in wb.sheetnames:
        matrix_sheet = wb['Employee-Course Matrix']
        wb.move_sheet(matrix_sheet, offset=-wb.index(matrix_sheet))
    
    # Find position of first individual course sheet (starts with LAR)
    course_sheet_idx = None
    for idx, sheet_name in enumerate(wb.sheetnames):
        if sheet_name.startswith('LAR'):
            course_sheet_idx = idx
            break
    
    # Move DCS Training to just before the individual course sheets
    if 'DCS Training' in wb.sheetnames and course_sheet_idx is not None:
        dcs_sheet = wb['DCS Training']
        current_position = wb.index(dcs_sheet)
        wb.move_sheet(dcs_sheet, offset=course_sheet_idx - current_position)
    
    # Move Summary to just before DCS Training (or before course sheets if DCS doesn't exist)
    if 'Summary' in wb.sheetnames:
        # Recalculate positions after moving DCS Training
        target_idx = None
        if 'DCS Training' in wb.sheetnames:
            target_idx = wb.index(wb['DCS Training'])
        elif course_sheet_idx is not None:
            # If DCS Training doesn't exist, find course sheets again
            for idx, sheet_name in enumerate(wb.sheetnames):
                if sheet_name.startswith('LAR'):
                    target_idx = idx
                    break
        
        if target_idx is not None:
            summary_sheet = wb['Summary']
            current_position = wb.index(summary_sheet)
            wb.move_sheet(summary_sheet, offset=target_idx - current_position)
    
    # Format all sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format headers (first row)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Special handling for Employee-Course Matrix to make it fit on one page
        if sheet_name == 'Employee-Course Matrix':
            # Set narrower column widths for course columns
            for col_idx, column in enumerate(ws.columns, start=1):
                column_letter = get_column_letter(col_idx)
                
                # Employee column - wider
                if col_idx == 1:
                    ws.column_dimensions[column_letter].width = 20
                # Job Duty column - medium
                elif col_idx == 2:
                    ws.column_dimensions[column_letter].width = 15
                # DCS Training column - wider to accommodate multiple courses
                elif col_idx == 3:
                    ws.column_dimensions[column_letter].width = 25
                # Course columns - narrow to fit more
                else:
                    ws.column_dimensions[column_letter].width = 12
            
            # Enable text wrapping for all cells in DCS Training column
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Increase header row height for wrapped text
            ws.row_dimensions[1].height = 45
        
        # Special handling for DCS Training sheet
        elif sheet_name == 'DCS Training':
            # Set column widths for manual input
            ws.column_dimensions['A'].width = 20  # Employee
            ws.column_dimensions['B'].width = 15  # Job Duty
            ws.column_dimensions['C'].width = 40  # Course Name
            ws.column_dimensions['D'].width = 12  # Training Date
            ws.column_dimensions['E'].width = 12  # Status
            ws.column_dimensions['F'].width = 20  # Instructor
            ws.column_dimensions['G'].width = 30  # Notes
            
            # Increase header row height
            ws.row_dimensions[1].height = 30
        
        else:
            # Auto-fit columns for other sheets
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze top row
        ws.freeze_panes = 'A2'
    
    # Special formatting for Summary sheet
    if 'Summary' in wb.sheetnames:
        ws_summary = wb['Summary']
        ws_summary['A1'].font = Font(size=14, bold=True)
        ws_summary['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws_summary['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    
    wb.save(output_file)
    
    print(f"\n📊 Training tracker created: {output_file}")
    print("\n✨ The tracker includes:")
    print("   • Summary - Overview of all training data")
    print("   • Master Data - All enrollment records combined")
    print(f"   • Individual course sheets ({len(course_data)} courses)")
    print("   • Employee-Course Matrix - Quick view of who took what")
    print("   • Employee Training Detail - Complete training record per person")
    print("   • Completion Timeline - Chronological completion data")
    
    print("\n📈 Quick Statistics:")
    print(f"   Total Enrollments: {len(df_master)}")
    if 'Enrollment Status' in df_master.columns:
        # Enrolled = Completed (manual attendance tracking)
        completed = ((df_master['Enrollment Status'] == 'Enrolled') |
                    (df_master['Enrollment Status'] == 'Completed') |
                    (df_master['Completion Status'] == 'Complete')).sum()
        dropped = (df_master['Enrollment Status'] == 'Dropped').sum()
        print(f"   Completed: {completed} ({completed/len(df_master)*100:.1f}%)")
        print(f"   Dropped: {dropped} ({dropped/len(df_master)*100:.1f}%)")
    
    if 'Employee_Name' in df_master.columns:
        unique_employees = df_master['Employee_Name'].nunique()
        print(f"   Unique Employees: {unique_employees}")
else:
    print("\n⚠️  No data collected to create tracker")

print("\n" + "=" * 80)
