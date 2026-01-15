import pandas as pd
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Training files to process
training_files = [
    "Training/BC-LAR-ENGPRO001.xlsx",
    "Training/BC-LAR-ENGPRO002.xlsx",
    "Training/BC-LAR-ENGPRO003.xlsx",
    "Training/BC-LAR-ENGPRO006.xlsx",
    "Training/BC-LAR-ENGPRO007.xlsx"
]

# Dictionary to store all training data
all_attendees = []

print("Processing training files...")
print("=" * 80)

for file in training_files:
    file_path = Path(file)
    if not file_path.exists():
        print(f"⚠️  File not found: {file}")
        continue
    
    print(f"\n📁 Processing: {file_path.name}")
    
    try:
        # Read the Excel file
        xl = pd.ExcelFile(file_path)
        
        # Display sheet names
        print(f"   Sheets found: {', '.join(xl.sheet_names)}")
        
        # Try to find a sheet with attendee data
        for sheet_name in xl.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            if df.empty:
                continue
            
            print(f"\n   Sheet: '{sheet_name}' ({len(df)} rows)")
            print(f"   Columns: {list(df.columns)[:5]}...")  # Show first 5 columns
            
            # Extract basic info (this will need to be customized based on actual structure)
            # For now, let's extract any columns that might contain names, dates, or attendance
            for idx, row in df.iterrows():
                if idx >= 5:  # Only show first 5 rows for preview
                    break
                row_dict = row.to_dict()
                # Store all data for later processing
                row_dict['source_file'] = file_path.stem
                row_dict['sheet_name'] = sheet_name
                all_attendees.append(row_dict)
    
    except Exception as e:
        print(f"   ❌ Error reading file: {e}")

print("\n" + "=" * 80)
print(f"\n✅ Total records collected: {len(all_attendees)}")

# Create a consolidated dataframe
if all_attendees:
    df_all = pd.DataFrame(all_attendees)
    
    # Create the tracking spreadsheet
    output_file = "Training/Training_Attendance_Tracker.xlsx"
    
    # Create a new workbook with formatting
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary['A1'] = "Training Attendance Tracker"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A4'] = f"Total Training Sessions: {len(training_files)}"
    ws_summary['A5'] = f"Total Records: {len(all_attendees)}"
    
    ws_summary['A7'] = "Training Files Processed:"
    ws_summary['A7'].font = Font(bold=True)
    
    for idx, file in enumerate(training_files, start=8):
        ws_summary[f'A{idx}'] = f"• {Path(file).name}"
    
    # Create Master Attendance sheet
    ws_master = wb.create_sheet("Master Attendance", 1)
    
    # Define headers
    headers = [
        "Employee Name",
        "Training Course",
        "Training Date",
        "Status",
        "Instructor",
        "Location",
        "Duration (Hours)",
        "Certificate Issued",
        "Expiry Date",
        "Notes",
        "Source File"
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_master.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add some sample/template rows
    sample_data = [
        ["[Enter Name]", "BC-LAR-ENGPRO001", datetime.now().strftime('%Y-%m-%d'), "Completed", "", "", "", "Yes", "", "", "BC-LAR-ENGPRO001"],
        ["[Enter Name]", "BC-LAR-ENGPRO002", datetime.now().strftime('%Y-%m-%d'), "Completed", "", "", "", "Yes", "", "", "BC-LAR-ENGPRO002"],
        ["[Enter Name]", "BC-LAR-ENGPRO003", datetime.now().strftime('%Y-%m-%d'), "Completed", "", "", "", "Yes", "", "", "BC-LAR-ENGPRO003"],
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_master.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Auto-fit columns
    for column in ws_master.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws_master.column_dimensions[column_letter].width = adjusted_width
    
    # Create individual sheets for each training course
    for file in training_files:
        course_name = Path(file).stem
        ws_course = wb.create_sheet(course_name)
        
        # Headers
        course_headers = [
            "Employee Name",
            "Date Attended",
            "Status",
            "Score/Grade",
            "Certificate Number",
            "Notes"
        ]
        
        for col_idx, header in enumerate(course_headers, start=1):
            cell = ws_course.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add template rows
        for row_idx in range(2, 12):
            ws_course.cell(row=row_idx, column=1).value = "[Enter Name]"
            ws_course.cell(row=row_idx, column=2).value = datetime.now().strftime('%Y-%m-%d')
            ws_course.cell(row=row_idx, column=3).value = "Pending"
        
        # Auto-fit columns
        for column in ws_course.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 40)
            ws_course.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_file)
    print(f"\n📊 Training tracker created: {output_file}")
    print("\n✨ The tracker includes:")
    print("   • Summary sheet with overview")
    print("   • Master Attendance sheet for consolidated tracking")
    print(f"   • Individual sheets for each training course ({len(training_files)} courses)")
    print("\n💡 Next steps:")
    print("   1. Open the tracker and review the template")
    print("   2. Replace sample data with actual attendee information")
    print("   3. Update status, dates, and other fields as needed")
else:
    print("\n⚠️  No data collected to create tracker")

print("\n" + "=" * 80)
