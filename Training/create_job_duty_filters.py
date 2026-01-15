import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

print("Creating filtered views by Job Duty...")
print("=" * 80)

# Read the existing tracker with Job Duty filled in
tracker_file = "Training/Training_Attendance_Tracker.xlsx"

try:
    # Read the Employee-Course Matrix sheet
    df_matrix = pd.read_excel(tracker_file, sheet_name='Employee-Course Matrix')
    
    print(f"Loaded Employee-Course Matrix with {len(df_matrix)} employees")
    
    # Check if Job Duty column exists and has data
    if 'Job Duty' not in df_matrix.columns:
        print("❌ Error: Job Duty column not found in Employee-Course Matrix")
        exit(1)
    
    # Get unique job duties (excluding empty values)
    job_duties = df_matrix['Job Duty'].dropna().unique()
    job_duties = [jd for jd in job_duties if str(jd).strip() != '']
    
    if len(job_duties) == 0:
        print("❌ Error: No job duties found. Please fill in the Job Duty column first.")
        exit(1)
    
    print(f"Found {len(job_duties)} job duties:")
    for jd in sorted(job_duties):
        count = len(df_matrix[df_matrix['Job Duty'] == jd])
        print(f"  • {jd}: {count} employees")
    
    # Load the existing workbook to add new sheets
    wb = openpyxl.load_workbook(tracker_file)
    
    # Remove old job duty sheets if they exist (to avoid duplicates)
    sheets_to_remove = [s for s in wb.sheetnames if s.startswith('JD-')]
    for sheet_name in sheets_to_remove:
        wb.remove(wb[sheet_name])
        print(f"  Removed old sheet: {sheet_name}")
    
    # Create filtered sheets for each job duty
    print("\nCreating filtered sheets...")
    
    # Create new sheets directly with openpyxl
    for job_duty in sorted(job_duties):
        # Filter data for this job duty
        df_filtered = df_matrix[df_matrix['Job Duty'] == job_duty].copy()
        
        # Create sheet name (Excel limit is 31 chars, no brackets allowed)
        sheet_name = f"JD-{job_duty}"[:31]
        
        # Create new sheet
        ws = wb.create_sheet(sheet_name)
        
        # Write headers
        for col_idx, col_name in enumerate(df_filtered.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = col_name
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Write data
        for row_idx, row_data in enumerate(df_filtered.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        print(f"  ✓ Created: {sheet_name} ({len(df_filtered)} employees)")
    
    # Format and position the new sheets
    for sheet_name in sorted([s for s in wb.sheetnames if s.startswith('JD-')]):
        ws = wb[sheet_name]
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Auto-fit columns
        for col_idx, column in enumerate(ws.columns, start=1):
            column_letter = get_column_letter(col_idx)
            
            # Employee column - wider
            if col_idx == 1:
                ws.column_dimensions[column_letter].width = 20
            # Job Duty column - medium
            elif col_idx == 2:
                ws.column_dimensions[column_letter].width = 15
            # Course columns - narrow to fit more
            else:
                ws.column_dimensions[column_letter].width = 12
        
        # Increase header row height for wrapped text
        ws.row_dimensions[1].height = 45
        
        # Freeze top row
        ws.freeze_panes = 'A2'
        
        # Move sheet to after Employee-Course Matrix
        current_idx = wb.sheetnames.index(sheet_name)
        target_idx = wb.sheetnames.index('Employee-Course Matrix') + 1
        # Count how many JD sheets should come before this one
        jd_sheets_before = len([s for s in sorted([x for x in wb.sheetnames if x.startswith('JD-')]) if s < sheet_name])
        final_target = target_idx + jd_sheets_before
        if current_idx > final_target:
            wb.move_sheet(ws, offset=final_target - current_idx)
    
    wb.save(tracker_file)
    
    print(f"\n✅ Successfully created {len(job_duties)} job duty filtered sheets!")
    print(f"\n📊 Updated tracker: {tracker_file}")
    print("\n✨ New sheets added (after Employee-Course Matrix):")
    for jd in sorted(job_duties):
        print(f"   • JD-{jd}")
    
except FileNotFoundError:
    print(f"❌ Error: Could not find {tracker_file}")
    print("   Please make sure the training tracker exists.")
except Exception as e:
    print(f"❌ Error: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 80)
