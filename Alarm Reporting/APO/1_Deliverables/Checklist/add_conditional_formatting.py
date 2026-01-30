from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font
from copy import copy

# Load Barbara's template to see conditional formatting
template_file = 'C:/Users/GF99/Documentation/Alarm Reporting/APO/2_Source_Documents/Working_Versions/APO_Deployment_Workflow_Checklist.xlsx'
wb_template = load_workbook(template_file)
ws_template = wb_template['Follow Up']

print("Conditional formatting rules in Barbara's template:")
print(f"Sheet: Follow Up")
print(f"Number of rules: {len(ws_template.conditional_formatting._cf_rules)}")

for rule_range, rules_list in ws_template.conditional_formatting._cf_rules.items():
    print(f"\nRange: {rule_range}")
    for rule in rules_list:
        print(f"  Type: {type(rule).__name__}")
        print(f"  Rule: {rule}")
        if hasattr(rule, 'formula'):
            print(f"  Formula: {rule.formula}")
        if hasattr(rule, 'dxf'):
            print(f"  Format: {rule.dxf}")
            if rule.dxf and rule.dxf.fill:
                print(f"    Fill: {rule.dxf.fill}")
            if rule.dxf and rule.dxf.font:
                print(f"    Font: {rule.dxf.font}")

wb_template.close()

# Now let's add conditional formatting to our converted file
wb_converted = load_workbook('ACM_to_APO_Migration_Comprehensive_Checklist_Barbara_Format.xlsx')

# Define color fills matching Barbara's exact template colors
custom_fill = PatternFill(start_color='FFFF7C80', end_color='FFFF7C80', fill_type='solid')  # Light red/pink
vendor_fill = PatternFill(start_color='FF92D050', end_color='FF92D050', fill_type='solid')  # Light green
itot_fill = PatternFill(start_color='FF00B0F0', end_color='FF00B0F0', fill_type='solid')    # Light blue
site_fill = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')    # Orange/amber

# Apply conditional formatting to all converted sheets (except Executive Summary)
for sheet_name in wb_converted.sheetnames:
    if sheet_name == 'Executive Summary':
        continue
    
    ws = wb_converted[sheet_name]
    
    # Get the maximum row with data
    max_row = ws.max_row
    
    # Apply conditional formatting to the entire data range
    # Format based on Owner column (column D)
    range_str = f'D2:D{max_row}'
    
    # Rule 1: If Owner = "Custom" - Light red/pink
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator='equal', formula=['"Custom"'], fill=custom_fill))
    
    # Rule 2: If Owner = "Vendor" - Light green
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator='equal', formula=['"Vendor"'], fill=vendor_fill))
    
    # Rule 3: If Owner = "IT/OT" or "IT\\OT" - Light blue
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator='equal', formula=['"IT\\\\OT"'], fill=itot_fill))
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator='equal', formula=['"IT/OT"'], fill=itot_fill))
    
    # Rule 4: If Owner = "Site" - Orange/amber
    ws.conditional_formatting.add(range_str,
        CellIsRule(operator='equal', formula=['"Site"'], fill=site_fill))
    
    print(f"✓ Applied conditional formatting to: {sheet_name}")

wb_converted.save('ACM_to_APO_Migration_Comprehensive_Checklist_Barbara_Format.xlsx')
print(f"\n✓ Conditional formatting applied successfully!")

wb_converted.close()
