from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

# Load source (comprehensive) checklist
source_file = 'ACM_to_APO_Migration_Comprehensive_Checklist.xlsx'
wb_source = load_workbook(source_file)

# Load Barbara's template to get exact styling
template_file = 'C:/Users/GF99/Documentation/Alarm Reporting/APO/2_Source_Documents/Working_Versions/APO_Deployment_Workflow_Checklist.xlsx'
wb_template = load_workbook(template_file)
template_sheet = wb_template['Follow Up']

# Get styling from template
header_fill = PatternFill(start_color='FF1F4E78', end_color='FF1F4E78', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')
category_fill = PatternFill(start_color='FFE7E6E6', end_color='FFE7E6E6', fill_type='solid')
category_font = Font(bold=True)
normal_font = Font(bold=False)

# Sheets to convert (skip Executive Summary)
sheets_to_convert = [
    ('Pre-Assessment', '0-Pre-Assessment'),
    ('Cleanup', '0-Cleanup'),
    ('Custom-Solutions', '1-Custom Solutions'),
    ('Pre-Planning', '2-Pre-Planning'),
    ('OSW-Completion', '3-OSW Completion'),
    ('Deployment-Readiness', '4-Deployment Readiness'),
    ('Migration-Execute', '5-Migration Execute'),
    ('Validation', '6-Validation & Go-Live'),
    ('Cutover', '7-Cutover'),
    ('Post-Migration', '8-Post-Migration')
]

# Create new workbook
from openpyxl import Workbook
wb_new = Workbook()
wb_new.remove(wb_new.active)

# Copy Executive Summary as-is
if 'Executive Summary' in wb_source.sheetnames:
    ws_source = wb_source['Executive Summary']
    ws_new = wb_new.create_sheet('Executive Summary')
    
    for row in ws_source.iter_rows():
        for cell in row:
            new_cell = ws_new[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.alignment = copy(cell.alignment)
    
    for col_letter, col_dim in ws_source.column_dimensions.items():
        ws_new.column_dimensions[col_letter].width = col_dim.width

# Convert each sheet
for source_name, target_name in sheets_to_convert:
    if source_name not in wb_source.sheetnames:
        continue
    
    ws_source = wb_source[source_name]
    ws_new = wb_new.create_sheet(target_name)
    
    # Set column widths
    ws_new.column_dimensions['A'].width = 5
    ws_new.column_dimensions['B'].width = 10
    ws_new.column_dimensions['C'].width = 60
    ws_new.column_dimensions['D'].width = 12
    ws_new.column_dimensions['E'].width = 15
    ws_new.column_dimensions['F'].width = 30
    ws_new.column_dimensions['G'].width = 15
    
    # Add headers
    headers = ['#', 'Item #', 'Task', 'Owner', 'Status', 'Notes', 'Prerequisites']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_new.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Process source data
    current_row = 2
    category_num = 1
    
    for row_idx, row in enumerate(ws_source.iter_rows(min_row=2, values_only=True), 2):
        task_col = row[1] if len(row) > 1 else None
        
        if not task_col:
            continue
        
        # Check if this is a category row (ends with " Tasks" or numbered section)
        is_category = False
        if task_col:
            task_str = str(task_col)
            # Category indicators: ends with "Tasks", or contains "**" markdown, or starts with number followed by period
            if (task_str.endswith('Tasks') or 
                task_str.startswith('**') or 
                (len(task_str) > 2 and task_str[0].isdigit() and task_str[1] == '.')):
                is_category = True
        
        if is_category:
            # Category row: number in A, text in B
            ws_new.cell(current_row, 1, category_num).font = category_font
            ws_new.cell(current_row, 1, category_num).fill = category_fill
            
            # Clean up task text
            clean_task = task_str.replace('**', '').strip()
            if clean_task and clean_task[0].isdigit() and clean_task[1] == '.':
                clean_task = clean_task[3:].strip()  # Remove "1. " prefix
            
            ws_new.cell(current_row, 2, clean_task).font = category_font
            ws_new.cell(current_row, 2, clean_task).fill = category_fill
            
            category_num += 1
            current_row += 1
        else:
            # Regular task row
            prev_row = current_row - 1
            
            # Column A: Auto-numbering formula (checks if PREVIOUS row B is text/category)
            ws_new.cell(current_row, 1, f'=IF(ISTEXT(B{prev_row}), A{prev_row}+1, A{prev_row})')
            
            # Column B: Item number (resets to 1 after category, otherwise increments)
            # Formula checks if PREVIOUS row B was text (category), if so start at 1, else increment
            ws_new.cell(current_row, 2, f'=IF(ISTEXT(B{prev_row}), 1, B{prev_row}+1)')
            
            # Column C: Task
            ws_new.cell(current_row, 3, task_col)
            
            # Column D: Owner
            if len(row) > 2:
                ws_new.cell(current_row, 4, row[2])
            
            # Column E: Status
            if len(row) > 3:
                ws_new.cell(current_row, 5, row[3])
            
            # Column F: Notes
            if len(row) > 4:
                ws_new.cell(current_row, 6, row[4])
            
            # Column G: Prerequisites
            if len(row) > 5:
                ws_new.cell(current_row, 7, row[5])
            
            current_row += 1

# Save output
output_file = 'ACM_to_APO_Migration_Comprehensive_Checklist_Barbara_Format.xlsx'
wb_new.save(output_file)
print(f"✓ Converted checklist saved to: {output_file}")
print(f"✓ Converted {len(sheets_to_convert)} sheets to Barbara's format")

wb_source.close()
wb_template.close()
wb_new.close()
