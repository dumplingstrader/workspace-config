import openpyxl

wb = openpyxl.load_workbook('C:/Users/GF99/Documentation/Alarm Reporting/APO/2_Source_Documents/Working_Versions/APO_Deployment_Workflow_Checklist.xlsx')

print(f"Sheets in Barbara's format: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    
    print(f"Dimensions: {sheet.dimensions}")
    
    # Read all data
    data = []
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            data.append(row)
    
    if data:
        print(f"\nHeaders: {data[0]}")
        print(f"Total rows: {len(data)}")
        
        # Print first 15 rows
        print(f"\nFirst {min(15, len(data))} rows:")
        for i, row in enumerate(data[:15], 1):
            print(f"Row {i}: {row}")
        
        if len(data) > 15:
            print(f"\n... ({len(data) - 15} more rows)")
    
    # Check for any formatting, colors, or special features
    print("\n--- Cell Formatting Sample (A1:F3) ---")
    for row_idx in range(1, min(4, sheet.max_row + 1)):
        for col_idx in range(1, min(7, sheet.max_column + 1)):
            cell = sheet.cell(row_idx, col_idx)
            if cell.value:
                print(f"  {cell.coordinate}: value='{cell.value}', font={cell.font.bold if cell.font else 'None'}, fill={cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else 'None'}")

wb.close()
