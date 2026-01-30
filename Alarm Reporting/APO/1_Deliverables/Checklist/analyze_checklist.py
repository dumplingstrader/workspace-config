import openpyxl
import json

wb = openpyxl.load_workbook('ACM_to_APO_Migration_Comprehensive_Checklist.xlsx')

print(f"Sheets in workbook: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")
    
    # Get dimensions
    print(f"Dimensions: {sheet.dimensions}")
    
    # Read all data
    data = []
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            data.append(row)
    
    # Print headers if available
    if data:
        print(f"\nHeaders: {data[0]}")
        print(f"Total rows: {len(data)}")
        
        # Print first 10 rows
        print(f"\nFirst {min(10, len(data))} rows:")
        for i, row in enumerate(data[:10], 1):
            print(f"Row {i}: {row}")
        
        if len(data) > 10:
            print(f"\n... ({len(data) - 10} more rows)")

wb.close()
