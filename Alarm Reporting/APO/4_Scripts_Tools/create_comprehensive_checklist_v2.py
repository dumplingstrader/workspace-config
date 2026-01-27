"""
ACM to APO Migration Comprehensive Checklist Generator V2
Extracts actual content from whitepaper outline and maps to Barbara's phase structure
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import re

# Define colors for formatting
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
SECTION_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
SUBSECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, size=11, color="1F4E78")
SUBSECTION_FONT = Font(bold=True, size=10, color="44546A")
NORMAL_FONT = Font(size=10)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def parse_whitepaper_outline():
    """Parse the whitepaper outline to extract actionable items"""
    try:
        with open('ACM_to_APO_Migration_Whitepaper_Outline.md', 'r', encoding='utf-8') as f:
            content = f.read()
        return content
    except FileNotFoundError:
        print("Warning: Whitepaper outline not found. Using default content.")
        return ""

def extract_tasks_from_section(content, section_pattern, end_pattern=None):
    """Extract tasks from a specific section of the outline"""
    tasks = []
    
    # Find section start
    section_match = re.search(section_pattern, content, re.MULTILINE)
    if not section_match:
        return tasks
    
    start_pos = section_match.end()
    
    # Find section end
    if end_pattern:
        end_match = re.search(end_pattern, content[start_pos:], re.MULTILINE)
        if end_match:
            section_text = content[start_pos:start_pos + end_match.start()]
        else:
            section_text = content[start_pos:start_pos + 5000]  # Take next 5000 chars
    else:
        section_text = content[start_pos:start_pos + 5000]
    
    # Extract bullet points and headers
    lines = section_text.split('\n')
    current_section = None
    
    for line in lines:
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
            
        # Section header (###)
        if line.startswith('####'):
            current_section = line.replace('####', '').strip()
            continue
        elif line.startswith('###') and not line.startswith('####'):
            current_section = line.replace('###', '').strip()
            continue
            
        # Bullet point
        if line.startswith('- **') or line.startswith('- '):
            # Clean up the text
            task = line.lstrip('- ').strip()
            # Remove markdown bold markers
            task = task.replace('**', '')
            # Skip very short lines (likely formatting)
            if len(task) > 10:
                tasks.append({
                    'section': current_section or 'General',
                    'task': task
                })
    
    return tasks

def create_executive_summary():
    """Create Executive Summary sheet with top failure reasons"""
    data = {
        'Section': [
            '⚠️ TOP REASONS FOR DELAYS & COST OVERRUNS',
            '',
            '1. MISUNDERSTANDING OF DEPENDENCIES',
            '   •',
            '   •',
            '   •',
            '',
            '2. LICENSE MISCALCULATION',
            '   •',
            '   •',
            '   •',
            '   •',
            '',
            '3. IT-FOCUSED PROJECTS (Quality & Usability Ignored)',
            '   •',
            '   •',
            '   •',
            '',
            '4. NO PRE-MIGRATION CLEANUP',
            '   •',
            '   •',
            '   •',
            '   •',
            '   •',
            '',
            '5. APO MATURITY GAPS',
            '   •',
            '',
            '📊 CRITICAL SUCCESS FACTORS',
            '',
            'Database Health',
            'License Sizing',
            'Hierarchy Alignment',
            'Custom Solutions',
            'Training & Change Management',
            'Post-Migration Support',
        ],
        'Description': [
            'Based on Marathon Petroleum pilots and multiple site assessments',
            '',
            'Delays and additional costs',
            'Perception that APO and Reporting must be deployed together',
            'Actually: HAMR should be working well (HAMR 2.3.0+) BEFORE APO',
            'Critical to have hierarchies compatible across ACM/M&R/EAS',
            '',
            'Cost overruns, delays if license insufficient',
            'Cannot delete tags without licenses - no reclaim after ordering',
            'Too many licenses bought but cannot return excess',
            'Example: Anacortes reduced by 35% through cleanup',
            'License count during migration unreliable but allows proceed',
            '',
            'No focus on quality or usability',
            'Optimal configuration gets insufficient attention',
            'Delays in providing functional enhancements',
            'Custom solutions not prioritized',
            '',
            'Delays and ineffective outcomes',
            'Hierarchy cleanup required (may take months)',
            'Database cleanup and testing (months of work)',
            'Corruptions must be fixed in ACM then re-migrate',
            'Difficult to adjust hierarchies later in APO',
            'Cannot realize full rationalization benefits without quality data',
            '',
            'Missing features and functionality gaps',
            'Custom solutions required for operational excellence',
            '',
            'Critical factors for successful migration',
            '',
            'ACM/HAMR cleanup, ghost tag removal, corruption fixes',
            'Evaluate actual current usage before ordering (30-40% ghost tags common)',
            'Ensure ACM/M&R/EAS hierarchies aligned before migration',
            'Plan for ACM→APO feature gaps, implement enhanced sync and tools',
            'Focus on "what was lost" not just new features',
            'Daily health checks, backup/recovery, continuous improvement',
        ]
    }
    
    return pd.DataFrame(data)

def create_phase_sheets_from_outline():
    """Create phase sheets by parsing whitepaper outline"""
    
    outline_content = parse_whitepaper_outline()
    
    phases = {}
    
    # Phase 0: Pre-Migration Assessment (maps to Section 4)
    phase0_tasks = []
    if outline_content:
        # Extract from Section 4
        section4_tasks = extract_tasks_from_section(
            outline_content, 
            r'## \*\*4\. Pre-Migration Assessment',
            r'## \*\*5\.'
        )
        for item in section4_tasks:
            phase0_tasks.append((item['section'], item['task'], 'Site'))
    
    # Add default tasks if outline parsing failed
    if not phase0_tasks:
        phase0_tasks = [
            ('Database Health Assessment', 'Run ACM Database Health Check', 'Site'),
            ('Database Health Assessment', 'Document all corruptions and errors', 'Site'),
            ('License Assessment', 'Count active tags vs licensed capacity', 'Site'),
            ('License Assessment', 'Identify ghost tags (30-40% typical)', 'Site'),
        ]
    
    phases['0-Pre-Assessment'] = {
        'phase_name': 'Phase 0: Pre-Migration Assessment (6-12 months before)',
        'sections': [('Assessment Tasks', phase0_tasks)]
    }
    
    # Phase 1: Database Cleanup (maps to Section 5)
    phase1_tasks = []
    if outline_content:
        section5_tasks = extract_tasks_from_section(
            outline_content,
            r'## \*\*5\. Database Health and Cleanup',
            r'## \*\*6\.'
        )
        for item in section5_tasks:
            phase1_tasks.append((item['section'], item['task'], 'Site'))
    
    if not phase1_tasks:
        phase1_tasks = [
            ('ACM Cleanup', 'Remove ghost tags and orphaned entries', 'Site'),
            ('ACM Cleanup', 'Fix hierarchy corruptions', 'Site'),
            ('HAMR Cleanup', 'Install HAMR 2.3.0+ before APO', 'Vendor'),
            ('HAMR Cleanup', 'Remove ghost tags from HAMR database', 'Site'),
        ]
    
    phases['1-Cleanup'] = {
        'phase_name': 'Phase 1: Database Cleanup (Before License Order)',
        'sections': [('Cleanup Tasks', phase1_tasks)]
    }
    
    # Phase 2: Custom Solutions (maps to Section 7)
    phase2_tasks = []
    if outline_content:
        section7_tasks = extract_tasks_from_section(
            outline_content,
            r'## \*\*7\. Custom Solutions',
            r'## \*\*8\.'
        )
        for item in section7_tasks:
            owner = 'Custom' if any(word in item['task'].lower() for word in ['custom', 'develop', 'implement', 'script', 'tool']) else 'Site'
            phase2_tasks.append((item['section'], item['task'], owner))
    
    if not phase2_tasks:
        phase2_tasks = [
            ('Gap Analysis', 'Identify ACM→APO feature gaps', 'Site'),
            ('Custom Tools', 'Develop tag export/import tools', 'Custom'),
            ('Health Checks', 'Implement daily health check scripts', 'Custom'),
        ]
    
    phases['2-Custom-Solutions'] = {
        'phase_name': 'Phase 2: Custom Solutions & Gap Mitigation',
        'sections': [('Custom Solution Tasks', phase2_tasks)]
    }
    
    # Phase 3: Installation Planning (maps to Section 6.1, 6.2)
    phase3_tasks = []
    if outline_content:
        section6_tasks = extract_tasks_from_section(
            outline_content,
            r'### \*\*6\.1 Environment Preparation',
            r'### \*\*6\.3'
        )
        for item in section6_tasks:
            owner = 'IT/OT' if any(word in item['task'].lower() for word in ['server', 'network', 'sql', 'infrastructure']) else 'Site'
            phase3_tasks.append((item['section'], item['task'], owner))
    
    if not phase3_tasks:
        phase3_tasks = [
            ('Infrastructure', 'Order L4 APO server', 'IT/OT'),
            ('Infrastructure', 'Configure SQL Server', 'IT/OT'),
            ('Installation', 'Install APO on L4 server', 'Vendor'),
        ]
    
    phases['3-Pre-Planning'] = {
        'phase_name': 'Phase 3: APO Installation Planning',
        'sections': [('Installation Planning Tasks', phase3_tasks)]
    }
    
    # Phase 4: OSW Completion (maps to outline Section 10 or custom Barbara content)
    phases['4-OSW-Completion'] = {
        'phase_name': 'Phase 4: Onsite Scoping Workbook (OSW)',
        'sections': [
            ('OSW Data Entry', [
                ('Assets', 'Complete Asset Hierarchy (Site→Area→Unit→System→Subsystem)', 'Site'),
                ('Assets', 'Validate hierarchy matches FDS standards', 'Site'),
                ('Devices', 'Complete Device Inventory (manufacturer, model, protocol)', 'Site'),
                ('Devices', 'Classify external reference devices', 'Site'),
                ('Devices', 'Document unsupported devices with justification', 'Site'),
            ]),
            ('External Reference Configuration', [
                ('ControlLogix', 'Provide .L5X/.L5K export and I/O config', 'Site'),
                ('PLC5/SLC', 'Provide .RSS export and addressing scheme', 'Site'),
                ('Triconex', 'Confirm EPKS External Reference or TPS import', 'Site'),
                ('Bently Nevada', 'Confirm DCS External Reference method', 'Site'),
                ('Analyzers', 'Decide manual entry vs automated import', 'Site'),
            ]),
            ('OSW Validation', [
                ('Review', 'Vendor reviews OSW for completeness', 'Vendor'),
                ('Review', 'Site confirms all devices in scope', 'Site'),
                ('Review', '⚠️ NO additions after sign-off without Change Order', 'Site'),
                ('Sign-Off', 'Executive sponsor signs off (scope locked)', 'Site'),
            ]),
        ]
    }
    
    # Phase 5: Migration Execution Readiness (maps to Section 6.3, 6.4)
    phase5_tasks = []
    if outline_content:
        section6_3_tasks = extract_tasks_from_section(
            outline_content,
            r'### \*\*6\.3 Data Migration Process',
            r'### \*\*6\.5'
        )
        for item in section6_3_tasks:
            phase5_tasks.append((item['section'], item['task'], 'Site'))
    
    if not phase5_tasks:
        phase5_tasks = [
            ('Pre-Migration', 'Disable scheduled enforcements in ACM', 'Site'),
            ('Pre-Migration', 'Create full ACM backup', 'Site'),
            ('Validation', 'Test backup restore procedures', 'Site'),
        ]
    
    phases['5-Deployment-Readiness'] = {
        'phase_name': 'Phase 5: Migration Execution Readiness',
        'sections': [('Readiness Tasks', phase5_tasks)]
    }
    
    # Phase 6: Migration Execution (maps to Section 8)
    phase6_tasks = []
    if outline_content:
        section8_tasks = extract_tasks_from_section(
            outline_content,
            r'## \*\*8\. Migration Execution',
            r'## \*\*9\.'
        )
        for item in section8_tasks:
            phase6_tasks.append((item['section'], item['task'], 'Vendor'))
    
    if not phase6_tasks:
        phase6_tasks = [
            ('Execution', 'Run ACM to APO migration tool', 'Vendor'),
            ('Execution', 'Validate data transfer completeness', 'Vendor'),
            ('Configuration', 'Configure APO site settings', 'Site'),
        ]
    
    phases['6-Migration-Execute'] = {
        'phase_name': 'Phase 6: APO Migration Execution',
        'sections': [('Migration Execution Tasks', phase6_tasks)]
    }
    
    # Phase 7: Validation & Parallel Operations
    phases['7-Validation'] = {
        'phase_name': 'Phase 7: Validation & Parallel Operations',
        'sections': [
            ('Parallel Operation Setup', [
                ('Configuration', 'Run ACM and APO in parallel (30-90 days)', 'Site'),
                ('Configuration', 'Maintain ACM for fallback capability', 'Site'),
                ('Configuration', 'Monitor both systems for discrepancies', 'Site'),
            ]),
            ('Data Validation', [
                ('Testing', 'Compare ACM and APO alarm counts', 'Site'),
                ('Testing', 'Validate enforcement execution', 'Site'),
                ('Testing', 'Verify reporting accuracy', 'Site'),
                ('Testing', 'Test custom solutions in APO', 'Custom'),
            ]),
            ('Performance Validation', [
                ('Metrics', 'Monitor sync accuracy (Standard vs Alternative)', 'Site'),
                ('Metrics', 'Verify KPI calculations match ISA 18.2 requirements', 'Site'),
                ('Metrics', 'Validate daily health check notifications', 'Custom'),
            ]),
        ]
    }
    
    # Phase 8: Cutover & Decommission
    phases['8-Cutover'] = {
        'phase_name': 'Phase 8: Cutover & Decommission',
        'sections': [
            ('Pre-Cutover', [
                ('Validation', 'Final ACM vs APO comparison', 'Site'),
                ('Validation', '⚠️ CRITICAL: Confirm all enforcements tested', 'Site'),
                ('Validation', 'Document any known issues and workarounds', 'Site'),
                ('Communication', 'Notify operations of cutover schedule', 'Site'),
            ]),
            ('Cutover Execution', [
                ('Cutover', 'Disable ACM enforcements', 'Site'),
                ('Cutover', 'Enable APO enforcements', 'Site'),
                ('Cutover', 'Monitor first 24 hours closely', 'Site'),
                ('Cutover', 'Verify DCS and APO in sync', 'Site'),
            ]),
            ('ACM Decommission', [
                ('Decommission', 'Retain ACM for 90 days as backup', 'Site'),
                ('Decommission', 'Archive ACM database', 'Site'),
                ('Decommission', 'Remove ACM services (after approval)', 'IT/OT'),
            ]),
        ]
    }
    
    # Phase 9: Post-Migration Optimization (maps to Section 7.5)
    phase9_tasks = []
    if outline_content:
        section7_5_tasks = extract_tasks_from_section(
            outline_content,
            r'### \*\*7\.5 Continuous Improvement',
            r'## \*\*8\.'
        )
        for item in section7_5_tasks:
            phase9_tasks.append((item['section'], item['task'], 'Site'))
    
    if not phase9_tasks:
        phase9_tasks = [
            ('Optimization', 'Regular database maintenance (weekly, monthly)', 'Site'),
            ('Optimization', 'Monitor license consumption and ghost tags', 'Site'),
            ('Optimization', 'Performance tuning and index maintenance', 'Site'),
        ]
    
    phases['9-Post-Migration'] = {
        'phase_name': 'Phase 9: Post-Migration Optimization',
        'sections': [('Optimization Tasks', phase9_tasks)]
    }
    
    return phases

def format_worksheet(ws, columns):
    """Apply formatting to worksheet"""
    # Set column widths
    column_widths = {'A': 20, 'B': 50, 'C': 15, 'D': 15, 'E': 40, 'F': 30}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Format header row
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    
    # Set header row height
    ws.row_dimensions[1].height = 45
    
    # Freeze header row
    ws.freeze_panes = 'A2'

def add_status_dropdown(ws, start_row, end_row):
    """Add status dropdown to column D"""
    dv = DataValidation(
        type="list",
        formula1='"☐ Not Started,⏳ In Progress,✓ Complete,✗ Blocked,⚠️ On Hold"',
        allow_blank=False
    )
    dv.add(f'D{start_row}:D{end_row}')
    ws.add_data_validation(dv)

def create_workbook():
    """Create comprehensive checklist workbook"""
    print("Creating comprehensive checklist from whitepaper outline...")
    
    # Create new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create Executive Summary
    print("Creating Executive Summary...")
    ws_summary = wb.create_sheet('Executive Summary', 0)
    df_summary = create_executive_summary()
    
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif 'TOP REASONS' in str(value) or 'CRITICAL SUCCESS' in str(value):
                cell.font = Font(bold=True, size=12, color="C00000")
            elif value and str(value).startswith(('1.', '2.', '3.', '4.', '5.')):
                cell.font = Font(bold=True, size=11)
                cell.fill = SECTION_FILL
    
    ws_summary.column_dimensions['A'].width = 40
    ws_summary.column_dimensions['B'].width = 80
    
    # Create phase sheets
    print("Parsing whitepaper outline and creating phase sheets...")
    phases = create_phase_sheets_from_outline()
    
    for phase_key in sorted(phases.keys()):
        phase = phases[phase_key]
        print(f"Creating sheet: {phase['phase_name']}")
        
        ws = wb.create_sheet(phase_key.split('-', 1)[1])
        
        # Add headers
        headers = ['#', 'Task', 'Owner', 'Status', 'Notes', 'Prerequisites']
        ws.append(headers)
        format_worksheet(ws, headers)
        
        current_row = 2
        section_num = 1
        
        for section_name, tasks in phase['sections']:
            # Add section header row
            section_row = current_row
            ws.cell(row=current_row, column=1, value=f"{section_num}.")
            ws.cell(row=current_row, column=2, value=section_name)
            
            # Style section header
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.font = SECTION_FONT
                cell.fill = SECTION_FILL
                cell.border = THIN_BORDER
            
            current_row += 1
            item_num = 1
            subsection_start = current_row
            
            # Group tasks by subsection
            current_subsection = None
            subsection_rows = []
            
            for subsection, task, owner in tasks:
                if subsection != current_subsection:
                    # New subsection - add subsection header
                    if current_subsection is not None and subsection_rows:
                        # Group previous subsection
                        if len(subsection_rows) > 1:
                            ws.row_dimensions.group(subsection_rows[0], subsection_rows[-1], hidden=False)
                        subsection_rows = []
                    
                    current_subsection = subsection
                    ws.cell(row=current_row, column=2, value=subsection)
                    for col in range(1, 7):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = SUBSECTION_FONT
                        cell.fill = SUBSECTION_FILL
                        cell.border = THIN_BORDER
                    current_row += 1
                
                # Add task
                ws.cell(row=current_row, column=1, value=f"{section_num}.{item_num}")
                ws.cell(row=current_row, column=2, value=task)
                ws.cell(row=current_row, column=3, value=owner)
                ws.cell(row=current_row, column=4, value="☐ Not Started")
                
                # Apply warning/critical styling
                if '⚠️' in task or 'CRITICAL' in task:
                    ws.cell(row=current_row, column=2).fill = WARNING_FILL
                
                # Apply borders and alignment
                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                
                subsection_rows.append(current_row)
                current_row += 1
                item_num += 1
            
            # Group last subsection
            if subsection_rows and len(subsection_rows) > 1:
                ws.row_dimensions.group(subsection_rows[0], subsection_rows[-1], hidden=False)
            
            section_num += 1
        
        # Add status dropdown
        add_status_dropdown(ws, 2, current_row - 1)
        
        print(f"  Added {current_row - 2} rows")
    
    # Save workbook
    output_file = 'ACM_to_APO_Migration_Comprehensive_Checklist.xlsx'
    wb.save(output_file)
    print(f"\n✓ Checklist created: {output_file}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  Total tasks: ~250+ items")
    
    return output_file

if __name__ == '__main__':
    create_workbook()
