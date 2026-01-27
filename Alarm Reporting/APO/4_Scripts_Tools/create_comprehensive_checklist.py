"""
ACM to APO Migration Comprehensive Checklist Generator
Combines Barbara's workflow structure with whitepaper outline content
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Define colors for formatting (Enhanced color scheme)
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue
SECTION_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")  # Light blue
SUBSECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Lighter blue
WARNING_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Yellow
CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
COMPLETE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
PROGRESS_FILL = PatternFill(start_color="FFF4CE", end_color="FFF4CE", fill_type="solid")  # Light yellow

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, size=11, color="1F4E78")
SUBSECTION_FONT = Font(bold=True, size=10, color="44546A")
NORMAL_FONT = Font(size=10)

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_executive_summary():
    """Create Executive Summary sheet"""
    data = {
        'Section': [
            '⚠️ TOP REASONS FOR DELAYS & COST OVERRUNS',
            '',
            '1. MISUNDERSTANDING OF DEPENDENCIES',
            '   •',
            '   •',
            '   •',
            '',
            '2. LICENSE MISCALCULATION',
            '   •',
            '   •',
            '   •',
            '   •',
            '',
            '3. IT-FOCUSED PROJECTS (Quality & Usability Ignored)',
            '   •',
            '   •',
            '   •',
            '',
            '4. NO PRE-MIGRATION CLEANUP',
            '   •',
            '   •',
            '   •',
            '   •',
            '   •',
            '',
            '5. APO MATURITY GAPS',
            '   •',
            '',
            '📊 CRITICAL SUCCESS FACTORS',
            '',
            'Database Health',
            'License Sizing',
            'Hierarchy Alignment',
            'Custom Solutions',
            'Training & Change Management',
            'Post-Migration Support',
        ],
        'Description': [
            'Based on Marathon Petroleum pilots and multiple site assessments',
            '',
            'Delays and additional costs',
            'Perception that APO and Reporting must be deployed together',
            'Actually: HAMR should be working well (HAMR 2.3.0+) BEFORE APO',
            'Critical to have hierarchies compatible across ACM/M&R/EAS',
            '',
            'Cost overruns, delays if license insufficient',
            'Cannot delete tags without licenses - no reclaim after ordering',
            'Too many licenses bought but cannot return excess',
            'Example: Anacortes reduced by 35% through cleanup',
            'License count during migration unreliable but allows proceed',
            '',
            'No focus on quality or usability',
            'Optimal configuration gets insufficient attention',
            'Delays in providing functional enhancements',
            'Custom solutions not prioritized',
            '',
            'Delays and ineffective outcomes',
            'Hierarchy cleanup required (may take months)',
            'Database cleanup and testing (months of work)',
            'Corruptions must be fixed in ACM then re-migrate',
            'Difficult to adjust hierarchies later in APO',
            'Cannot realize full rationalization benefits without quality data',
            '',
            'Management tools lacking',
            'Example: LAR Wilmington non-operational for months due to APO 1.1 gaps',
            '',
            'Green = Ready | Yellow = Caution | Red = Blocker',
            '',
            'ACM DB health, M&R optimization, EMDB integrity',
            'Ghost tags removed, accurate count before ordering',
            'ACM/M&R/EAS paths compatible, consoles mapped to OPs',
            'Backup scripts, notifications, gap mitigation',
            'Workflow changes documented, missing features mitigated',
            'Daily monitoring, issue response, custom tool migration',
        ]
    }
    return pd.DataFrame(data)

def create_phase_sheets():
    """Create phase-based checklist sheets"""
    
    phases = {
        '0-Pre-Assessment': {
            'phase_name': 'Phase 0: Pre-Migration Assessment (6-12 months)',
            'sections': [
                ('1. Database Health Assessment', [
                    ('ACM DB Health Check', 'Run health check scripts (ghost tags, invalid tags, corruptions)', 'Custom'),
                    ('ACM DB Health Check', 'Identify tags not in DCS (historical/test tags)', 'Site'),
                    ('ACM DB Health Check', 'Check tags in wrong assets (hierarchy misalignment)', 'Site'),
                    ('ACM DB Health Check', 'Validate alarm priorities follow rationalization rules', 'Site'),
                    ('ACM DB Health Check', 'Check corrupted parameters (limits, deadbands, delays)', 'Site'),
                    ('ACM DB Health Check', 'Review SCADA tag quality and corruptions', 'Site'),
                    ('ACM DB Health Check', 'Validate tag descriptions completeness and accuracy', 'Site'),
                    ('ACM DB Health Check', 'Document ACM console to HAMR OP mapping', 'Site'),
                    ('M&R DB Health Check', 'Run license consumption analysis (identify ghost tags)', 'Custom'),
                    ('M&R DB Health Check', 'Check for tag merging opportunities (Redirection Index)', 'Site'),
                    ('M&R DB Health Check', 'Identify old/corrupted assets for removal', 'Site'),
                    ('M&R DB Health Check', 'Review path arrangements and optimization needs', 'Site'),
                    ('M&R DB Health Check', 'Validate Operating Positions for APO console mapping', 'Site'),
                    ('M&R DB Health Check', 'Close invalid alarms in database', 'Custom'),
                    ('M&R DB Health Check', 'Remove invalid suppressed states', 'Custom'),
                    ('EAS EMDB Check', 'Run EMDB comparison tools (find discrepancies)', 'Custom'),
                    ('EAS EMDB Check', 'Check for inconsistent naming conventions', 'Site'),
                    ('EAS EMDB Check', 'Identify orphaned assets', 'Site'),
                    ('EAS EMDB Check', 'Validate hierarchy correctness', 'Site'),
                ]),
                ('2. Licensing Analysis', [
                    ('License Calculation', 'Calculate true tag count (exclude ghosts/invalids)', 'Custom'),
                    ('License Calculation', 'Validate license count before ordering', 'Site'),
                    ('License Calculation', 'Add 5-10% contingency buffer', 'Site'),
                    ('License Calculation', '⚠️ CRITICAL: Cannot reclaim licenses after ordering', 'Site'),
                    ('Cost Analysis', 'Project cleanup cost vs license cost', 'Site'),
                    ('Cost Analysis', 'Document 30-40% ghost tag potential savings', 'Site'),
                ]),
                ('3. Current System Dependencies', [
                    ('Alarm Help', 'Document Alarm Help Router connections', 'Site'),
                    ('Alarm Help', 'Identify custom Alarm Help integrations', 'Site'),
                    ('Enforcements', 'Catalog all static enforcement mappings', 'Site'),
                    ('Enforcements', 'Identify dynamic mode enforcements (highest risk)', 'Site'),
                    ('Enforcements', 'Document reverse enforcement scenarios', 'Site'),
                    ('Reporting', 'List M&R reporting schedules and dependencies', 'Site'),
                    ('Reporting', 'Identify custom reports and dashboards', 'Site'),
                ]),
            ]
        },
        
        '1-Cleanup-Pre-License': {
            'phase_name': 'Phase 1: Database Cleanup (Before License Order)',
            'sections': [
                ('1. ACM Database Cleanup', [
                    ('Temp Environment', 'Install temporary ACM Manager Server', 'Custom'),
                    ('Temp Environment', 'Install temporary Enforcer Server', 'Custom'),
                    ('Temp Environment', 'Verify Alarm Help and enforcement still work', 'Site'),
                    ('Tag Cleanup', 'Clean EAS EMDB and load to ACM', 'Site'),
                    ('Tag Cleanup', 'Move tags to correct assets', 'Site'),
                    ('Tag Cleanup', 'Export and delete tags with invalid structure', 'Site'),
                    ('Tag Cleanup', 'Delete unused Control Modules (CMs)', 'Site'),
                    ('Tag Cleanup', 'Clean SCADA tag corruptions', 'Site'),
                    ('Tag Cleanup', 'Fix corrupted notes', 'Site'),
                    ('Tag Cleanup', 'Fix invalid parameters', 'Site'),
                    ('Tag Cleanup', 'Address reconfiguration errors', 'Site'),
                    ('Tag Cleanup', 'Address enforcement errors', 'Site'),
                    ('Tag Cleanup', 'Validate all tag descriptions', 'Site'),
                    ('Enforcement Prep', 'Change all enforcements to monitor mode', 'Site'),
                    ('Enforcement Prep', 'Adjust schedules to avoid OPC overload', 'Site'),
                    ('Enforcement Prep', '⚠️ DISABLE enforcements in ACM DB before migration', 'Site'),
                    ('Final Validation', 'Run final health check', 'Custom'),
                    ('Final Validation', 'Obtain sign-off from stakeholders', 'Site'),
                ]),
                ('2. M&R Database Optimization', [
                    ('Hierarchy', 'Optimize M&R hierarchy structure', 'Site'),
                    ('Hierarchy', 'Remove ghost tags (no events in M&R)', 'Custom'),
                    ('Hierarchy', 'Consolidate duplicate assets', 'Site'),
                    ('Performance', 'Close invalid alarms', 'Custom'),
                    ('Performance', 'Remove invalid suppressed states', 'Custom'),
                    ('Performance', 'Adjust retention periods', 'Site'),
                    ('Performance', 'Run performance optimization scripts', 'Custom'),
                ]),
                ('3. Hierarchy Alignment Validation', [
                    ('Alignment Check', 'Validate ACM, M&R, EAS hierarchies compatible', 'Site'),
                    ('Alignment Check', 'Confirm ACM consoles map to HAMR OPs', 'Site'),
                    ('Alignment Check', 'Document any remaining misalignments', 'Site'),
                ]),
            ]
        },
        
        '2-Custom-Solutions': {
            'phase_name': 'Phase 2: Custom Solutions & Gap Mitigation',
            'sections': [
                ('1. Identify ACM→APO Feature Gaps', [
                    ('Gap Analysis', 'Tags import/export (only Excel tool in APO)', 'Site'),
                    ('Gap Analysis', 'EMDB import/export (only manual in APO)', 'Site'),
                    ('Gap Analysis', 'Tag movement between consoles (not available)', 'Site'),
                    ('Gap Analysis', 'Drag-and-drop hierarchy UI (not available)', 'Site'),
                    ('Gap Analysis', 'TagSync functionality (missing)', 'Site'),
                    ('Gap Analysis', 'BMA support (critical gap)', 'Site'),
                    ('Gap Analysis', 'Advanced query and bulk operations', 'Site'),
                    ('Gap Analysis', 'Document user impact of each gap', 'Site'),
                ]),
                ('2. Marathon Custom Solutions', [
                    ('Backups', 'Implement synchronized backup across all servers', 'Custom'),
                    ('Backups', 'Create backup scripts (files, DBs, exports)', 'Custom'),
                    ('Backups', 'Test restore procedures', 'Custom'),
                    ('Notifications', 'Set up "you know before users notice" alerts', 'Custom'),
                    ('Notifications', 'Configure health monitoring', 'Custom'),
                    ('Health Checks', 'Deploy daily health check scripts', 'Custom'),
                    ('Tag Management', 'Build tag export/import tools', 'Custom'),
                    ('Tag Management', 'Create EMDB manipulation utilities', 'Custom'),
                    ('Sync Tools', 'Develop Tag Sync replacement (if needed)', 'Custom'),
                ]),
                ('3. Recommended Settings', [
                    ('HAMR Configuration', 'Configure HAMR rules/processing', 'Site'),
                    ('ACM Configuration', 'Optimize ACM site settings', 'Site'),
                    ('Documentation', 'Document all custom solutions', 'Custom'),
                ]),
            ]
        },
        
        '3-Pre-Planning': {
            'phase_name': 'Phase 3: APO Installation Planning',
            'sections': [
                ('1. Infrastructure Readiness', [
                    ('Servers', 'Order L4 APO server', 'IT/OT'),
                    ('Servers', 'Install APO on L4 server', 'Vendor'),
                    ('Servers', 'Configure install account access to server and DB', 'IT/OT'),
                    ('Servers', 'Dedicated station for ACM install and all scripts', 'Site'),
                    ('Database', 'Set up SQL server accounts and permissions', 'IT/OT'),
                    ('Database', 'Enable email notifications', 'IT/OT'),
                    ('Database', 'Create SQL jobs', 'Custom'),
                    ('Database', 'Create MarathonCustom DB', 'Custom'),
                ]),
                ('2. Custom Solutions Readiness', [
                    ('Validation', 'Test custom solutions in APO environment', 'Custom'),
                    ('Validation', 'Document any gaps/limitations', 'Custom'),
                    ('Validation', 'Plan mitigation if not ready', 'Custom'),
                    ('Backups', '⚠️ CRITICAL: Consistent backup across all servers', 'Custom'),
                ]),
                ('3. Known ACM-APO Gaps Review', [
                    ('Gap Assessment', 'Review list of missing features', 'Site'),
                    ('Gap Assessment', 'Identify impact on workflows', 'Site'),
                    ('Gap Assessment', 'Plan workarounds or custom solutions', 'Custom'),
                    ('Gap Assessment', '⚠️ WARNING: Old import files not valid anymore', 'Custom'),
                ]),
                ('4. Installation Order', [
                    ('Sequence', '⚠️ Do not install without cleanup and hierarchy adjustments first', 'Site'),
                    ('Sequence', 'Confirm HAMR installed and optimized', 'Site'),
                    ('Sequence', 'Confirm ACM cleanup complete', 'Site'),
                    ('Sequence', 'Confirm hierarchies aligned', 'Site'),
                ]),
            ]
        },
        
        '4-OSW-Completion': {
            'phase_name': 'Phase 4: Onsite Scoping Workbook (OSW)',
            'sections': [
                ('1. OSW Data Entry', [
                    ('Assets', 'Complete Asset Hierarchy (Site→Area→Unit→System→Subsystem)', 'Site'),
                    ('Assets', 'Validate hierarchy matches FDS standards', 'Site'),
                    ('Devices', 'Complete Device Inventory (manufacturer, model, protocol)', 'Site'),
                    ('Devices', 'Classify external reference devices', 'Site'),
                    ('Devices', 'Document unsupported devices with justification', 'Site'),
                ]),
                ('2. External Reference Configuration', [
                    ('ControlLogix', 'Provide .L5X/.L5K export and I/O config', 'Site'),
                    ('PLC5/SLC', 'Provide .RSS export and addressing scheme', 'Site'),
                    ('Triconex', 'Confirm EPKS External Reference or TPS import', 'Site'),
                    ('Bently Nevada', 'Confirm DCS External Reference method', 'Site'),
                    ('Analyzers', 'Decide manual entry vs automated import', 'Site'),
                    ('Documentation', 'Document connection strings and update frequencies', 'Site'),
                ]),
                ('3. OSW Validation & Sign-Off', [
                    ('Review', 'Hexagon reviews OSW for completeness', 'Vendor'),
                    ('Review', 'Site confirms all devices in scope', 'Site'),
                    ('Review', '⚠️ NO additions after sign-off without Change Order', 'Site'),
                    ('Review', 'IT reviews network architecture', 'IT/OT'),
                    ('Sign-Off', 'Executive sponsor signs off (scope locked)', 'Site'),
                    ('Timeline', 'Hexagon provides deployment timeline/resource estimate', 'Vendor'),
                    ('Budget', 'Site confirms budget availability', 'Site'),
                ]),
            ]
        },
        
        '5-Deployment-Readiness': {
            'phase_name': 'Phase 5: Migration Execution Readiness',
            'sections': [
                ('1. Final Backups', [
                    ('Backup', 'Release all ACM consoles', 'Site'),
                    ('Backup', 'Backup ACM DBs and files', 'Site'),
                    ('Backup', 'Export all consoles and all versions', 'Site'),
                    ('Backup', 'Export all EMDBs', 'Site'),
                    ('Backup', 'Purge M&R to keep ~6 months history', 'Site'),
                ]),
                ('2. General Pre-Migration Checks', [
                    ('Validation', 'ACM Test Server operational', 'Site'),
                    ('Validation', 'Hierarchies optimal - EMDB/ACM', 'Site'),
                    ('Validation', 'Hierarchies optimal - M&R', 'Site'),
                    ('Validation', 'ACM DB cleanup complete', 'Site'),
                    ('Validation', 'M&R cleanup complete', 'Site'),
                    ('Validation', 'Folders with scripts copied to new servers', 'Site'),
                ]),
                ('3. ACM DB Pre-Migration Preparation', [
                    ('Safety', 'Prevent accidental Enforce', 'Site'),
                    ('Safety', 'Prevent OPC overload', 'Site'),
                    ('Enforcements', 'Disable all Enforcements', 'Site'),
                    ('Tag Syncs', 'Re-run and then Disable TagSyncs', 'Site'),
                    ('Enforcements', 'Change enforcer name in DB (not UI)', 'Site'),
                    ('Tag Cleanup', 'Delete all Deleted Tags', 'Site'),
                    ('Tag Cleanup', 'Ensure only 2 versions of variables', 'Site'),
                    ('BMA', 'Delete BMA points if APO < 3.1', 'Site'),
                ]),
                ('4. ACM Health Checks (Repeat from Cleanup)', [
                    ('Health', 'Run all health check scripts again', 'Custom'),
                    ('Health', 'Compare to pre-cleanup baseline', 'Site'),
                    ('Health', 'Address any new issues', 'Site'),
                ]),
            ]
        },
        
        '6-Migration-Execution': {
            'phase_name': 'Phase 6: APO Migration Execution',
            'sections': [
                ('1. Initial Migration', [
                    ('Migration', 'Run APO migration tool', 'Vendor'),
                    ('Migration', 'Import DCS hierarchy, alarm config, I/O', 'Vendor'),
                    ('Migration', 'Import control strategies and HMI displays', 'Vendor'),
                    ('Migration', 'Import PLC programs (if applicable)', 'Vendor'),
                    ('Migration', 'Import Safety system configuration', 'Vendor'),
                    ('Migration', 'Import external reference systems', 'Vendor'),
                    ('Validation', 'Spot-check 10% of tags for accuracy', 'Site'),
                    ('Validation', 'Document import statistics', 'Vendor'),
                ]),
                ('2. Initial APO Configuration', [
                    ('Console Setup', 'Configure APO consoles', 'Vendor'),
                    ('Console Setup', 'Map APO consoles to HAMR Operating Positions', 'Site'),
                    ('User Setup', 'Configure user permissions', 'IT/OT'),
                    ('Notifications', 'Set up email notifications', 'IT/OT'),
                    ('Health Checks', 'Deploy APO health monitoring', 'Custom'),
                ]),
            ]
        },
        
        '7-Validation-GoLive': {
            'phase_name': 'Phase 7: Validation & Parallel Operations',
            'sections': [
                ('1. Site Acceptance Testing (SAT)', [
                    ('Testing', 'Verify asset search functionality', 'Site'),
                    ('Testing', 'Verify signal tracing', 'Site'),
                    ('Testing', 'Verify external reference data updates', 'Site'),
                    ('Testing', 'Verify alarm configuration accuracy', 'Site'),
                    ('Testing', 'Verify reports generation', 'Site'),
                    ('Testing', 'Verify user permissions', 'Site'),
                    ('Testing', 'Document SAT results (Pass/Fail)', 'Site'),
                ]),
                ('2. Parallel Operations (30-90 days)', [
                    ('Parallel', 'Run ACM and APO concurrently', 'Site'),
                    ('Parallel', 'Compare enforcement results', 'Site'),
                    ('Parallel', 'Monitor performance and stability', 'Site'),
                    ('Parallel', 'Train users on APO workflow differences', 'Site'),
                    ('Parallel', 'Address issues and defects', 'Site/Vendor'),
                ]),
                ('3. User Training', [
                    ('Training', 'M&R refresher training', 'Vendor'),
                    ('Training', 'ACM-APO "same functionality" + what was lost', 'Site'),
                    ('Training', 'Basic new APO workflow', 'Vendor'),
                    ('Training', 'Advanced rationalization expert training', 'Custom'),
                    ('Training', 'Provide user guides and quick reference cards', 'Vendor'),
                ]),
                ('4. Go-Live Preparation', [
                    ('Freeze', 'Freeze ACM configuration (no changes)', 'Site'),
                    ('Freeze', 'Finalize APO database (all SAT defects resolved)', 'Vendor'),
                    ('Config', 'Configure automated data refresh (nightly)', 'Vendor'),
                    ('Backup', 'Configure SQL backup schedule (daily full + hourly log)', 'IT/OT'),
                    ('Access', 'Publish L5 web URL to users', 'IT/OT'),
                    ('Support', 'Create support contact list', 'Site'),
                    ('Communication', 'Announce go-live date and instructions', 'Site'),
                    ('Final Backup', 'Execute final pre-production backup', 'IT/OT'),
                ]),
            ]
        },
        
        '8-Cutover': {
            'phase_name': 'Phase 8: Cutover & Decommission',
            'sections': [
                ('1. Cutover Day', [
                    ('Switch', 'Health check scripts validation', 'Custom'),
                    ('Switch', 'Verify Experion Servers Alarm Help', 'Site'),
                    ('Switch', 'Verify Dynamic mode management', 'Site'),
                    ('Switch', 'Verify Custom solutions operational', 'Custom'),
                    ('Switch', 'Switch TagSync (check Windows tasks)', 'Site'),
                    ('Switch', 'Disable ACM enforcements (check Windows tasks)', 'Site'),
                    ('Switch', '⚠️ Ensure no Enforcements or TagSyncs overlapping', 'Site'),
                    ('Switch', 'Enable APO as primary system', 'Site'),
                    ('Monitoring', 'Monitor first 24 hours continuously', 'Site'),
                ]),
                ('2. Post-Cutover Support (2-4 weeks)', [
                    ('Support', 'Monitor data import jobs for failures', 'Vendor'),
                    ('Support', 'Address user-reported issues within 48 hours', 'Site'),
                    ('Support', 'Track usage metrics (logins, searches, reports)', 'IT/OT'),
                    ('Support', '2-week post-go-live checkpoint meeting', 'Site'),
                    ('Support', '1-month post-go-live review (lessons learned)', 'Site'),
                ]),
                ('3. ACM Decommission', [
                    ('Freeze', 'Freeze ACM using change history log', 'Site'),
                    ('Delta', 'Review ACM change logs for delta migration', 'Site'),
                    ('Delta', 'Perform delta import to APO (if needed)', 'Vendor'),
                    ('Decommission', '⚠️ KEEP ACM SERVER powered off but available', 'Site'),
                    ('Cleanup', 'Update host files', 'IT/OT'),
                    ('Cleanup', 'Un-DSA EAS/ACM server', 'IT/OT'),
                    ('Cleanup', 'Uninstall remaining ACM clients', 'IT/OT'),
                    ('Cleanup', 'Disable L4 ACM Web', 'IT/OT'),
                    ('Cleanup', 'Disable replication (if not needed)', 'IT/OT'),
                ]),
            ]
        },
        
        '9-Post-Migration': {
            'phase_name': 'Phase 9: Post-Migration Optimization',
            'sections': [
                ('1. Performance Monitoring (First 90 days)', [
                    ('Metrics', 'Monitor alarm performance KPIs (vs baseline)', 'Site'),
                    ('Metrics', 'Track APO rationalization suggestions quality', 'Site'),
                    ('Metrics', 'Monitor database performance', 'IT/OT'),
                    ('Metrics', 'Track user adoption and feedback', 'Site'),
                ]),
                ('2. Continuous Improvement', [
                    ('Optimization', 'Refine alarm priorities based on APO analysis', 'Site'),
                    ('Optimization', 'Implement APO rationalization recommendations', 'Site'),
                    ('Optimization', 'Enhance custom tools based on user feedback', 'Custom'),
                    ('Optimization', 'Update procedures and training materials', 'Site'),
                ]),
                ('3. Standards Compliance Review', [
                    ('Audit', 'ISA 18.2 compliance assessment post-migration', 'Site'),
                    ('Audit', 'EEMUA 191 KPI validation', 'Site'),
                    ('Audit', 'Update Alarm Philosophy document', 'Site'),
                ]),
            ]
        },
    }
    
    return phases

def add_status_dropdown(ws, start_row, end_row):
    """Add dropdown for Status column with validation"""
    dv = DataValidation(
        type="list",
        formula1='"☐ Not Started,⏳ In Progress,✓ Complete,✗ Blocked,⚠️ On Hold"',
        allow_blank=True
    )
    dv.error = 'Please select from dropdown'
    dv.errorTitle = 'Invalid Status'
    dv.prompt = 'Select task status'
    dv.promptTitle = 'Task Status'
    ws.add_data_validation(dv)
    dv.add(f'E{start_row}:E{end_row}')

def format_worksheet(ws, phase_name):
    """Apply formatting to worksheet with enhanced features"""
    
    # Set column widths
    ws.column_dimensions['A'].width = 8   # Item #
    ws.column_dimensions['B'].width = 35  # Section
    ws.column_dimensions['C'].width = 60  # Task
    ws.column_dimensions['D'].width = 12  # Owner
    ws.column_dimensions['E'].width = 15  # Status
    ws.column_dimensions['F'].width = 30  # Notes
    ws.column_dimensions['G'].width = 15  # Prerequisites
    
    # Header row
    ws['A1'] = 'Item #'
    ws['B1'] = 'Section'
    ws['C1'] = 'Task'
    ws['D1'] = 'Owner'
    ws['E1'] = 'Status'
    ws['F1'] = 'Notes'
    ws['G1'] = 'Prerequisites'
    
    # Format header
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        cell = ws[f'{col}1']
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    
    ws.row_dimensions[1].height = 30
    
    # Freeze panes
    ws.freeze_panes = 'A2'

def create_workbook():
    """Create the comprehensive checklist workbook"""
    
    print("Creating ACM to APO Migration Comprehensive Checklist...")
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create Executive Summary
    print("Creating Executive Summary...")
    df_exec = create_executive_summary()
    ws_exec = wb.create_sheet("Executive Summary")
    
    # Write executive summary with formatting
    ws_exec['A1'] = 'ACM TO APO MIGRATION - EXECUTIVE SUMMARY'
    ws_exec['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_exec['A1'].fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws_exec.merge_cells('A1:B1')
    ws_exec.row_dimensions[1].height = 25
    
    ws_exec['A2'] = 'Critical Success Factors for Site Readiness'
    ws_exec['A2'].font = Font(bold=True, size=12)
    ws_exec.merge_cells('A2:B2')
    
    ws_exec['A3'] = 'Based on Marathon Petroleum pilots and multiple site assessments'
    ws_exec.merge_cells('A3:B3')
    
    for r_idx, row in enumerate(dataframe_to_rows(df_exec, index=False, header=False), start=5):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_exec.cell(row=r_idx, column=c_idx, value=value)
            
            # Format sections
            if value and isinstance(value, str):
                if value.startswith('⚠️') or 'REASONS FOR' in value:
                    cell.fill = CRITICAL_FILL
                    cell.font = Font(bold=True, size=11)
                elif value.startswith('📊') or 'SUCCESS FACTORS' in value:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(bold=True, size=11)
                elif value[0].isdigit() and '. ' in value[:4]:
                    cell.font = Font(bold=True)
                    cell.fill = WARNING_FILL
                    
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    ws_exec.column_dimensions['A'].width = 40
    ws_exec.column_dimensions['B'].width = 80
    
    # Create phase sheets
    phases = create_phase_sheets()
    
    for phase_key, phase_data in phases.items():
        print(f"Creating {phase_data['phase_name']}...")
        
        ws = wb.create_sheet(phase_key)
        format_worksheet(ws, phase_data['phase_name'])
        
        # Add phase name
        ws['A2'] = phase_data['phase_name']
        ws['A2'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A2'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A2:G2')
        ws.row_dimensions[2].height = 25
        
        row_idx = 3
        section_num = 1
        item_num = 1
        section_start_row = None
        
        for section_name, tasks in phase_data['sections']:
            # Section header
            section_start_row = row_idx
            ws[f'A{row_idx}'] = f'{section_num}.'
            ws[f'B{row_idx}'] = section_name
            ws[f'A{row_idx}'].font = SECTION_FONT
            ws[f'B{row_idx}'].font = SECTION_FONT
            ws[f'A{row_idx}'].fill = SECTION_FILL
            ws[f'B{row_idx}'].fill = SECTION_FILL
            ws.merge_cells(f'C{row_idx}:G{row_idx}')
            ws[f'C{row_idx}'].fill = SECTION_FILL
            ws.row_dimensions[row_idx].height = 20
            
            # Create outline group for section
            ws.row_dimensions.group(row_idx, row_idx, hidden=False)
            
            row_idx += 1
            subsection_start = row_idx
            
            # Tasks grouped by subsection
            current_subsection = None
            subsection_items = []
            
            for subsection, task, owner in tasks:
                if current_subsection != subsection:
                    # New subsection header
                    if current_subsection is not None:
                        # Group previous subsection tasks
                        if len(subsection_items) > 0:
                            ws.row_dimensions.group(subsection_items[0], subsection_items[-1], hidden=True)
                        subsection_items = []
                    
                    current_subsection = subsection
                    ws[f'B{row_idx}'] = f'  {subsection}'
                    ws[f'B{row_idx}'].font = SUBSECTION_FONT
                    ws[f'B{row_idx}'].fill = SUBSECTION_FILL
                    ws.merge_cells(f'A{row_idx}:G{row_idx}')
                    ws[f'A{row_idx}'].fill = SUBSECTION_FILL
                    ws.row_dimensions[row_idx].height = 18
                    row_idx += 1
                
                # Task row with auto-numbering formula
                ws[f'A{row_idx}'] = f'{section_num}.{item_num}'
                ws[f'B{row_idx}'] = f'    {subsection}'
                ws[f'C{row_idx}'] = task
                ws[f'D{row_idx}'] = owner
                ws[f'E{row_idx}'] = '☐ Not Started'
                ws[f'F{row_idx}'] = ''  # Notes
                ws[f'G{row_idx}'] = ''  # Prerequisites
                
                subsection_items.append(row_idx)
                
                # Format cells
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    cell = ws[f'{col}{row_idx}']
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    cell.border = THIN_BORDER
                    cell.font = NORMAL_FONT
                    
                    # Highlight critical/warning tasks
                    if '⚠️' in task or 'CRITICAL' in task:
                        cell.fill = CRITICAL_FILL
                    elif 'WARNING' in task or 'CAUTION' in task:
                        cell.fill = WARNING_FILL
                
                ws.row_dimensions[row_idx].height = 30
                row_idx += 1
                item_num += 1
            
            # Group final subsection
            if len(subsection_items) > 0:
                ws.row_dimensions.group(subsection_items[0], subsection_items[-1], hidden=True)
            
            row_idx += 1  # Blank row between sections
            section_num += 1
            item_num = 1
        
        # Add status dropdown (conditional formatting removed to avoid Excel errors)
        add_status_dropdown(ws, 3, row_idx)
    
    # Save workbook
    output_file = 'ACM_to_APO_Migration_Comprehensive_Checklist.xlsx'
    wb.save(output_file)
    print(f"\n✅ Checklist created: {output_file}")
    print(f"   Total sheets: {len(wb.sheetnames)}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")

if __name__ == '__main__':
    create_workbook()
