"""Final summary of usage data in PKS sheet"""
from openpyxl import load_workbook

wb = load_workbook('data/output/Experion_License_Report_20260128_143237.xlsx')
ws = wb['PKS']

license_types_with_usage = []
license_types_without_usage = []

# Scan all rows
for row_num in range(1, ws.max_row + 1):
    cell_value = ws.cell(row_num, 1).value
    
    if cell_value and isinstance(cell_value, str) and cell_value.strip() == '  - Used':
        # Get license type from previous row
        if row_num > 1:
            license_type = ws.cell(row_num - 1, 1).value
            if license_type and isinstance(license_type, str) and not license_type.startswith('  -'):
                # Check if this row has any usage data
                has_data = False
                for col in range(2, ws.max_column + 1):
                    val = ws.cell(row_num, col).value
                    if val and val != 0:
                        has_data = True
                        break
                
                if has_data:
                    license_types_with_usage.append(license_type.strip())
                else:
                    license_types_without_usage.append(license_type.strip())

print("="*70)
print(f" PKS Sheet Usage Data Summary")
print("="*70)
print(f"\n✅ License types WITH usage data displayed ({len(license_types_with_usage)}):\n")
for lt in license_types_with_usage:
    print(f"   ✓ {lt}")

if license_types_without_usage:
    print(f"\n❌ License types WITHOUT usage data ({len(license_types_without_usage)}):\n")
    for lt in license_types_without_usage:
        print(f"   ✗ {lt}")

print(f"\n{'='*70}")
total = len(license_types_with_usage) + len(license_types_without_usage)
if total > 0:
    pct = 100 * len(license_types_with_usage) / total
    print(f"Total: {len(license_types_with_usage)}/{total} license types have usage data ({pct:.1f}%)")
print("="*70)
