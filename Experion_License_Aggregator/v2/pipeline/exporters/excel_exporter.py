"""Excel exporter for license data with multi-sheet support and formatting."""

from pathlib import Path
from typing import Dict, List, Optional
from datetime import datetime
import yaml

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .base_exporter import BaseExporter, ExportResult
from ...models.license import LicenseData
from ...models.usage import UsageData
from ...models.cost import CostCalculation
from ...models.transfer import TransferCandidate


class ExcelExporter(BaseExporter):
    """Export license data to Excel with multiple formatted sheets."""
    
    def __init__(self, output_dir: Path):
        """Initialize Excel exporter.
        
        Args:
            output_dir: Directory for Excel files
        """
        super().__init__(output_dir)
        self.format = "xlsx"
        
        # Load cluster friendly names
        self.cluster_names = self._load_cluster_names()
        
        # XML→CSV license type mapping (from field_mappings.yaml)
        # XML files use different names than CSV files for some license types
        # This maps XML license type → standardized CSV name used in usage data
        self.xml_to_csv_mapping = {
            'DIRECTSTATIONS': 'CONSOLE_STATIONS',        # Console station(s)
            'MULTISTATIONS': 'MULTISTATIONS',             # Multi window flex station(s)
            'OPER_TOUCH_PANEL': 'OPER_TOUCH_PANEL',      # Operator touch panel(s)
            'MODICON': 'MODICON',                         # Modbus
            'OPCCLIENT': 'OPCCLIENT',                     # OPC client interface
            'DIRECTCLIENTS': 'DIRECTCLIENTS',             # Console extension station(s)
            'MULTI_COUNT': 'MULTI_COUNT',                 # Distributed server(s)
            'CDA_IO_ANA': 'CDA_IO_ANA',                   # Analog IO Point(s)
            'CDA_IO_DIG': 'CDA_IO_DIG',                   # Digital IO Point(s)
            'STATIONS': 'STATIONS',                       # Flex station(s)
            'PROCESSPOINTS': 'PROCESSPOINTS',             # Process point(s)
            'SCADAPOINTS': 'SCADAPOINTS',                 # SCADA point(s)
            # Most other fields map 1:1
        }
        
        # Style definitions
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.data_font = Font(size=10)
        self.data_alignment = Alignment(horizontal="left", vertical="center")
        self.number_alignment = Alignment(horizontal="right", vertical="center")
        
        self.border_side = Side(style="thin", color="000000")
        self.cell_border = Border(
            left=self.border_side,
            right=self.border_side,
            top=self.border_side,
            bottom=self.border_side
        )
    
    def _load_cluster_names(self) -> Dict[str, str]:
        """Load cluster friendly name mappings from config.
        
        Returns:
            Dictionary mapping technical cluster names to friendly names.
            Falls back to original name if config not found or cluster not mapped.
        """
        config_path = Path('config/cluster_names.yaml')
        if not config_path.exists():
            return {}  # Return empty dict, will use original names
        
        try:
            with open(config_path, 'r') as f:
                config = yaml.safe_load(f)
                return config.get('cluster_names', {})
        except Exception as e:
            print(f"[WARNING] Failed to load cluster names config: {e}")
            return {}
    
    def export(self, data, filename: str, **options) -> ExportResult:
        """Export data to Excel (implements BaseExporter abstract method).
        
        This method routes to export_licenses or export_comprehensive based on data type.
        
        Args:
            data: License data (list or dict)
            filename: Output filename
            **options: Export options (usage_data, costs, transfers for comprehensive)
        
        Returns:
            ExportResult with export details
        """
        if isinstance(data, dict):
            # Comprehensive export with multiple data types
            return self.export_comprehensive(
                licenses=data.get('licenses', []),
                usage_data=data.get('usage_data'),
                costs=data.get('costs'),
                transfers=data.get('transfers'),
                output_filename=filename
            )
        else:
            # Simple license export
            return self.export_licenses(data, filename)
    
    def export_licenses(
        self,
        licenses: List[LicenseData],
        output_filename: str = "licenses.xlsx"
    ) -> ExportResult:
        """Export license data to Excel.
        
        Args:
            licenses: List of license records
            output_filename: Output file name
            
        Returns:
            ExportResult with export details
        """
        try:
            self._ensure_output_dir()
            output_path = self.output_dir / output_filename
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Create PKS sheet
            ws_pks = wb.create_sheet("PKS Licenses", 0)
            self._write_pks_sheet(ws_pks, licenses)
            
            # Create summary sheet
            ws_summary = wb.create_sheet("Summary", 1)
            self._write_summary_sheet(ws_summary, licenses)
            
            # Save workbook
            wb.save(output_path)
            
            return self._create_success_result(
                output_path=output_path,
                format=self.format,
                record_count=len(licenses),
                sheets=wb.sheetnames,
                total_licenses=len(licenses)
            )
            
        except Exception as e:
            return self._create_error_result(
                format=self.format,
                error_message=f"Failed to export licenses: {str(e)}"
            )
    
    def export_comprehensive(
        self,
        licenses: List[LicenseData],
        usage_data: Optional[List[UsageData]] = None,
        costs: Optional[List[CostCalculation]] = None,
        transfers: Optional[List[TransferCandidate]] = None,
        output_filename: str = "comprehensive_report.xlsx"
    ) -> ExportResult:
        """Export comprehensive report with multiple sheets.
        
        Args:
            licenses: License data
            usage_data: Optional usage data
            costs: Optional cost calculations
            transfers: Optional transfer candidates
            output_filename: Output file name
            
        Returns:
            ExportResult with export details
        """
        try:
            self._ensure_output_dir()
            output_path = self.output_dir / output_filename
            wb = Workbook()
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            sheet_index = 0
            sheets_created = []
            
            # PKS sheet (transposed format - MSIDs as columns)
            if licenses:
                ws_pks = wb.create_sheet("PKS", sheet_index)
                self._write_pks_transposed_sheet(ws_pks, licenses, usage_data)
                sheets_created.append("PKS")
                sheet_index += 1
            
            # Summary sheet
            ws_summary = wb.create_sheet("Summary", sheet_index)
            self._write_summary_sheet(ws_summary, licenses, usage_data, costs, transfers)
            sheets_created.append("Summary")
            sheet_index += 1
            
            # Usage sheet
            if usage_data:
                ws_usage = wb.create_sheet("Usage Data", sheet_index)
                self._write_usage_sheet(ws_usage, usage_data)
                sheets_created.append("Usage Data")
                sheet_index += 1
            
            # Costs sheet
            if costs:
                ws_costs = wb.create_sheet("Cost Analysis", sheet_index)
                self._write_costs_sheet(ws_costs, costs)
                sheets_created.append("Cost Analysis")
                sheet_index += 1
            
            # Transfer Candidates sheet
            if transfers:
                ws_transfers = wb.create_sheet("Transfer Candidates", sheet_index)
                self._write_transfers_sheet(ws_transfers, transfers)
                sheets_created.append("Transfer Candidates")
                sheet_index += 1
            
            # Save workbook
            wb.save(output_path)
            
            total_records = (
                len(licenses) +
                (len(usage_data) if usage_data else 0) +
                (len(costs) if costs else 0) +
                (len(transfers) if transfers else 0)
            )
            
            return self._create_success_result(
                output_path=output_path,
                format=self.format,
                record_count=total_records,
                sheets=sheets_created,
                license_count=len(licenses),
                usage_count=len(usage_data) if usage_data else 0,
                cost_count=len(costs) if costs else 0,
                transfer_count=len(transfers) if transfers else 0
            )
            
        except Exception as e:
            return self._create_error_result(
                format=self.format,
                error_message=f"Failed to export comprehensive report: {str(e)}"
            )
    
    def _write_pks_sheet(self, ws, licenses: List[LicenseData]) -> None:
        """Write PKS licenses to worksheet.
        
        Args:
            ws: Worksheet object
            licenses: License data
        """
        # Headers
        headers = ["MSID", "System Number", "Cluster", "Release"]
        
        # Collect all unique license types across all licenses
        all_license_types = set()
        for lic in licenses:
            all_license_types.update(lic.licensed.keys())
        
        license_type_list = sorted(all_license_types)
        headers.extend(license_type_list)
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border
        
        # Write data rows
        for row_idx, lic in enumerate(licenses, start=2):
            ws.cell(row=row_idx, column=1, value=lic.msid).border = self.cell_border
            ws.cell(row=row_idx, column=2, value=lic.system_number).border = self.cell_border
            ws.cell(row=row_idx, column=3, value=lic.cluster).border = self.cell_border
            ws.cell(row=row_idx, column=4, value=lic.release).border = self.cell_border
            
            # Write license counts for each type
            for col_idx, lic_type in enumerate(license_type_list, start=5):
                count = lic.licensed.get(lic_type, 0)
                cell = ws.cell(row=row_idx, column=col_idx, value=count)
                cell.alignment = self.number_alignment
                cell.border = self.cell_border
        
        # Column widths
        ws.column_dimensions['A'].width = 15  # MSID
        ws.column_dimensions['B'].width = 15  # System Number
        ws.column_dimensions['C'].width = 12  # Cluster
        ws.column_dimensions['D'].width = 10  # Release
        for col_idx in range(5, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        
        # Freeze panes
        ws.freeze_panes = "E2"
    
    def _write_summary_sheet(
        self,
        ws,
        licenses: List[LicenseData],
        usage_data: Optional[List[UsageData]] = None,
        costs: Optional[List[CostCalculation]] = None,
        transfers: Optional[List[TransferCandidate]] = None
    ) -> None:
        """Write summary statistics to worksheet.
        
        Args:
            ws: Worksheet object
            licenses: License data
            usage_data: Optional usage data
            costs: Optional cost calculations
            transfers: Optional transfer candidates
        """
        row = 1
        
        # Title
        title_cell = ws.cell(row=row, column=1, value="License Report Summary")
        title_cell.font = Font(bold=True, size=14)
        row += 2
        
        # Timestamp
        ws.cell(row=row, column=1, value="Generated:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        row += 2
        
        # License counts
        ws.cell(row=row, column=1, value="License Statistics").font = Font(bold=True, size=12)
        row += 1
        
        ws.cell(row=row, column=1, value="Total Systems:")
        ws.cell(row=row, column=2, value=len(licenses))
        row += 1
        
        # Cluster breakdown
        clusters = {}
        for lic in licenses:
            clusters[lic.cluster] = clusters.get(lic.cluster, 0) + 1
        
        ws.cell(row=row, column=1, value="By Cluster:")
        row += 1
        for cluster, count in sorted(clusters.items()):
            ws.cell(row=row, column=2, value=f"{cluster}:")
            ws.cell(row=row, column=3, value=count)
            row += 1
        
        row += 1
        
        # Usage statistics
        if usage_data:
            ws.cell(row=row, column=1, value="Usage Statistics").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="Total Usage Records:")
            ws.cell(row=row, column=2, value=len(usage_data))
            row += 2
        
        # Cost statistics
        if costs:
            ws.cell(row=row, column=1, value="Cost Analysis").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="Total Cost Records:")
            ws.cell(row=row, column=2, value=len(costs))
            row += 1
            
            total_cost = sum(cost.total_cost for cost in costs)
            ws.cell(row=row, column=1, value="Total Cost:")
            cost_cell = ws.cell(row=row, column=2, value=total_cost)
            cost_cell.number_format = '"$"#,##0.00'
            row += 2
        
        # Transfer statistics
        if transfers:
            ws.cell(row=row, column=1, value="Transfer Opportunities").font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="Total Transfer Candidates:")
            ws.cell(row=row, column=2, value=len(transfers))
            row += 1
            
            total_excess_value = sum(t.excess_value for t in transfers)
            ws.cell(row=row, column=1, value="Total Excess Value:")
            value_cell = ws.cell(row=row, column=2, value=total_excess_value)
            value_cell.number_format = '"$"#,##0.00'
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    def _write_usage_sheet(self, ws, usage_data: List[UsageData]) -> None:
        """Write usage data to worksheet.
        
        Args:
            ws: Worksheet object
            usage_data: Usage records
        """
        # Headers
        headers = ["MSID", "License Type", "Used Count"]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border
        
        # Data rows
        for row_idx, usage in enumerate(usage_data, start=2):
            ws.cell(row=row_idx, column=1, value=usage.msid)
            ws.cell(row=row_idx, column=2, value=usage.license_type)
            ws.cell(row=row_idx, column=3, value=usage.used_quantity)
            
            # Apply formatting
            for col_idx in range(1, 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = self.data_font
                cell.border = self.cell_border
                if col_idx == 3:
                    cell.alignment = self.number_alignment
                else:
                    cell.alignment = self.data_alignment
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        
        ws.freeze_panes = "A2"
    
    def _write_costs_sheet(self, ws, costs: List[CostCalculation]) -> None:
        """Write cost data to worksheet.
        
        Args:
            ws: Worksheet object
            costs: Cost calculations
        """
        # Headers
        headers = [
            "MSID", "System Number", "License Type",
            "Licensed Quantity", "Unit Price", "Total Cost", "Price Source"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border
        
        # Data rows
        for row_idx, cost in enumerate(costs, start=2):
            ws.cell(row=row_idx, column=1, value=cost.msid).border = self.cell_border
            ws.cell(row=row_idx, column=2, value=cost.system_number).border = self.cell_border
            ws.cell(row=row_idx, column=3, value=cost.license_type).border = self.cell_border
            
            cell = ws.cell(row=row_idx, column=4, value=cost.licensed_quantity)
            cell.alignment = self.number_alignment
            cell.border = self.cell_border
            
            cell = ws.cell(row=row_idx, column=5, value=cost.unit_price)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = self.number_alignment
            cell.border = self.cell_border
            
            cell = ws.cell(row=row_idx, column=6, value=cost.total_cost)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = self.number_alignment
            cell.border = self.cell_border
            
            ws.cell(row=row_idx, column=7, value=cost.price_source).border = self.cell_border
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        
        ws.freeze_panes = "A2"
    
    def _write_transfers_sheet(self, ws, transfers: List[TransferCandidate]) -> None:
        """Write transfer candidates to worksheet.
        
        Args:
            ws: Worksheet object
            transfers: Transfer candidates
        """
        # Headers
        headers = [
            "MSID", "System Number", "Cluster", "License Type",
            "Licensed", "Used", "Excess", "Excess Value", "Unit Price", "Priority"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border
        
        # Data rows
        for row_idx, transfer in enumerate(transfers, start=2):
            ws.cell(row=row_idx, column=1, value=transfer.msid).border = self.cell_border
            ws.cell(row=row_idx, column=2, value=transfer.system_number).border = self.cell_border
            ws.cell(row=row_idx, column=3, value=transfer.cluster).border = self.cell_border
            ws.cell(row=row_idx, column=4, value=transfer.license_type).border = self.cell_border
            
            cell = ws.cell(row=row_idx, column=5, value=int(transfer.licensed_quantity) if transfer.licensed_quantity else 0)
            cell.alignment = self.number_alignment
            cell.border = self.cell_border
            
            cell = ws.cell(row=row_idx, column=6, value=int(transfer.used_quantity) if transfer.used_quantity else 0)
            cell.alignment = self.number_alignment
            cell.border = self.cell_border
            
            cell = ws.cell(row=row_idx, column=7, value=int(transfer.excess_quantity) if transfer.excess_quantity else 0)
            cell.alignment = self.number_alignment
            cell.border = self.cell_border
            
            cell = ws.cell(row=row_idx, column=8, value=transfer.excess_value)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = self.number_alignment
            cell.border = self.cell_border
            
            cell = ws.cell(row=row_idx, column=9, value=transfer.unit_price)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = self.number_alignment
            cell.border = self.cell_border
            
            priority_cell = ws.cell(row=row_idx, column=10, value=transfer.priority)
            priority_cell.border = self.cell_border
            
            # Color code by priority
            if transfer.priority == "HIGH":
                priority_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                priority_cell.font = Font(bold=True, color="FFFFFF")
            elif transfer.priority == "MEDIUM":
                priority_cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 12
        
        ws.freeze_panes = "E2"
    
    def _write_pks_transposed_sheet(self, ws, licenses: List[LicenseData], usage_data: Optional[List] = None) -> None:
        """Write PKS licenses in transposed format (MSIDs as columns, fields as rows).
        
        This format matches the EXAMPLE_Output.xlsx PKS sheet structure where:
        - Column 1: Field names (MSID, Cluster, Release, license types, etc.)
        - Columns 2+: Each system's data
        - Includes usage and utilization % rows for each license type
        - Shows only the latest version of each MSID
        
        Args:
            ws: Worksheet object
            licenses: License data
            usage_data: Optional usage data for showing utilization
        """
        # Deduplicate licenses - keep only the latest version (highest file_version) for each unique system
        # Key: (cluster, msid, system_number) - treats same MSID with different system numbers as separate
        msid_to_latest = {}
        for lic in licenses:
            key = (lic.cluster, lic.msid, lic.system_number)
            if key not in msid_to_latest or lic.file_version > msid_to_latest[key].file_version:
                msid_to_latest[key] = lic
        
        # Filter out "00000" system numbers if real system numbers exist for the same MSID
        # "00000" means no system number - these are older/incomplete XML files
        filtered_licenses = {}
        msid_has_real_numbers = {}  # Track which (cluster, msid) pairs have non-00000 systems
        
        # First pass: identify which MSIDs have real system numbers
        for key, lic in msid_to_latest.items():
            cluster_msid = (lic.cluster, lic.msid)
            if lic.system_number != "00000":
                msid_has_real_numbers[cluster_msid] = True
        
        # Second pass: exclude 00000 systems if real system numbers exist for that MSID
        for key, lic in msid_to_latest.items():
            cluster_msid = (lic.cluster, lic.msid)
            # Skip 00000 systems if this MSID has real system numbers
            if lic.system_number == "00000" and msid_has_real_numbers.get(cluster_msid):
                continue  # Skip this duplicate
            filtered_licenses[key] = lic
        
        # Use filtered list
        licenses = list(filtered_licenses.values())
        licenses.sort(key=lambda x: (x.cluster, x.msid, x.system_number))  # Sort for consistent order
        
        # Build usage lookup: (system_name, msid, license_type) -> used_quantity
        # Note: usage data uses CSV field names (e.g., CONSOLE_STATIONS)
        # while licenses use XML field names (e.g., DIRECTSTATIONS)
        # We need to map XML → CSV when looking up usage
        usage_lookup = {}
        if usage_data:
            for usage in usage_data:
                key = (usage.system_name, usage.msid, usage.license_type)
                usage_lookup[key] = usage.used_quantity
        
        # Create reverse lookup: (system_name, msid, xml_license_type) -> used_quantity
        # This maps XML license type → CSV license type for lookup
        usage_lookup_by_xml = {}
        for lic in licenses:
            for xml_license_type in lic.licensed.keys():
                # Map XML type to CSV type (e.g., DIRECTSTATIONS → CONSOLE_STATIONS)
                csv_license_type = self.xml_to_csv_mapping.get(xml_license_type, xml_license_type)
                
                # Check if usage data exists for this CSV type
                key_csv = (lic.system_name, lic.msid, csv_license_type)
                if key_csv in usage_lookup:
                    # Store in XML-keyed lookup for easy access
                    key_xml = (lic.system_name, lic.msid, xml_license_type)
                    usage_lookup_by_xml[key_xml] = usage_lookup[key_csv]
        
        # Define utilization threshold colors (from DATA_PROCESSING_GUIDE.md)
        util_color_high = PatternFill(start_color="FFC00000", end_color="FFC00000", fill_type="solid")      # Red: >= 80%
        util_color_medium = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")    # Yellow: >= 50%
        util_color_low = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")       # Green: < 50%
        
        # Define data cell color (light green background for most cells)
        data_light_green = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
        
        # Column 1 is "Field" header (light blue like original)
        field_cell = ws.cell(row=1, column=1, value="Field")
        field_cell.font = self.header_font
        field_cell.fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
        field_cell.alignment = self.header_alignment
        field_cell.border = self.cell_border
        
        # Columns 2+ are system names (MSID/cluster/system) - no color (default white)
        for col_idx, lic in enumerate(licenses, start=2):
            # Only show system number if it's not 00000 (default/unknown)
            if lic.system_number == "00000":
                system_name = f"{lic.msid} ({lic.cluster})"
            else:
                system_name = f"{lic.msid}-{lic.system_number} ({lic.cluster})"
            header_cell = ws.cell(row=1, column=col_idx, value=system_name)
            header_cell.font = self.header_font
            header_cell.fill = self.header_fill  # Standard dark blue header
            header_cell.alignment = self.header_alignment
            header_cell.border = self.cell_border
        
        # Define field rows
        row_idx = 2
        
        # LICENSE INFO section header
        section_cell = ws.cell(row=row_idx, column=1, value="LICENSE INFO")
        section_cell.font = Font(bold=True, size=11)
        section_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        section_cell.border = self.cell_border
        row_idx += 1
        
        # Basic info fields
        basic_fields = [
            ("Cluster", lambda lic: lic.system_name if lic.system_name else lic.cluster),  # Use system name (ESVT0, HCU, etc.)
            ("MSID", lambda lic: lic.msid),
            ("System Number", lambda lic: lic.system_number),
            ("Release", lambda lic: lic.release),
            ("License Version", lambda lic: Path(lic.file_path).name if lic.file_path else ""),  # XML filename
            ("Customer", lambda lic: lic.customer or "")
        ]
        
        for field_name, field_func in basic_fields:
            field_cell = ws.cell(row=row_idx, column=1, value=field_name)
            field_cell.font = Font(bold=True)
            field_cell.border = self.cell_border
            
            for col_idx, lic in enumerate(licenses, start=2):
                value = field_func(lic)
                data_cell = ws.cell(row=row_idx, column=col_idx, value=value)
                data_cell.border = self.cell_border
                data_cell.alignment = self.data_alignment
                # Light green background for LICENSE INFO section
                data_cell.fill = data_light_green
            
            row_idx += 1
        
        row_idx += 1  # Blank row
        
        # Collect all license types and organize by category
        all_license_types = set()
        for lic in licenses:
            all_license_types.update(lic.licensed.keys())
        
        # Define category mappings - which category each type belongs to
        # This allows ANY license type from XML to appear in the right category
        type_to_category = {
            # POINTS
            "PROCESSPOINTS": "POINTS", "SCADAPOINTS": "POINTS", 
            "CDA_IO_ANA": "POINTS", "CDA_IO_DIG": "POINTS",
            # STATIONS
            "STATIONS": "STATIONS", "DIRECTSTATIONS": "STATIONS", 
            "MULTISTATIONS": "STATIONS", "DIRECTCLIENTS": "STATIONS",
            "CONSOLE_STATIONS": "STATIONS", "OPER_TOUCH_PANEL": "STATIONS", 
            "COLLABORATION_STATIONS": "STATIONS",
            # REDUNDANCY
            "DUAL": "REDUNDANCY", "MULTI_SERVER": "REDUNDANCY",
            # INTERFACES
            "TPS": "INTERFACES", "OPC_DA": "INTERFACES", "OPC_UA_CLIENT": "INTERFACES",
            "OPC_HDA": "INTERFACES", "OPC_AE": "INTERFACES", "FSC": "INTERFACES",
            "MODICON": "INTERFACES", "AB": "INTERFACES", "AB_ETH": "INTERFACES",
            "AB_INT": "INTERFACES", "DNP3": "INTERFACES", "CDA": "INTERFACES",
            # TPS Options (ELCN controllers and access points)
            "TPS_ELCN": "INTERFACES", "TPS_AM_AP": "INTERFACES", "TPS_HM_AP": "INTERFACES",
            "TPS_ENIM_AP_RED": "INTERFACES", "TPS_EPLCG_AP": "INTERFACES", "TPS_VIRT_AP": "INTERFACES",
            # FEATURES
            "DAS": "FEATURES", "API": "FEATURES", "SQL": "FEATURES",
            "DSPBLD": "FEATURES", "LAS": "FEATURES", "MEDE": "FEATURES",
            "LOTRPT": "FEATURES", "EQUIPMENT_DISPLAY": "FEATURES",
            "DISPLAY_REPOSITORY": "FEATURES", "USR_SCAN": "FEATURES"
        }
        
        # Group license types by category (all types from XML included)
        categories = {"POINTS": [], "STATIONS": [], "REDUNDANCY": [], "INTERFACES": [], "FEATURES": [], "OTHER": []}
        for lic_type in sorted(all_license_types):
            category = type_to_category.get(lic_type, "OTHER")
            categories[category].append(lic_type)
        
        # Write categorized license types
        for category in ["POINTS", "STATIONS", "REDUNDANCY", "INTERFACES", "FEATURES", "OTHER"]:
            existing_types = categories[category]
            
            if existing_types:
                # Category header
                cat_cell = ws.cell(row=row_idx, column=1, value=category)
                cat_cell.font = Font(bold=True, size=11)
                cat_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cat_cell.border = self.cell_border
                row_idx += 1
                
                # License types in category
                for lic_type in existing_types:
                    # License quantity row
                    field_cell = ws.cell(row=row_idx, column=1, value=lic_type)
                    field_cell.font = Font(bold=True)
                    field_cell.border = self.cell_border
                    
                    for col_idx, lic in enumerate(licenses, start=2):
                        count = lic.licensed.get(lic_type, 0)
                        data_cell = ws.cell(row=row_idx, column=col_idx, value=count)
                        data_cell.border = self.cell_border
                        data_cell.alignment = self.number_alignment
                        # All license quantity cells get light green background
                        data_cell.fill = data_light_green
                    
                    row_idx += 1
                    
                    # Usage row ("  - Used")
                    if usage_data:
                        used_cell = ws.cell(row=row_idx, column=1, value="  - Used")
                        used_cell.font = Font(italic=True)
                        used_cell.border = self.cell_border
                        
                        for col_idx, lic in enumerate(licenses, start=2):
                            # Use XML-keyed lookup (maps XML type to CSV usage data)
                            used_qty = usage_lookup_by_xml.get((lic.system_name, lic.msid, lic_type), 0)
                            data_cell = ws.cell(row=row_idx, column=col_idx, value=used_qty if used_qty > 0 else None)
                            data_cell.border = self.cell_border
                            data_cell.alignment = self.number_alignment
                            # Usage rows get light green
                            data_cell.fill = data_light_green
                        
                        row_idx += 1
                        
                        # Utilization % row ("  - Utilization %") with color coding by threshold
                        util_cell = ws.cell(row=row_idx, column=1, value="  - Utilization %")
                        util_cell.font = Font(italic=True)
                        util_cell.border = self.cell_border
                        
                        for col_idx, lic in enumerate(licenses, start=2):
                            licensed_qty = lic.licensed.get(lic_type, 0)
                            # Use XML-keyed lookup (maps XML type to CSV usage data)
                            used_qty = usage_lookup_by_xml.get((lic.system_name, lic.msid, lic_type), 0)
                            
                            if licensed_qty > 0 and used_qty > 0:
                                utilization = (used_qty / licensed_qty) * 100
                                data_cell = ws.cell(row=row_idx, column=col_idx, value=f"{utilization:.1f}%")
                                
                                # Apply color based on utilization threshold
                                if utilization >= 80:
                                    data_cell.fill = util_color_high      # Red: High utilization
                                elif utilization >= 50:
                                    data_cell.fill = util_color_medium    # Yellow: Medium utilization
                                else:
                                    data_cell.fill = util_color_low       # Green: Low utilization
                            else:
                                data_cell = ws.cell(row=row_idx, column=col_idx, value=None)
                            
                            data_cell.border = self.cell_border
                            data_cell.alignment = self.number_alignment
                        
                        row_idx += 1
                
                row_idx += 1  # Blank row after category
        
        # Set column widths
        ws.column_dimensions['A'].width = 25  # Field names
        for col_idx in range(2, len(licenses) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Freeze panes at B2
        ws.freeze_panes = "B2"
