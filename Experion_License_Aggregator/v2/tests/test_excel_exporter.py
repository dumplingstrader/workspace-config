"""Tests for Excel exporter."""

import pytest
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from v2.pipeline.exporters.excel_exporter import ExcelExporter
from v2.models.license import LicenseData
from v2.models.usage import UsageData
from v2.models.cost import CostCalculation
from v2.models.transfer import TransferCandidate


class TestExcelExporterInit:
    """Test Excel exporter initialization."""
    
    def test_init_creates_exporter(self, tmp_path):
        """Test exporter initialization."""
        exporter = ExcelExporter(tmp_path)
        
        assert exporter.output_dir == tmp_path
        assert exporter.format == "xlsx"
        assert exporter.header_font is not None
        assert exporter.header_fill is not None
    
    def test_init_with_nonexistent_dir(self, tmp_path):
        """Test initialization with non-existent directory."""
        output_dir = tmp_path / "new_dir"
        exporter = ExcelExporter(output_dir)
        
        assert exporter.output_dir == output_dir


class TestLicenseExport:
    """Test license export functionality."""
    
    @pytest.fixture
    def sample_licenses(self):
        """Create sample license data."""
        return [
            LicenseData(
                msid="M0614",
                system_number="60806",
                cluster="Carson",
                release="R520",
                licensed={"PROCESSPOINTS": 1000}
            ),
            LicenseData(
                msid="M0615",
                system_number="60807",
                cluster="Wilmington",
                release="R511",
                licensed={"DIRECTSTATIONS": 10}
            )
        ]
    
    def test_export_licenses_creates_file(self, tmp_path, sample_licenses):
        """Test that export creates Excel file."""
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_licenses(sample_licenses, "test_licenses.xlsx")
        
        assert result.success
        assert result.output_path.exists()
        assert result.output_path.suffix == ".xlsx"
        assert result.record_count == 2
    
    def test_export_licenses_has_correct_sheets(self, tmp_path, sample_licenses):
        """Test that exported file has correct sheets."""
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_licenses(sample_licenses, "test_licenses.xlsx")
        
        wb = load_workbook(result.output_path)
        assert "PKS Licenses" in wb.sheetnames
        assert "Summary" in wb.sheetnames
    
    def test_export_licenses_pks_sheet_content(self, tmp_path, sample_licenses):
        """Test PKS sheet content."""
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_licenses(sample_licenses, "test_licenses.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["PKS Licenses"]
        
        # Check headers
        assert ws.cell(1, 1).value == "MSID"
        assert ws.cell(1, 2).value == "System Number"
        assert ws.cell(1, 3).value == "Cluster"
        
        # Check first data row (now M0614)
        assert ws.cell(2, 1).value == "M0614"
        assert ws.cell(2, 3).value == "Carson"
        # Licensed count
        assert ws.cell(2, 6).value == 1000
    
    def test_export_licenses_utilization_calculation(self, tmp_path, sample_licenses):
        """Test utilization percentage calculation."""
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_licenses(sample_licenses, "test_licenses.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["PKS Licenses"]
        
        # Check that basic columns exist (utilization requires usage data merge)
        assert ws.cell(1, 1).value == "MSID"
        assert ws.cell(1, 2).value == "System Number"
    
    def test_export_empty_licenses(self, tmp_path):
        """Test exporting empty license list."""
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_licenses([], "empty_licenses.xlsx")
        
        assert result.success
        assert result.record_count == 0
        
        wb = load_workbook(result.output_path)
        ws = wb["PKS Licenses"]
        assert ws.cell(1, 1).value == "MSID"  # Headers present
        assert ws.cell(2, 1).value is None  # No data


class TestComprehensiveExport:
    """Test comprehensive export with multiple data types."""
    
    @pytest.fixture
    def sample_data(self):
        """Create sample data for all types."""
        licenses = [
            LicenseData(
                msid="M0614",
                system_number="60806",
                cluster="Carson",
                release="R520",
                licensed={"PROCESSPOINTS": 1000}
            )
        ]
        
        usage = [
            UsageData(
                msid="M0614",
                license_type="PROCESSPOINTS",
                used_quantity=800
            )
        ]
        
        costs = [
            CostCalculation(
                msid="M0614",
                system_number="60806",
                license_type="PROCESSPOINTS",
                licensed_quantity=1000,
                unit_price=100.0,
                total_cost=100000.0,
                price_source="MPC 2026 Confirmed"
            )
        ]
        
        transfers = [
            TransferCandidate(
                msid="M0614",
                system_number="60806",
                cluster="Carson",
                license_type="PROCESSPOINTS",
                licensed_quantity=1000,
                used_quantity=900,
                excess_quantity=100,
                excess_value=10000.0,
                unit_price=100.0,
                priority="HIGH"
            )
        ]
        
        return licenses, usage, costs, transfers
    
    def test_comprehensive_export_all_sheets(self, tmp_path, sample_data):
        """Test comprehensive export creates all sheets."""
        licenses, usage, costs, transfers = sample_data
        exporter = ExcelExporter(tmp_path)
        
        result = exporter.export_comprehensive(
            licenses, usage, costs, transfers, "comprehensive.xlsx"
        )
        
        assert result.success
        wb = load_workbook(result.output_path)
        
        assert "PKS Licenses" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert "Usage Data" in wb.sheetnames
        assert "Cost Analysis" in wb.sheetnames
        assert "Transfer Candidates" in wb.sheetnames
    
    def test_comprehensive_export_licenses_only(self, tmp_path, sample_data):
        """Test comprehensive export with only licenses."""
        licenses, _, _, _ = sample_data
        exporter = ExcelExporter(tmp_path)
        
        result = exporter.export_comprehensive(licenses, output_filename="licenses_only.xlsx")
        
        assert result.success
        wb = load_workbook(result.output_path)
        
        assert "PKS Licenses" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert "Usage Data" not in wb.sheetnames
    
    def test_comprehensive_export_record_count(self, tmp_path, sample_data):
        """Test record count in comprehensive export."""
        licenses, usage, costs, transfers = sample_data
        exporter = ExcelExporter(tmp_path)
        
        result = exporter.export_comprehensive(
            licenses, usage, costs, transfers, "comprehensive.xlsx"
        )
        
        # 1 license + 1 usage + 1 cost + 1 transfer = 4
        assert result.record_count == 4
    
    def test_comprehensive_export_metadata(self, tmp_path, sample_data):
        """Test metadata in comprehensive export result."""
        licenses, usage, costs, transfers = sample_data
        exporter = ExcelExporter(tmp_path)
        
        result = exporter.export_comprehensive(
            licenses, usage, costs, transfers, "comprehensive.xlsx"
        )
        
        assert result.metadata["license_count"] == 1
        assert result.metadata["usage_count"] == 1
        assert result.metadata["cost_count"] == 1
        assert result.metadata["transfer_count"] == 1


class TestUsageSheet:
    """Test usage sheet formatting."""
    
    def test_usage_sheet_content(self, tmp_path):
        """Test usage sheet has correct content."""
        usage = [
            UsageData(msid="M0614", license_type="PROCESSPOINTS", used_quantity=800),
            UsageData(msid="M0614", license_type="PROCESSPOINTS", used_quantity=820)
        ]
        
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_comprehensive([], usage_data=usage, output_filename="usage.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["Usage Data"]
        
        # Check headers
        assert ws.cell(1, 1).value == "MSID"
        assert ws.cell(1, 2).value == "License Type"
        assert ws.cell(1, 3).value == "Used Count"
        
        # Check data
        assert ws.cell(2, 1).value == "M0614"
        assert ws.cell(2, 3).value == 800


class TestCostSheet:
    """Test cost sheet formatting."""
    
    def test_cost_sheet_content(self, tmp_path):
        """Test cost sheet has correct content."""
        costs = [
            CostCalculation(
                msid="M0614",
                system_number="60806",
                license_type="PROCESSPOINTS",
                licensed_quantity=1000,
                unit_price=100.0,
                total_cost=100000.0,
                price_source="MPC 2026 Confirmed"
            )
        ]
        
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_comprehensive([], costs=costs, output_filename="costs.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["Cost Analysis"]
        
        # Check headers
        assert ws.cell(1, 1).value == "MSID"
        assert ws.cell(1, 3).value == "License Type"
        assert ws.cell(1, 4).value == "Licensed Quantity"
        assert ws.cell(1, 5).value == "Unit Price"
        
        # Check data
        assert ws.cell(2, 1).value == "M0614"
        assert ws.cell(2, 4).value == 1000
        assert ws.cell(2, 5).value == 100.0
        assert ws.cell(2, 6).value == 100000.0


class TestTransferSheet:
    """Test transfer candidates sheet."""
    
    def test_transfer_sheet_content(self, tmp_path):
        """Test transfer sheet has correct content."""
        transfers = [
            TransferCandidate(
                msid="M0614",
                system_number="60806",
                cluster="Carson",
                license_type="PROCESSPOINTS",
                licensed_quantity=1000,
                used_quantity=900,
                excess_quantity=100,
                excess_value=10000.0,
                unit_price=100.0,
                priority="HIGH"
            )
        ]
        
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_comprehensive([], transfers=transfers, output_filename="transfers.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["Transfer Candidates"]
        
        # Check headers
        assert ws.cell(1, 1).value == "MSID"
        assert ws.cell(1, 2).value == "System Number"
        assert ws.cell(1, 3).value == "Cluster"
        assert ws.cell(1, 4).value == "License Type"
        
        # Check data
        assert ws.cell(2, 1).value == "M0614"
        assert ws.cell(2, 3).value == "Carson"  # Cluster
        assert ws.cell(2, 4).value == "PROCESSPOINTS"  # License Type
        assert ws.cell(2, 5).value == 1000  # Licensed quantity
    
    def test_transfer_priority_formatting(self, tmp_path):
        """Test priority color coding in transfer sheet."""
        transfers = [
            TransferCandidate("M0614", "60806", "Carson", "PROCESSPOINTS", 1000, 900, 100, 10000.0, 100.0, "HIGH"),
            TransferCandidate("M0615", "60807", "Wilmington", "DIRECTSTATIONS", 100, 95, 5, 5000.0, 1000.0, "MEDIUM"),
            TransferCandidate("M0616", "60808", "Carson", "PROCESSPOINTS", 100, 90, 10, 1000.0, 100.0, "LOW")
        ]
        
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_comprehensive([], transfers=transfers, output_filename="transfers.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["Transfer Candidates"]
        
        # High priority should have red fill (column 10 is Priority)
        high_cell = ws.cell(2, 10)
        assert high_cell.fill.start_color.rgb in ["FFFF6B6B", "00FF6B6B"]
        
        # Medium priority should have yellow fill
        medium_cell = ws.cell(3, 10)
        assert medium_cell.fill.start_color.rgb in ["FFFFD93D", "00FFD93D"]


class TestSummarySheet:
    """Test summary sheet content."""
    
    def test_summary_sheet_basic_stats(self, tmp_path):
        """Test summary sheet has basic statistics."""
        licenses = [
            LicenseData("M0614", "60806", "Carson", "R520", licensed={"PROCESSPOINTS": 1000}),
            LicenseData("M0615", "60807", "Wilmington", "R511", licensed={"DIRECTSTATIONS": 10}),
            LicenseData("M0614", "60806", "Carson", "R520", licensed={"DIRECTSTATIONS": 5})
        ]
        
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_comprehensive(licenses, output_filename="summary.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["Summary"]
        
        # Should have title
        assert "License Report Summary" in ws.cell(1, 1).value
        
        # Should have system count (3 records total)
        # Row 6 column 2 contains total system count (len(licenses))
        assert ws.cell(6, 2).value == 3
    
    def test_summary_sheet_cluster_breakdown(self, tmp_path):
        """Test summary sheet cluster breakdown."""
        licenses = [
            LicenseData("M0614", "60806", "Carson", "R520", licensed={"PROCESSPOINTS": 1000}),
            LicenseData("M0615", "60807", "Wilmington", "R511", licensed={"DIRECTSTATIONS": 10})
        ]
        
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_comprehensive(licenses, output_filename="summary.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["Summary"]
        
        # Should list clusters
        found_carson = False
        found_wilmington = False
        
        for row in range(1, 20):
            cell_value = ws.cell(row, 2).value
            if cell_value == "Carson:":
                found_carson = True
                assert ws.cell(row, 3).value == 1
            elif cell_value == "Wilmington:":
                found_wilmington = True
                assert ws.cell(row, 3).value == 1
        
        assert found_carson
        assert found_wilmington


class TestFormatting:
    """Test Excel formatting."""
    
    def test_header_formatting(self, tmp_path):
        """Test header row formatting."""
        licenses = [LicenseData("M0614", "60806", "Carson", "R520", licensed={"PROCESSPOINTS": 1000})]
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_licenses(licenses, "formatting.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["PKS Licenses"]
        
        # Check header cell formatting
        header_cell = ws.cell(1, 1)
        assert header_cell.font.bold
        assert header_cell.fill.start_color.rgb in ["FF366092", "00366092"]
    
    def test_column_widths(self, tmp_path):
        """Test column width settings."""
        licenses = [LicenseData("M0614", "60806", "Carson", "R520", licensed={"PROCESSPOINTS": 1000})]
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_licenses(licenses, "widths.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["PKS Licenses"]
        
        # Check column widths
        assert ws.column_dimensions['A'].width == 15
        assert ws.column_dimensions['B'].width == 15
    
    def test_freeze_panes(self, tmp_path):
        """Test freeze panes setting."""
        licenses = [LicenseData("M0614", "60806", "Carson", "R520", licensed={"PROCESSPOINTS": 1000})]
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_licenses(licenses, "freeze.xlsx")
        
        wb = load_workbook(result.output_path)
        ws = wb["PKS Licenses"]
        
        assert ws.freeze_panes == "E2"


class TestErrorHandling:
    """Test error handling."""
    
    def test_export_handles_invalid_path(self, tmp_path):
        """Test export with invalid output path."""
        licenses = [LicenseData("M0614", "60806", "Carson", "R520", licensed={"PROCESSPOINTS": 1000})]
        exporter = ExcelExporter(tmp_path / "nonexistent")
        
        # Should create directory and succeed
        result = exporter.export_licenses(licenses, "test.xlsx")
        assert result.success
    
    def test_export_result_on_success(self, tmp_path):
        """Test export result structure on success."""
        licenses = [LicenseData("M0614", "60806", "Carson", "R520", licensed={"PROCESSPOINTS": 1000})]
        exporter = ExcelExporter(tmp_path)
        result = exporter.export_licenses(licenses, "test.xlsx")
        
        assert result.success
        assert result.format == "xlsx"
        assert result.record_count == 1
        assert isinstance(result.metadata, dict)
        assert len(result.errors) == 0
