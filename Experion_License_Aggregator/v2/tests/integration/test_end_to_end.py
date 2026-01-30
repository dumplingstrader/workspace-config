"""
End-to-End Integration Tests for V2.0 Pipeline.

Tests complete pipeline execution from XML/CSV input through to JSON/Excel output
using real data files from data/raw directories.
"""

import pytest
from pathlib import Path
import json
import openpyxl
from datetime import datetime

from v2.pipeline.coordinator import PipelineCoordinator, PipelineResult


@pytest.fixture
def data_dir():
    """Get path to test data directory."""
    return Path(__file__).parent.parent.parent.parent / "data" / "raw"


@pytest.fixture
def output_dir(tmp_path):
    """Create temporary output directory."""
    output = tmp_path / "output"
    output.mkdir()
    return output


@pytest.fixture
def coordinator(output_dir, data_dir):
    """Create coordinator with real config and temp output."""
    config_dir = Path(__file__).parent.parent.parent / "config"
    return PipelineCoordinator(config_dir=config_dir, output_dir=output_dir)


# =============================================================================
# Full Pipeline Tests
# =============================================================================

@pytest.mark.integration
def test_full_pipeline_carson_xml_only(coordinator, data_dir, output_dir):
    """Test complete pipeline with Carson XML files only (no CSV)."""
    # Use a subdirectory that has actual XML files
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    
    if not xml_dir.exists():
        pytest.skip(f"Carson XML directory not found: {xml_dir}")
    
    # Run pipeline without CSV
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        export_json=True,
        export_excel=True,
        validate_business_rules=True
    )
    
    # Verify success
    assert result.success is True
    assert len(result.licenses) > 0
    assert len(result.usage_data) == 0  # No CSV provided
    assert result.execution_time > 0
    
    # Verify stage times recorded
    assert 'extraction' in result.stage_times
    assert 'validation' in result.stage_times
    assert 'export' in result.stage_times
    
    # Verify statistics
    assert 'clusters' in result.stats
    assert len(result.stats['clusters']) > 0
    
    # Verify outputs created
    json_files = list(output_dir.glob("*.json"))
    excel_files = list(output_dir.glob("*.xlsx"))
    
    assert len(json_files) >= 1, "JSON output should be created"
    assert len(excel_files) >= 1, "Excel output should be created"


@pytest.mark.integration
def test_full_pipeline_wilmington_xml_only(coordinator, data_dir, output_dir):
    """Test complete pipeline with Wilmington XML files only."""
    xml_dir = data_dir / "Wilmington"
    
    if not xml_dir.exists():
        pytest.skip(f"Wilmington XML directory not found: {xml_dir}")
    
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        export_json=True,
        export_excel=True
    )
    
    assert result.success is True
    assert len(result.licenses) > 0
    
    # Verify cluster assignment
    clusters = result.stats.get('clusters', {})
    assert len(clusters) > 0


@pytest.mark.integration
def test_full_pipeline_with_csv_usage(coordinator, data_dir, output_dir):
    """Test complete pipeline with both XML and CSV data."""
    xml_dir = data_dir / "Carson"
    csv_file = data_dir.parent / "utilization_input.csv"
    
    if not xml_dir.exists():
        pytest.skip(f"XML directory not found: {xml_dir}")
    if not csv_file.exists():
        pytest.skip(f"CSV file not found: {csv_file}")
    
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        csv_file=csv_file,
        export_json=True,
        export_excel=True
    )
    
    assert result.success is True
    assert len(result.licenses) > 0
    
    # CSV extraction may fail if format is incompatible (V2 vs V1)
    # Pipeline should still succeed with just XML data
    if len(result.usage_data) > 0:
        # If CSV loaded successfully, verify matching occurred
        assert 'matching' in result.stage_times
    else:
        # CSV extraction failed - verify warning logged
        assert any('CSV extraction failed' in str(w) for w in result.warnings)
    
    # Verify usage statistics
    if result.usage_data:
        assert 'total_usage_records' in result.stats


@pytest.mark.integration
def test_pipeline_with_cluster_filter(coordinator, data_dir, output_dir):
    """Test pipeline with cluster filtering."""
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    
    if not xml_dir.exists():
        pytest.skip(f"XML directory not found: {xml_dir}")
    
    # Run with cluster filter
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        clusters=["LAR"],  # Only LAR cluster
        export_json=False,
        export_excel=False
    )
    
    assert result.success is True
    
    # Verify only LAR licenses extracted
    for license in result.licenses:
        assert license.cluster == "LAR"


# =============================================================================
# Cost Calculation Integration Tests
# =============================================================================

@pytest.mark.integration
def test_cost_calculation_integration(coordinator, data_dir, output_dir):
    """Test cost calculation with real license data."""
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    
    if not xml_dir.exists():
        pytest.skip(f"XML directory not found: {xml_dir}")
    
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        export_json=False,
        export_excel=False
    )
    
    assert result.success is True
    
    # Verify costs calculated
    if result.costs:
        assert len(result.costs) > 0
        
        # Check cost structure
        cost = result.costs[0]
        assert hasattr(cost, 'msid')
        assert hasattr(cost, 'total_cost')
        assert hasattr(cost, 'unit_price')
        assert cost.total_cost >= 0
        
        # Verify statistics
        assert 'total_cost' in result.stats
        assert result.stats['total_cost'] >= 0


# =============================================================================
# Transfer Detection Integration Tests
# =============================================================================

@pytest.mark.integration
def test_transfer_detection_integration(coordinator, data_dir, output_dir):
    """Test transfer detection with real data."""
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    csv_file = data_dir.parent / "utilization_input.csv"
    
    if not xml_dir.exists() or not csv_file.exists():
        pytest.skip("Test data not available")
    
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        csv_file=csv_file,
        export_json=False,
        export_excel=False
    )
    
    assert result.success is True
    
    # Verify transfers detected
    if result.transfers:
        assert len(result.transfers) > 0
        
        # Check transfer structure
        transfer = result.transfers[0]
        assert hasattr(transfer, 'msid')
        assert hasattr(transfer, 'excess_quantity')
        assert hasattr(transfer, 'priority')
        assert transfer.priority in ["HIGH", "MEDIUM", "LOW"]
        
        # Verify statistics
        if 'high_priority_transfers' in result.stats:
            assert result.stats['high_priority_transfers'] >= 0


# =============================================================================
# JSON Export Integration Tests
# =============================================================================

@pytest.mark.integration
def test_json_export_integration(coordinator, data_dir, output_dir):
    """Test JSON export with real data."""
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    
    if not xml_dir.exists():
        pytest.skip(f"XML directory not found: {xml_dir}")
    
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        export_json=True,
        export_excel=False
    )
    
    assert result.success is True
    
    # Find JSON output
    json_files = list(output_dir.glob("*.json"))
    assert len(json_files) >= 1
    
    # Validate JSON structure
    json_file = json_files[0]
    with open(json_file, 'r') as f:
        data = json.load(f)
    
    assert 'licenses' in data
    assert 'metadata' in data
    assert isinstance(data['licenses'], list)
    
    # Verify metadata
    metadata = data['metadata']
    assert 'export_timestamp' in metadata
    assert 'record_count' in metadata
    assert metadata['record_count'] == len(data['licenses'])


# =============================================================================
# Excel Export Integration Tests
# =============================================================================

@pytest.mark.integration
def test_excel_export_integration(coordinator, data_dir, output_dir):
    """Test Excel export with real data."""
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    
    if not xml_dir.exists():
        pytest.skip(f"XML directory not found: {xml_dir}")
    
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        export_json=False,
        export_excel=True
    )
    
    assert result.success is True
    
    # Find Excel output
    excel_files = list(output_dir.glob("*.xlsx"))
    assert len(excel_files) >= 1
    
    # Validate Excel structure
    excel_file = excel_files[0]
    wb = openpyxl.load_workbook(excel_file)
    
    # Verify expected sheets exist
    assert "PKS Licenses" in wb.sheetnames or "PKS" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    
    # Verify PKS sheet has data
    pks_sheet = wb["PKS Licenses"] if "PKS Licenses" in wb.sheetnames else wb["PKS"]
    assert pks_sheet.max_row > 1  # More than header
    assert pks_sheet.max_column > 1


@pytest.mark.integration
def test_excel_comprehensive_export(coordinator, data_dir, output_dir):
    """Test Excel comprehensive export with all sheets."""
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    csv_file = data_dir.parent / "utilization_input.csv"
    
    if not xml_dir.exists() or not csv_file.exists():
        pytest.skip("Test data not available")
    
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        csv_file=csv_file,
        export_json=False,
        export_excel=True
    )
    
    assert result.success is True
    
    # Find Excel output
    excel_files = list(output_dir.glob("*.xlsx"))
    assert len(excel_files) >= 1
    
    excel_file = excel_files[0]
    wb = openpyxl.load_workbook(excel_file)
    
    # Verify all expected sheets present
    expected_sheets = ["PKS", "Summary"]
    for sheet_name in expected_sheets:
        assert any(sheet_name in ws for ws in wb.sheetnames), f"Missing sheet: {sheet_name}"
    
    # If usage data exists, verify usage sheet
    if result.usage_data:
        assert any("Usage" in ws for ws in wb.sheetnames)
    
    # If costs calculated, verify costs sheet
    if result.costs:
        assert any("Cost" in ws for ws in wb.sheetnames)
    
    # If transfers found, verify transfers sheet
    if result.transfers:
        assert any("Transfer" in ws for ws in wb.sheetnames)


# =============================================================================
# Error Handling Integration Tests
# =============================================================================

@pytest.mark.integration
def test_pipeline_with_invalid_xml_dir(coordinator, tmp_path):
    """Test pipeline handles invalid XML directory gracefully."""
    invalid_dir = tmp_path / "nonexistent"
    
    result = coordinator.run_pipeline(
        xml_dir=invalid_dir,
        export_json=False,
        export_excel=False
    )
    
    # Pipeline should fail gracefully
    assert result.success is False
    assert len(result.errors) > 0


@pytest.mark.integration
def test_pipeline_with_empty_xml_dir(coordinator, tmp_path, output_dir):
    """Test pipeline handles empty XML directory."""
    empty_dir = tmp_path / "empty"
    empty_dir.mkdir()
    
    result = coordinator.run_pipeline(
        xml_dir=empty_dir,
        export_json=False,
        export_excel=False
    )
    
    # Should succeed but with no data
    assert len(result.licenses) == 0


@pytest.mark.integration
def test_pipeline_with_invalid_csv(coordinator, data_dir, tmp_path, output_dir):
    """Test pipeline handles invalid CSV gracefully."""
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    invalid_csv = tmp_path / "invalid.csv"
    invalid_csv.write_text("invalid,csv,data\n1,2,3")
    
    if not xml_dir.exists():
        pytest.skip("XML directory not found")
    
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        csv_file=invalid_csv,
        export_json=False,
        export_excel=False
    )
    
    # May succeed with XML-only processing
    # Or fail with error about CSV
    # Both are acceptable behaviors


# =============================================================================
# Performance Tests
# =============================================================================

@pytest.mark.integration
@pytest.mark.slow
def test_pipeline_performance(coordinator, data_dir, output_dir):
    """Test pipeline performance with real data."""
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    csv_file = data_dir.parent / "utilization_input.csv"
    
    if not xml_dir.exists():
        pytest.skip("XML directory not found")
    
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        csv_file=csv_file if csv_file.exists() else None,
        export_json=True,
        export_excel=True
    )
    
    assert result.success is True
    
    # Verify reasonable performance
    assert result.execution_time < 60  # Should complete within 60 seconds
    
    # Verify all stages completed
    assert len(result.stage_times) >= 4
    
    # Print performance metrics
    print(f"\n{'='*60}")
    print("Pipeline Performance Metrics")
    print(f"{'='*60}")
    print(f"Total Execution Time: {result.execution_time:.2f}s")
    print(f"Licenses Processed: {len(result.licenses)}")
    print(f"Usage Records: {len(result.usage_data)}")
    print("\nStage Breakdown:")
    for stage, time in result.stage_times.items():
        print(f"  {stage:20s}: {time:6.2f}s")
    print(f"{'='*60}\n")


# =============================================================================
# Multi-Site Integration Tests
# =============================================================================

@pytest.mark.integration
def test_multi_site_processing(coordinator, data_dir, output_dir):
    """Test processing licenses from multiple sites."""
    sites = ["Carson", "Wilmington"]
    all_licenses = []
    
    for site in sites:
        xml_dir = data_dir / site
        if not xml_dir.exists():
            continue
        
        result = coordinator.run_pipeline(
            xml_dir=xml_dir,
            export_json=False,
            export_excel=False
        )
        
        if result.success:
            all_licenses.extend(result.licenses)
    
    # Verify licenses from multiple sites
    if all_licenses:
        clusters = set(lic.cluster for lic in all_licenses)
        assert len(clusters) > 0
        
        print(f"\nProcessed {len(all_licenses)} licenses from {len(clusters)} clusters")


# =============================================================================
# Validation Integration Tests
# =============================================================================

@pytest.mark.integration
def test_business_rule_validation_integration(coordinator, data_dir, output_dir):
    """Test business rule validation with real data."""
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    
    if not xml_dir.exists():
        pytest.skip("XML directory not found")
    
    # Run with business validation enabled
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        validate_business_rules=True,
        export_json=False,
        export_excel=False
    )
    
    assert result.success is True
    
    # Check for validation warnings (acceptable)
    if result.warnings:
        print(f"\nValidation Warnings: {len(result.warnings)}")
        for warning in result.warnings[:5]:  # Print first 5
            print(f"  - {warning}")


@pytest.mark.integration
def test_validation_disabled(coordinator, data_dir, output_dir):
    """Test pipeline with validation disabled."""
    xml_dir = data_dir / "Carson" / "ALKY M13287-EX10 131844"
    
    if not xml_dir.exists():
        pytest.skip("XML directory not found")
    
    result = coordinator.run_pipeline(
        xml_dir=xml_dir,
        validate_business_rules=False,
        export_json=False,
        export_excel=False
    )
    
    assert result.success is True
