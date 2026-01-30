"""Check PKS sheet row structure"""
from openpyxl import load_workbook

wb = load_workbook('data/output/Experion_License_Report_20260128_143237.xlsx')
ws = wb['PKS']

print("First 40 rows, column A:")
for i in range(1, 41):
    cell_value = ws.cell(i, 1).value
    repr_value = repr(cell_value)
    print(f"Row {i:2d}: {repr_value}")
    if cell_value and isinstance(cell_value, str) and 'Used' in cell_value:
        # Check if row has any data
        has_data = False
        for col in range(2, 10):
            val = ws.cell(i, col).value
            if val and val != 0:
                has_data = True
                break
        print(f"         ^ Has usage data: {has_data}")
