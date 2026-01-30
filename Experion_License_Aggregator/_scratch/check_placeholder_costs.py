"""Check placeholder costs in production data"""
from openpyxl import load_workbook
import json

# Load placeholder list
with open('v2/config/cost_catalog.json') as f:
    catalog = json.load(f)

placeholders = {}
for key, value in catalog.items():
    if not key.startswith('_') and isinstance(value, dict) and value.get('is_placeholder'):
        placeholders[key] = {
            'unit_cost': value.get('unit_cost', 100),
            'per': value.get('per', 1),
            'category': value.get('category', 'Unknown')
        }

print(f'Total placeholder entries in catalog: {len(placeholders)}')
print()

# Load Excel report and check PKS sheet
wb = load_workbook('data/output/Experion_License_Report_20260128_143237.xlsx', data_only=True)
ws_pks = wb['PKS']

# Get license type names from row 1 (skipping first column which is labels)
license_types_in_use = set()
for row in range(2, ws_pks.max_row + 1):
    license_type = ws_pks.cell(row, 1).value
    if license_type and not license_type.startswith('$'):  # Skip dollar rows
        license_types_in_use.add(license_type)

# Find which placeholders are actually used
used_placeholders = {}
for lic_type in license_types_in_use:
    if lic_type in placeholders:
        used_placeholders[lic_type] = placeholders[lic_type]

print('PLACEHOLDER COSTS IN USE (NEED VENDOR QUOTES):')
print('=' * 70)
print(f'Found {len(used_placeholders)} license types with placeholder pricing in production')
print()

# Group by category
by_category = {}
for lic_type, info in used_placeholders.items():
    cat = info['category']
    if cat not in by_category:
        by_category[cat] = []
    by_category[cat].append(lic_type)

for cat in sorted(by_category.keys()):
    items = by_category[cat]
    print(f'{cat} ({len(items)} types):')
    for lic_type in sorted(items):
        info = used_placeholders[lic_type]
        print(f'  {lic_type:<35} ${info["unit_cost"]:>8,.2f} per {info["per"]}')
    print()

# Also list unused placeholders
unused = set(placeholders.keys()) - set(used_placeholders.keys())
print('=' * 70)
print(f'UNUSED PLACEHOLDERS: {len(unused)} types not found in production data')
print('(These can remain as placeholders until needed)')
print()

unused_by_cat = {}
for lic_type in unused:
    cat = placeholders[lic_type]['category']
    if cat not in unused_by_cat:
        unused_by_cat[cat] = []
    unused_by_cat[cat].append(lic_type)

for cat in sorted(unused_by_cat.keys()):
    items = unused_by_cat[cat]
    print(f'{cat}: {len(items)} types - {", ".join(sorted(items)[:5])}{"..." if len(items) > 5 else ""}')

