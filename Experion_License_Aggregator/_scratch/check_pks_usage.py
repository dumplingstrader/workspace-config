"""Check which license types have usage data displayed in PKS sheet"""
from openpyxl import load_workbook
import pandas as pd
from pathlib import Path
import sys

# Ensure output is not buffered
sys.stdout.reconfigure(line_buffering=True)

# Find most recent report
output_dir = Path('data/output')
reports = sorted(output_dir.glob('Experion_License_Report_*.xlsx'), key=lambda x: x.stat().st_mtime, reverse=True)
if not reports:
    print("No reports found")
    exit(1)

report_file = reports[0]
print(f"Checking report: {report_file.name}\n")

wb = load_workbook(report_file)
ws = wb['PKS']

# Read all data
data = [[cell.value for cell in row] for row in ws.iter_rows()]
df = pd.DataFrame(data)

# Find license types that have usage data
license_types_with_usage = []

for idx in range(len(df)):
    row = df.iloc[idx]
    if row[0] and isinstance(row[0], str):
        cell_text = str(row[0]).strip()
        
        # Check if this is a "Used" row
        if cell_text == '  - Used':
            print(f"DEBUG: Found Used row at idx {idx}")
            # Get the license type from the previous row
            if idx >= 1:
                licensed_row = df.iloc[idx-1]
                license_type = str(licensed_row[0]).strip() if licensed_row[0] else ''
                
                # Check if this Used row has any non-zero, non-None values
                has_values = False
                for val in row[1:]:
                    if pd.notna(val) and val != '' and val != 0:
                        has_values = True
                        break
                
                if has_values and license_type and not license_type.startswith('  -'):
                    license_types_with_usage.append(license_type)
                    print(f"DEBUG: Found {license_type} with usage data")
                else:
                    print(f"DEBUG: Row {idx}, cell_text='{cell_text}', has_values={has_values}, license_type='{license_type}'")

# Display results
unique_types = sorted(set(license_types_with_usage))
print(f"Found {len(unique_types)} license types with usage data displayed:\n")
for lt in unique_types:
    print(f"  ✓ {lt}")

print(f"\n{'='*60}")
print(f"Total: {len(unique_types)} license types showing usage data")
