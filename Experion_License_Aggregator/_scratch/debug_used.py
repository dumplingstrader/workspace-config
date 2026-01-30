from openpyxl import load_workbook
import pandas as pd

wb = load_workbook('data/output/Experion_License_Report_20260128_143237.xlsx')
ws = wb['PKS']

data = [[cell.value for cell in row] for row in ws.iter_rows()]
df = pd.DataFrame(data)

print("Looking for rows with 'Used' text:")
for i in range(len(df)):
    cell = df.iloc[i][0]
    if cell and isinstance(cell, str) and 'Used' in cell:
        print(f"Row {i}: '{cell}' (repr: {repr(cell)})")
        if i > 0:
            prev_cell = df.iloc[i-1][0]
            print(f"  Previous row: '{prev_cell}'")
        # Check values
        values = [v for v in df.iloc[i][1:] if pd.notna(v) and v != '' and v != 0]
        print(f"  Non-zero values: {len(values)}")
        if i < 20:
            print()
