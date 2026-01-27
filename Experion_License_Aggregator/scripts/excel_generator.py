"""
Excel Report Generator
Creates comprehensive Excel reports with multiple sheets and formatting.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict
import json


class ExcelReportGenerator:
    """Generate formatted Excel reports for license data."""
    
    # Color schemes
    COLOR_GREEN = 'C6EFCE'  # Transfer candidate
    COLOR_RED = 'FFC7CE'    # Stale license
    COLOR_ORANGE = 'FFEB9C' # Invalid customer
    COLOR_YELLOW = 'FFFF00' # Changed
    COLOR_HEADER = '4472C4'  # Blue header
    
    def __init__(self, system_names_file: str = None):
        """
        Initialize Excel generator.
        
        Args:
            system_names_file: Path to system_names.json for friendly names
        """
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self.system_names = self._load_system_names(system_names_file)
    
    def _load_system_names(self, system_names_file: str) -> Dict:
        """Load system name mappings."""
        if not system_names_file:
            return {}
        try:
            with open(system_names_file, 'r') as f:
                return json.load(f)
        except:
            return {}
    
    def _get_friendly_name(self, cluster: str, msid: str, system_number: str) -> str:
        """Get friendly name for system or return MSID."""
        key = f"{cluster}|{msid}|{system_number}"
        return self.system_names.get(key, msid)
    
    def _apply_header_style(self, ws, row: int, columns: int):
        """Apply header formatting to a row."""
        for col in range(1, columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, 
                                   end_color=self.COLOR_HEADER, 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def _auto_size_columns(self, ws):
        """Auto-size columns based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_executive_summary(self, summary_data: Dict):
        """Create Executive Summary sheet."""
        ws = self.wb.create_sheet('Executive Summary', 0)
        
        row = 1
        
        # Title
        ws['A1'] = 'Experion License Summary'
        ws['A1'].font = Font(size=16, bold=True)
        row += 2
        
        # Key metrics
        ws[f'A{row}'] = 'Generated:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        row += 1
        
        ws[f'A{row}'] = 'Total Systems:'
        ws[f'B{row}'] = summary_data.get('system_count', 0)
        row += 1
        
        ws[f'A{row}'] = 'Total Estimated Cost:'
        ws[f'B{row}'] = f"${summary_data.get('grand_total', 0):,.2f}"
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Cost by cluster
        ws[f'A{row}'] = 'Cost by Cluster'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Cluster'
        ws[f'B{row}'] = 'Total Cost'
        self._apply_header_style(ws, row, 2)
        row += 1
        
        for cluster, cost in summary_data.get('total_by_cluster', {}).items():
            ws[f'A{row}'] = cluster
            ws[f'B{row}'] = f"${cost:,.2f}"
            row += 1
        
        row += 1
        
        # Top 5 action items (placeholder)
        ws[f'A{row}'] = 'Top 5 Action Items'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        action_items = summary_data.get('action_items', [
            'Review stale licenses (>2 years old)',
            'Identify transfer candidates with excess capacity',
            'Update invalid customer names in license files',
            'Verify actual usage vs licensed capacity',
            'Update cost catalog with actual PO pricing'
        ])
        
        for i, item in enumerate(action_items[:5], 1):
            ws[f'A{row}'] = f"{i}. {item}"
            row += 1
        
        self._auto_size_columns(ws)
    
    def create_pks_sheet(self, licenses: List[Dict], changes: List[Dict] = None):
        """Create PKS systems sheet with conditional formatting."""
        ws = self.wb.create_sheet('PKS')
        
        # Headers
        headers = [
            'Cluster', 'System Name', 'MSID', 'System #', 'Product', 'Release',
            'Customer', 'License Date', 'PROCESSPOINTS', 'SCADAPOINTS',
            'STATIONS', 'MULTISTATIONS', 'DUAL', 'DAS', 'API', 'TPS',
            'Estimated Cost'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        
        # Data rows
        pks_licenses = [lic for lic in licenses if lic.get('product', 'PKS') == 'PKS']
        
        change_keys = set()
        if changes:
            change_keys = {(c['cluster'], c['msid'], c['system_number']) for c in changes}
        
        for row, lic in enumerate(pks_licenses, 2):
            cluster = lic.get('cluster', '')
            msid = lic.get('msid', '')
            sys_num = lic.get('system_number', '')
            
            ws.cell(row=row, column=1, value=cluster)
            ws.cell(row=row, column=2, value=self._get_friendly_name(cluster, msid, sys_num))
            ws.cell(row=row, column=3, value=msid)
            ws.cell(row=row, column=4, value=sys_num)
            ws.cell(row=row, column=5, value=lic.get('product', ''))
            ws.cell(row=row, column=6, value=lic.get('release', ''))
            ws.cell(row=row, column=7, value=lic.get('customer', ''))
            ws.cell(row=row, column=8, value=lic.get('license_date', ''))
            ws.cell(row=row, column=9, value=lic.get('PROCESSPOINTS', 0))
            ws.cell(row=row, column=10, value=lic.get('SCADAPOINTS', 0))
            ws.cell(row=row, column=11, value=lic.get('STATIONS', 0))
            ws.cell(row=row, column=12, value=lic.get('MULTISTATIONS', 0))
            ws.cell(row=row, column=13, value=lic.get('DUAL', 0))
            ws.cell(row=row, column=14, value=lic.get('DAS', 0))
            ws.cell(row=row, column=15, value=lic.get('API', 0))
            ws.cell(row=row, column=16, value=lic.get('TPS', 0))
            ws.cell(row=row, column=17, value=f"${lic.get('_estimated_cost', 0):,.2f}")
            
            # Apply conditional formatting
            self._apply_conditional_formatting(ws, row, lic, change_keys, len(headers))
        
        self._auto_size_columns(ws)
    
    def _apply_conditional_formatting(self, ws, row: int, lic: Dict, 
                                     change_keys: set, num_cols: int):
        """Apply color coding based on conditions."""
        cluster = lic.get('cluster', '')
        msid = lic.get('msid', '')
        sys_num = lic.get('system_number', '')
        
        fill = None
        
        # Check for changes (yellow)
        if (cluster, msid, sys_num) in change_keys:
            fill = PatternFill(start_color=self.COLOR_YELLOW, 
                             end_color=self.COLOR_YELLOW, fill_type='solid')
        
        # Check for stale license (red) - >2 years
        license_date = lic.get('license_date', '')
        if license_date:
            try:
                lic_dt = datetime.strptime(license_date, '%Y-%m-%d')
                days_old = (datetime.now() - lic_dt).days
                if days_old > 730:  # 2 years
                    fill = PatternFill(start_color=self.COLOR_RED,
                                     end_color=self.COLOR_RED, fill_type='solid')
            except:
                pass
        
        # Check for invalid customer (orange)
        customer = lic.get('customer', '').upper()
        if customer and not any(valid in customer for valid in ['MARATHON', 'MPC']):
            fill = PatternFill(start_color=self.COLOR_ORANGE,
                             end_color=self.COLOR_ORANGE, fill_type='solid')
        
        # Check for transfer candidate (green) - >500 excess process points
        process_points = lic.get('PROCESSPOINTS', 0)
        if process_points > 5000:  # Arbitrary threshold without actual usage
            fill = PatternFill(start_color=self.COLOR_GREEN,
                             end_color=self.COLOR_GREEN, fill_type='solid')
        
        # Apply fill to entire row
        if fill:
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).fill = fill
    
    def create_summary_sheet(self, summary_data: Dict):
        """Create Summary sheet with totals by cluster."""
        ws = self.wb.create_sheet('Summary')
        
        # Headers
        ws['A1'] = 'Cluster'
        ws['B1'] = 'System Count'
        ws['C1'] = 'Total Cost'
        self._apply_header_style(ws, 1, 3)
        
        # Data
        row = 2
        cluster_counts = {}
        for sys_cost in summary_data.get('system_costs', []):
            cluster = sys_cost.get('system', 'Unknown').split(' - ')[0]
            cluster_counts[cluster] = cluster_counts.get(cluster, 0) + 1
        
        for cluster, cost in summary_data.get('total_by_cluster', {}).items():
            ws[f'A{row}'] = cluster
            ws[f'B{row}'] = cluster_counts.get(cluster, 0)
            ws[f'C{row}'] = f"${cost:,.2f}"
            row += 1
        
        # Total row
        row += 1
        ws[f'A{row}'] = 'TOTAL'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = summary_data.get('system_count', 0)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f"${summary_data.get('grand_total', 0):,.2f}"
        ws[f'C{row}'].font = Font(bold=True)
        
        self._auto_size_columns(ws)
    
    def create_changes_sheet(self, changes: List[Dict]):
        """Create Changes sheet showing differences from previous run."""
        ws = self.wb.create_sheet('Changes')
        
        if not changes:
            ws['A1'] = 'No changes detected from previous run'
            return
        
        # Headers
        headers = ['Cluster', 'MSID', 'System #', 'Field', 'Previous', 'Current', 'Change']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        
        # Data
        for row, change in enumerate(changes, 2):
            ws.cell(row=row, column=1, value=change['cluster'])
            ws.cell(row=row, column=2, value=change['msid'])
            ws.cell(row=row, column=3, value=change['system_number'])
            ws.cell(row=row, column=4, value=change['field'])
            ws.cell(row=row, column=5, value=change.get('old_value', ''))
            ws.cell(row=row, column=6, value=change.get('new_value', ''))
            ws.cell(row=row, column=7, value=change.get('delta', ''))
        
        self._auto_size_columns(ws)
    
    def create_errors_sheet(self, errors: List[Dict]):
        """Create Errors sheet for files that couldn't be parsed."""
        ws = self.wb.create_sheet('Errors')
        
        if not errors:
            ws['A1'] = 'No errors - all files parsed successfully'
            return
        
        # Headers
        ws['A1'] = 'File'
        ws['B1'] = 'Error'
        self._apply_header_style(ws, 1, 2)
        
        # Data
        for row, error in enumerate(errors, 2):
            ws.cell(row=row, column=1, value=error['file'])
            ws.cell(row=row, column=2, value=error['error'])
        
        self._auto_size_columns(ws)
    
    def create_transfer_candidates_sheet(self, candidates: List[Dict]):
        """Create Transfer Candidates sheet."""
        ws = self.wb.create_sheet('Transfer Candidates')
        
        if not candidates:
            ws['A1'] = 'No transfer candidates identified'
            ws['A2'] = '(Systems with >500 excess points or >25% excess capacity)'
            return
        
        # Headers
        headers = ['Cluster', 'MSID', 'System #', 'Licensed Points', 
                  'Excess Points', 'Excess %', 'Cost Value']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        
        # Data
        for row, candidate in enumerate(candidates, 2):
            ws.cell(row=row, column=1, value=candidate['cluster'])
            ws.cell(row=row, column=2, value=candidate['msid'])
            ws.cell(row=row, column=3, value=candidate['system_number'])
            ws.cell(row=row, column=4, value=candidate['licensed_process_points'])
            ws.cell(row=row, column=5, value=candidate['excess_points'])
            ws.cell(row=row, column=6, value=f"{candidate['excess_percent']:.1f}%")
            ws.cell(row=row, column=7, value=f"${candidate['cost_value']:,.2f}")
            
            # Green fill for transfer candidates
            fill = PatternFill(start_color=self.COLOR_GREEN,
                             end_color=self.COLOR_GREEN, fill_type='solid')
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill
        
        self._auto_size_columns(ws)
    
    def save(self, output_path: str):
        """Save workbook to file."""
        self.wb.save(output_path)
        print(f"Excel report saved: {output_path}")


if __name__ == '__main__':
    # Test Excel generation
    generator = ExcelReportGenerator()
    
    summary_data = {
        'system_count': 3,
        'grand_total': 150000,
        'total_by_cluster': {'Carson': 100000, 'Wilmington': 50000},
        'system_costs': []
    }
    
    generator.create_executive_summary(summary_data)
    generator.create_summary_sheet(summary_data)
    generator.save('test_report.xlsx')
