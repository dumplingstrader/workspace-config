"""
Historical License Analysis Script
Analyzes license changes over time using version numbers as timestamps
"""
from pathlib import Path
from collections import defaultdict
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from v2.pipeline.coordinator import PipelineCoordinator
from v2.pipeline.extractors.xml_extractor import XmlExtractor
from v2.core.config import Config


class HistoricalAnalyzer:
    """Analyze license history by processing all versions."""
    
    def __init__(self, xml_dir: Path, output_dir: Path, usage_dir: Path = None):
        self.xml_dir = xml_dir
        self.output_dir = output_dir
        self.usage_dir = usage_dir or Path('data/raw/Usage')
        self.config = Config(config_dir=Path('v2/config'))
        self.extractor = XmlExtractor(self.config)
        
    def extract_all_versions(self):
        """Extract all versions without deduplication."""
        print("\n" + "=" * 60)
        print("EXTRACTING ALL LICENSE VERSIONS")
        print("=" * 60)
        
        all_versions = []
        xml_files = list(self.xml_dir.rglob("*.xml"))
        
        print(f"Found {len(xml_files)} XML files")
        
        for xml_file in xml_files:
            try:
                result = self.extractor.extract_from_file(xml_file)
                if not result.success:
                    print(f"  ⚠️  Skipped {xml_file.name}: {result.validation_errors}")
                    continue
                    
                license_data = result.data
                all_versions.append({
                    'file': xml_file.name,
                    'path': str(xml_file.parent),
                    'msid': license_data.msid,
                    'system_number': license_data.system_number,
                    'cluster': license_data.cluster,
                    'release': license_data.release,
                    'version': license_data.file_version or 0,
                    'license_date': license_data.license_date,
                    'licensed': license_data.licensed,
                    'license_obj': license_data
                })
            except Exception as e:
                print(f"  [WARN] Skipped {xml_file.name}: {e}")
                
        print(f"\n[OK] Extracted {len(all_versions)} licenses")
        return all_versions
    
    def extract_current_versions_from_usage(self):
        """Extract current version from Usage CSV files."""
        import csv
        
        current_versions = {}
        
        if not self.usage_dir.exists():
            print(f"[WARN] Usage directory not found: {self.usage_dir}")
            return current_versions
        
        for csv_file in self.usage_dir.glob('*.csv'):
            try:
                with open(csv_file, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    rows = list(reader)
                    
                    # Look for row with "License version"
                    # Format: License certificate,License version,License,<VERSION>
                    msid = None
                    version = None
                    system_name = csv_file.stem  # Filename without .csv
                    
                    for row in rows:
                        if len(row) >= 4:
                            if row[1] == 'MSID/ESID':
                                msid = row[3]
                            elif row[1] == 'License version':
                                try:
                                    version = int(row[3])
                                except ValueError:
                                    pass
                    
                    if msid and version is not None:
                        current_versions[msid] = {
                            'version': version,
                            'system_name': system_name,
                            'csv_file': csv_file.name
                        }
            except Exception as e:
                print(f"[WARN] Error reading {csv_file.name}: {e}")
        
        return current_versions
    
    def group_by_system_and_version(self, all_versions):
        """Group versions by system (MSID) and sort by version number."""
        systems = defaultdict(list)
        
        for version in all_versions:
            key = version['msid']
            systems[key].append(version)
        
        # Sort each system's versions
        for msid in systems:
            systems[msid].sort(key=lambda x: x['version'])
            
        return systems
    
    def analyze_changes(self, systems, current_versions=None):
        """Analyze changes between versions for each system."""
        print("\n" + "=" * 60)
        print("ANALYZING VERSION CHANGES")
        print("=" * 60)
        
        if current_versions is None:
            current_versions = {}
        
        changes = []
        
        for msid, versions in systems.items():
            if len(versions) < 2:
                current_info = current_versions.get(msid)
                if current_info:
                    print(f"  {msid}: Only 1 version (no history) - Current: v{current_info['version']}")
                else:
                    print(f"  {msid}: Only 1 version (no history)")
                continue
                
            current_info = current_versions.get(msid)
            if current_info:
                print(f"\n  {msid}: {len(versions)} versions (v{versions[0]['version']} → v{versions[-1]['version']}) - Current: v{current_info['version']}")
            else:
                print(f"\n  {msid}: {len(versions)} versions (v{versions[0]['version']} → v{versions[-1]['version']})")
            
            # Compare each version to previous
            for i in range(1, len(versions)):
                prev = versions[i-1]
                curr = versions[i]
                
                # Analyze license type changes
                for lic_type in set(list(prev['licensed'].keys()) + list(curr['licensed'].keys())):
                    prev_qty = prev['licensed'].get(lic_type, 0)
                    curr_qty = curr['licensed'].get(lic_type, 0)
                    
                    if prev_qty != curr_qty:
                        current_info = current_versions.get(msid)
                        change = {
                            'msid': msid,
                            'cluster': curr['cluster'],
                            'from_version': prev['version'],
                            'to_version': curr['version'],
                            'current_version': current_info['version'] if current_info else None,
                            'license_type': lic_type,
                            'prev_quantity': prev_qty,
                            'curr_quantity': curr_qty,
                            'change': curr_qty - prev_qty,
                            'change_pct': ((curr_qty - prev_qty) / prev_qty * 100) if prev_qty > 0 else 100,
                            'change_type': 'Addition' if curr_qty > prev_qty else 'Reduction'
                        }
                        changes.append(change)
                        
                        # Print significant changes
                        if abs(change['change']) >= 100 or lic_type in ['PROCESSPOINTS', 'SCADAPOINTS', 'DIRECTSTATIONS']:
                            print(f"    v{prev['version']} → v{curr['version']}: {lic_type} {prev_qty} → {curr_qty} ({change['change']:+d})")
        
        print(f"\n✓ Found {len(changes)} license changes across all systems")
        return changes
    
    def generate_historical_report(self, systems, changes):
        """Generate Excel report with historical analysis."""
        print("\n" + "=" * 60)
        print("GENERATING HISTORICAL REPORT")
        print("=" * 60)
        
        output_file = self.output_dir / f"License_History_Report_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Sheet 1: All Versions Summary
        ws_versions = wb.create_sheet("All Versions")
        version_data = []
        for msid, versions in sorted(systems.items()):
            for v in versions:
                version_data.append({
                    'MSID': v['msid'],
                    'Cluster': v['cluster'],
                    'System Number': v['system_number'],
                    'Version': v['version'],
                    'Release': v['release'],
                    'File': v['file'],
                    'Path': v['path'],
                    'Total License Types': len(v['licensed'])
                })
        
        df_versions = pd.DataFrame(version_data)
        for row in dataframe_to_rows(df_versions, index=False, header=True):
            ws_versions.append(row)
        
        # Format headers
        for cell in ws_versions[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Sheet 2: Change History
        ws_changes = wb.create_sheet("Change History")
        if changes:
            df_changes = pd.DataFrame(changes)
            # Reorder columns to include current_version
            column_order = ['msid', 'cluster', 'from_version', 'to_version', 'current_version', 
                          'license_type', 'prev_quantity', 'curr_quantity', 'change', 'change_pct', 'change_type']
            df_changes = df_changes[[col for col in column_order if col in df_changes.columns]]
            for row in dataframe_to_rows(df_changes, index=False, header=True):
                ws_changes.append(row)
            
            # Format headers
            for cell in ws_changes[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Highlight additions (green) and reductions (red)
            for row_idx, row in enumerate(ws_changes.iter_rows(min_row=2), start=2):
                change_type_cell = row[9]  # change_type column
                if change_type_cell.value == 'Addition':
                    for cell in row:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif change_type_cell.value == 'Reduction':
                    for cell in row:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Sheet 3: System History Timeline
        ws_timeline = wb.create_sheet("System Timeline")
        timeline_data = []
        for msid, versions in sorted(systems.items()):
            if len(versions) > 1:
                timeline_data.append({
                    'MSID': msid,
                    'Cluster': versions[0]['cluster'],
                    'First Version': versions[0]['version'],
                    'Latest Version': versions[-1]['version'],
                    'Total Versions': len(versions),
                    'Version Range': f"{versions[0]['version']}-{versions[-1]['version']}"
                })
        
        df_timeline = pd.DataFrame(timeline_data)
        for row in dataframe_to_rows(df_timeline, index=False, header=True):
            ws_timeline.append(row)
        
        # Format headers
        for cell in ws_timeline[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Sheet 4: Summary Statistics
        ws_summary = wb.create_sheet("Summary", 0)  # Make it first sheet
        
        summary_stats = [
            ['Historical License Analysis Report', ''],
            ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['', ''],
            ['Total Systems Analyzed', len(systems)],
            ['Systems with History (>1 version)', len([s for s in systems.values() if len(s) > 1])],
            ['Total License Versions', sum(len(v) for v in systems.values())],
            ['Total License Changes', len(changes)],
            ['', ''],
            ['Change Breakdown', ''],
            ['Additions', len([c for c in changes if c['change'] > 0])],
            ['Reductions', len([c for c in changes if c['change'] < 0])],
            ['', ''],
            ['Top Systems by Version Count', ''],
        ]
        
        # Add top systems
        top_systems = sorted(systems.items(), key=lambda x: len(x[1]), reverse=True)[:10]
        for msid, versions in top_systems:
            summary_stats.append([f"  {msid}", f"{len(versions)} versions"])
        
        for row in summary_stats:
            ws_summary.append(row)
        
        # Format summary sheet
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 25
        
        # Bold headers
        for row_idx in [1, 4, 9, 13]:
            ws_summary[f'A{row_idx}'].font = Font(bold=True, size=12)
        
        # Save workbook
        wb.save(output_file)
        print(f"\n✓ Report saved: {output_file}")
        
        return output_file
    
    def run_analysis(self):
        """Run complete historical analysis."""
        print("\n" + "=" * 70)
        print(" " * 15 + "HISTORICAL LICENSE ANALYSIS")
        print("=" * 70)
        
        # Extract all versions
        all_versions = self.extract_all_versions()
        
        if not all_versions:
            print("\n❌ No licenses found!")
            return None
        
        # Group by system
        systems = self.group_by_system_and_version(all_versions)
        
        # Analyze changes
        changes = self.analyze_changes(systems)
        
        # Generate report
        report_file = self.generate_historical_report(systems, changes)
        
        # Print summary
        print("\n" + "=" * 70)
        print("ANALYSIS COMPLETE")
        print("=" * 70)
        print(f"\n📊 Systems analyzed: {len(systems)}")
        print(f"📊 Total versions: {len(all_versions)}")
        print(f"📊 License changes: {len(changes)}")
        
        systems_with_history = len([s for s in systems.values() if len(s) > 1])
        print(f"📊 Systems with history: {systems_with_history}")
        
        if changes:
            additions = len([c for c in changes if c['change'] > 0])
            reductions = len([c for c in changes if c['change'] < 0])
            print(f"\n📈 Additions: {additions}")
            print(f"📉 Reductions: {reductions}")
            
            # Show biggest changes
            print("\n🔝 Top 5 Largest Changes:")
            top_changes = sorted(changes, key=lambda x: abs(x['change']), reverse=True)[:5]
            for i, change in enumerate(top_changes, 1):
                print(f"  {i}. {change['msid']} - {change['license_type']}: "
                      f"{change['prev_quantity']} → {change['curr_quantity']} "
                      f"({change['change']:+d}) [v{change['from_version']}→v{change['to_version']}]")
        
        print(f"\n📄 Report: {report_file}")
        print("=" * 70)
        
        return report_file


def main():
    """Main entry point."""
    analyzer = HistoricalAnalyzer(
        xml_dir=Path('data/raw/Carson'),
        output_dir=Path('data/output'),
        usage_dir=Path('data/raw/Usage')
    )
    
    # Extract current versions from Usage CSVs
    print("\n" + "=" * 70)
    print("READING CURRENT VERSIONS FROM USAGE DATA")
    print("=" * 70)
    current_versions = analyzer.extract_current_versions_from_usage()
    print(f"[OK] Found current versions for {len(current_versions)} systems")
    for msid, info in current_versions.items():
        print(f"  {msid}: v{info['version']} (from {info['csv_file']})")
    
    # Extract all versions
    all_versions = analyzer.extract_all_versions()
    
    # Group by system
    systems = analyzer.group_by_system_and_version(all_versions)
    
    # Analyze changes with current version info
    changes = analyzer.analyze_changes(systems, current_versions)
    
    # Generate report
    report_file = analyzer.generate_historical_report(systems, changes)
    
    # Print summary
    print("\n" + "=" * 70)
    print("ANALYSIS COMPLETE")
    print("=" * 70)
    print(f"\n📊 Systems analyzed: {len(systems)}")
    print(f"📊 Total versions: {len(all_versions)}")
    print(f"📊 License changes: {len(changes)}")
    
    systems_with_history = len([s for s in systems.values() if len(s) > 1])
    print(f"📊 Systems with history: {systems_with_history}")
    
    if changes:
        additions = len([c for c in changes if c['change'] > 0])
        reductions = len([c for c in changes if c['change'] < 0])
        print(f"\n📈 Additions: {additions}")
        print(f"📉 Reductions: {reductions}")
        
        # Show biggest changes
        print("\n🔝 Top 5 Largest Changes:")
        top_changes = sorted(changes, key=lambda x: abs(x['change']), reverse=True)[:5]
        for i, change in enumerate(top_changes, 1):
            print(f"  {i}. {change['msid']} - {change['license_type']}: "
                  f"{change['prev_quantity']} → {change['curr_quantity']} "
                  f"({change['change']:+d}) [v{change['from_version']}→v{change['to_version']}]")
    
    print(f"\n📄 Report: {report_file}")
    print("=" * 70)
    
    return report_file


if __name__ == '__main__':
    main()
