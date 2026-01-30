"""
Excel Report Generator
Creates comprehensive Excel reports with multiple sheets and formatting.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict
import json


class ExcelReportGenerator:
    """Generate formatted Excel reports for license data."""
    
    # Color schemes
    COLOR_GREEN = 'C6EFCE'  # Transfer candidate
    COLOR_RED = 'FFC7CE'    # Stale license
    COLOR_ORANGE = 'FFEB9C' # Invalid customer
    COLOR_YELLOW = 'FFFF00' # Changed
    COLOR_HEADER = '4472C4'  # Blue header
    
    def __init__(self, system_names_file: str = None):
        """
        Initialize Excel generator.
        
        Args:
            system_names_file: Path to system_names.json for friendly names
        """
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        self.system_names = self._load_system_names(system_names_file)
    
    def _load_system_names(self, system_names_file: str) -> Dict:
        """Load system name mappings."""
        if not system_names_file:
            return {}
        try:
            with open(system_names_file, 'r') as f:
                return json.load(f)
        except:
            return {}
    
    def _get_friendly_name(self, cluster: str, msid: str, system_number: str) -> str:
        """Get friendly name for system or return MSID."""
        key = f"{cluster}|{msid}|{system_number}"
        return self.system_names.get(key, msid)
    
    def _apply_header_style(self, ws, row: int, columns: int):
        """Apply header formatting to a row."""
        for col in range(1, columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color='FFFFFF', size=11)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, 
                                   end_color=self.COLOR_HEADER, 
                                   fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    def _auto_size_columns(self, ws):
        """Auto-size columns based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_executive_summary(self, summary_data: Dict):
        """Create Executive Summary sheet."""
        ws = self.wb.create_sheet('Executive Summary', 0)
        
        row = 1
        
        # Title
        ws['A1'] = 'Experion License Summary'
        ws['A1'].font = Font(size=16, bold=True)
        row += 2
        
        # Key metrics
        ws[f'A{row}'] = 'Generated:'
        ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        row += 1
        
        ws[f'A{row}'] = 'Total Systems:'
        ws[f'B{row}'] = summary_data.get('system_count', 0)
        row += 1
        
        ws[f'A{row}'] = 'Total Estimated Cost:'
        ws[f'B{row}'] = f"${summary_data.get('grand_total', 0):,.2f}"
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Cost by cluster
        ws[f'A{row}'] = 'Cost by Cluster'
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = 'Cluster'
        ws[f'B{row}'] = 'Total Cost'
        self._apply_header_style(ws, row, 2)
        row += 1
        
        for cluster, cost in summary_data.get('total_by_cluster', {}).items():
            ws[f'A{row}'] = cluster
            ws[f'B{row}'] = f"${cost:,.2f}"
            row += 1
        
        row += 1
        
        # Top 5 action items (placeholder)
        ws[f'A{row}'] = 'Top 5 Action Items'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        action_items = summary_data.get('action_items', [
            'Review stale licenses (>2 years old)',
            'Identify transfer candidates with excess capacity',
            'Update invalid customer names in license files',
            'Verify actual usage vs licensed capacity',
            'Update cost catalog with actual PO pricing'
        ])
        
        for i, item in enumerate(action_items[:5], 1):
            ws[f'A{row}'] = f"{i}. {item}"
            row += 1
        
        self._auto_size_columns(ws)
    
    def create_pks_sheet(self, licenses: List[Dict], changes: List[Dict] = None):
        """Create PKS systems sheet in pivot table format (systems as columns)."""
        ws = self.wb.create_sheet('PKS')
        
        # Group licenses by cluster
        pks_licenses = [lic for lic in licenses if lic.get('product', 'PKS') == 'PKS']
        pks_licenses.sort(key=lambda x: (x.get('cluster', ''), x.get('msid', ''), x.get('system_number', '')))
        
        # Build column headers (systems)
        systems = []
        for lic in pks_licenses:
            cluster_abbr = lic.get('cluster', '')[:3].upper()
            system_name = self._get_friendly_name(lic.get('cluster'), lic.get('msid'), lic.get('system_number'))
            systems.append({
                'header': f"{system_name} ({cluster_abbr})",
                'license': lic
            })
        
        # Header row - Column A is "Field", then system names
        ws.cell(1, 1, 'Field')
        for col_idx, sys in enumerate(systems, 2):
            ws.cell(1, col_idx, sys['header'])
        self._apply_header_style(ws, 1, len(systems) + 1)
        
        # Define rows in order
        rows = [
            ('LICENSE INFO', None, True),
            ('Cluster', 'cluster', False),
            ('MSID', 'msid', False),
            ('System Number', 'system_number', False),
            ('Release', 'release', False),
            ('License Version', 'version', False),
            ('Generated', 'license_date', False),
            ('Customer', 'customer', False),
            ('', None, True),  # Blank row
            ('POINTS', None, True),
            ('Process Points', 'PROCESSPOINTS', False),
            ('  - Used', 'PROCESSPOINTS_USED', False),
            ('  - Utilization %', None, False),  # Calculated
            ('SCADA Points', 'SCADAPOINTS', False),
            ('  - Used', 'SCADAPOINTS_USED', False),
            ('  - Utilization %', None, False),  # Calculated
            ('CDA Analog I/O', 'CDA_IO_ANA', False),
            ('  - Used', 'CDA_IO_ANA_USED', False),
            ('CDA Digital I/O', 'CDA_IO_DIG', False),
            ('  - Used', 'CDA_IO_DIG_USED', False),
            ('', None, True),  # Blank row
            ('STATIONS', None, True),
            ('Flex Stations', 'STATIONS', False),
            ('  - Used', 'STATIONS_USED', False),
            ('  - Utilization %', None, False),  # Calculated
            ('Multi-Window Stations', 'MULTISTATIONS', False),
            ('  - Used', 'MULTISTATIONS_USED', False),
            ('  - Utilization %', None, False),  # Calculated
            ('Console/Direct Stations', 'DIRECTSTATIONS', False),
            ('  - Used', 'CONSOLE_STATIONS_USED', False),  # CSV uses CONSOLE_STATIONS
            ('  - Utilization %', None, False),  # Calculated
            ('Console Extensions', 'CONSOLE_EXTENSION', False),  # Only in CSV, not XML
            ('  - Used', 'CONSOLE_EXTENSION_USED', False),
            ('Operator Touch Panels', 'OPER_TOUCH_PANEL', False),
            ('  - Used', 'OPERATOR_TOUCH_PANELS_USED', False),  # CSV uses OPERATOR_TOUCH_PANELS
            ('Read-Only Stations', 'READONLY_STATIONS', False),
            ('  - Used', 'READONLY_STATIONS_USED', False),
            ('', None, True),  # Blank row
            ('REDUNDANCY', None, True),
            ('Server Redundancy', 'DUAL', False),
            ('Distributed Servers', 'MULTI_SERVER', False),
            ('', None, True),  # Blank row
            ('FEATURES', None, True),
            ('Data Acquisition Service', 'DAS', False),
            ('API Access', 'API', False),
            ('SQL Database Access', 'SQL', False),
            ('License Admin Service', 'LAS', False),
            ('Display Builder', 'DSPBLD', False),
            ('', None, True),  # Blank row
            ('INTERFACES', None, True),
            ('CDA Subsystems', 'CDA', False),
            ('TPS Enabler', 'TPS', False),
            ('Safety Manager/FSC', 'FSC', False),
            ('Modbus', 'MODICON', False),
            ('Allen Bradley', 'AB', False),
            ('Allen Bradley Ethernet', 'AB_ETH', False),
            ('DNP3', 'DNP3', False),
            ('OPC DA Servers', 'OPC_DA', False),
            ('OPC UA Client', 'OPC_UA_CLIENT', False),
            ('', None, True),  # Blank row
            ('VIRTUALIZATION', None, True),
            ('Server CALs', 'VIRTUALIZATION', False),
            ('Client CALs', 'VIRTUALIZATION_CLIENT', False),
            ('', None, True),  # Blank row
            ('COST ESTIMATE', None, True),
            ('Total Cost', '_estimated_cost', False),
        ]
        
        # Fill in data rows
        for row_idx, (label, field, is_header) in enumerate(rows, 2):
            ws.cell(row_idx, 1, label)
            
            if is_header:
                # Format as section header
                ws.cell(row_idx, 1).font = Font(bold=True, size=11)
                ws.cell(row_idx, 1).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            elif field or label.endswith('Utilization %'):
                # Data row - fill in values for each system
                for col_idx, sys in enumerate(systems, 2):
                    lic = sys['license']
                    
                    # Calculate utilization percentage rows
                    if label.endswith('Utilization %'):
                        # Find current position in rows array
                        current_idx = None
                        for i, r in enumerate(rows):
                            if r[0] == label and r[2] == is_header:  # Match label and is_header
                                if current_idx is None or abs(i - (row_idx - 2)) < abs(current_idx - (row_idx - 2)):
                                    current_idx = i
                        
                        if current_idx is not None and current_idx >= 2:
                            licensed_idx = current_idx - 2
                            used_idx = current_idx - 1
                            
                            licensed_field = rows[licensed_idx][1]
                            used_field = rows[used_idx][1]
                            
                            if licensed_field and used_field:
                                licensed = lic.get(licensed_field, 0)
                                used = lic.get(used_field, 0)
                                
                                try:
                                    if licensed and int(licensed) > 0:
                                        util_pct = (int(used) / int(licensed)) * 100
                                        ws.cell(row_idx, col_idx, f"{util_pct:.1f}%")
                                        
                                        # Color code based on utilization
                                        if util_pct >= 80:
                                            ws.cell(row_idx, col_idx).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Red
                                        elif util_pct >= 50:
                                            ws.cell(row_idx, col_idx).fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Yellow
                                        else:
                                            ws.cell(row_idx, col_idx).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
                                    else:
                                        ws.cell(row_idx, col_idx, "")
                                except (ValueError, ZeroDivisionError):
                                    ws.cell(row_idx, col_idx, "")
                    elif field:
                        value = lic.get(field, 0 if field.isupper() else '')
                        
                        # Format cost
                        if field == '_estimated_cost':
                            ws.cell(row_idx, col_idx, f"${value:,.2f}" if value else "$0.00")
                        # Format boolean fields
                        elif field in ['DUAL', 'MULTI_SERVER', 'DAS', 'API', 'SQL', 'LAS', 'DSPBLD',
                                     'CDA', 'TPS', 'FSC', 'MODICON', 'AB', 'AB_ETH', 'DNP3', 'OPC_DA', 'OPC_UA_CLIENT']:
                            ws.cell(row_idx, col_idx, '✓' if value and int(value) > 0 else '')
                        else:
                            ws.cell(row_idx, col_idx, value if value else '')
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        for col_idx in range(2, len(systems) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Freeze panes (freeze first column and first row)
        ws.freeze_panes = 'B2'
    
    def _apply_conditional_formatting(self, ws, row: int, lic: Dict, 
                                     change_keys: set, num_cols: int):
        """Apply color coding based on conditions."""
        cluster = lic.get('cluster', '')
        msid = lic.get('msid', '')
        sys_num = lic.get('system_number', '')
        
        fill = None
        
        # Check for changes (yellow)
        if (cluster, msid, sys_num) in change_keys:
            fill = PatternFill(start_color=self.COLOR_YELLOW, 
                             end_color=self.COLOR_YELLOW, fill_type='solid')
        
        # Check for stale license (red) - >2 years
        license_date = lic.get('license_date', '')
        if license_date:
            try:
                lic_dt = datetime.strptime(license_date, '%Y-%m-%d')
                days_old = (datetime.now() - lic_dt).days
                if days_old > 730:  # 2 years
                    fill = PatternFill(start_color=self.COLOR_RED,
                                     end_color=self.COLOR_RED, fill_type='solid')
            except:
                pass
        
        # Check for invalid customer (orange)
        customer = lic.get('customer', '').upper()
        if customer and not any(valid in customer for valid in ['MARATHON', 'MPC']):
            fill = PatternFill(start_color=self.COLOR_ORANGE,
                             end_color=self.COLOR_ORANGE, fill_type='solid')
        
        # Check for transfer candidate (green) - >500 excess process points
        process_points = lic.get('PROCESSPOINTS', 0)
        if process_points > 5000:  # Arbitrary threshold without actual usage
            fill = PatternFill(start_color=self.COLOR_GREEN,
                             end_color=self.COLOR_GREEN, fill_type='solid')
        
        # Apply fill to entire row
        if fill:
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).fill = fill
    
    def create_summary_sheet(self, summary_data: Dict):
        """Create Summary sheet with totals by cluster."""
        ws = self.wb.create_sheet('Summary')
        
        # Headers
        ws['A1'] = 'Cluster'
        ws['B1'] = 'System Count'
        ws['C1'] = 'Total Cost'
        self._apply_header_style(ws, 1, 3)
        
        # Data
        row = 2
        cluster_counts = {}
        for sys_cost in summary_data.get('system_costs', []):
            cluster = sys_cost.get('system', 'Unknown').split(' - ')[0]
            cluster_counts[cluster] = cluster_counts.get(cluster, 0) + 1
        
        for cluster, cost in summary_data.get('total_by_cluster', {}).items():
            ws[f'A{row}'] = cluster
            ws[f'B{row}'] = cluster_counts.get(cluster, 0)
            ws[f'C{row}'] = f"${cost:,.2f}"
            row += 1
        
        # Total row
        row += 1
        ws[f'A{row}'] = 'TOTAL'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = summary_data.get('system_count', 0)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'C{row}'] = f"${summary_data.get('grand_total', 0):,.2f}"
        ws[f'C{row}'].font = Font(bold=True)
        
        self._auto_size_columns(ws)
    
    def create_changes_sheet(self, changes: List[Dict]):
        """Create Changes sheet showing differences from previous run."""
        ws = self.wb.create_sheet('Changes')
        
        if not changes:
            ws['A1'] = 'No changes detected from previous run'
            return
        
        # Headers
        headers = ['Cluster', 'MSID', 'System #', 'Field', 'Previous', 'Current', 'Change']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        
        # Data
        for row, change in enumerate(changes, 2):
            ws.cell(row=row, column=1, value=change['cluster'])
            ws.cell(row=row, column=2, value=change['msid'])
            ws.cell(row=row, column=3, value=change['system_number'])
            ws.cell(row=row, column=4, value=change['field'])
            ws.cell(row=row, column=5, value=change.get('old_value', ''))
            ws.cell(row=row, column=6, value=change.get('new_value', ''))
            ws.cell(row=row, column=7, value=change.get('delta', ''))
        
        self._auto_size_columns(ws)
    
    def create_errors_sheet(self, errors: List[Dict]):
        """Create Errors sheet for files that couldn't be parsed."""
        ws = self.wb.create_sheet('Errors')
        
        if not errors:
            ws['A1'] = 'No errors - all files parsed successfully'
            return
        
        # Headers
        ws['A1'] = 'File'
        ws['B1'] = 'Error'
        self._apply_header_style(ws, 1, 2)
        
        # Data
        for row, error in enumerate(errors, 2):
            ws.cell(row=row, column=1, value=error['file'])
            ws.cell(row=row, column=2, value=error['error'])
        
        self._auto_size_columns(ws)
    
    def create_transfer_candidates_sheet(self, candidates: List[Dict]):
        """Create Transfer Candidates sheet."""
        ws = self.wb.create_sheet('Transfer Candidates')
        
        if not candidates:
            ws['A1'] = 'No transfer candidates identified'
            ws['A2'] = '(Systems with >500 excess points or >25% excess capacity)'
            return
        
        # Headers
        headers = [
            'Cluster', 'System Name', 'MSID', 'System #',
            'License Type', 'Licensed', 'Used', 'Excess', 'Utilization %',
            'Excess Value', 'Has Usage Data'
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        self._apply_header_style(ws, 1, len(headers))
        
        # Data rows - one row per excess item
        row = 2
        for candidate in candidates:
            cluster = candidate['cluster']
            msid = candidate['msid']
            sys_num = candidate['system_number']
            system_name = self._get_friendly_name(cluster, msid, sys_num)
            has_util_data = candidate.get('has_utilization_data', False)
            
            for excess_item in candidate.get('excess_items', []):
                ws.cell(row=row, column=1, value=cluster)
                ws.cell(row=row, column=2, value=system_name)
                ws.cell(row=row, column=3, value=msid)
                ws.cell(row=row, column=4, value=sys_num)
                ws.cell(row=row, column=5, value=excess_item['type'])
                ws.cell(row=row, column=6, value=excess_item['licensed'])
                ws.cell(row=row, column=7, value=excess_item['used'])
                ws.cell(row=row, column=8, value=excess_item['excess'])
                ws.cell(row=row, column=9, value=f"{excess_item['utilization_percent']:.1f}%")
                ws.cell(row=row, column=10, value=f"${excess_item['excess_value']:,.2f}")
                ws.cell(row=row, column=11, value='Yes' if has_util_data else 'No')
                
                # Green fill for transfer candidates
                fill = PatternFill(start_color=self.COLOR_GREEN,
                                 end_color=self.COLOR_GREEN, fill_type='solid')
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = fill
                
                row += 1
        
        self._auto_size_columns(ws)
    
    def save(self, output_path: str):
        """Save workbook to file."""
        self.wb.save(output_path)
        print(f"Excel report saved: {output_path}")


if __name__ == '__main__':
    # Test Excel generation
    generator = ExcelReportGenerator()
    
    summary_data = {
        'system_count': 3,
        'grand_total': 150000,
        'total_by_cluster': {'Carson': 100000, 'Wilmington': 50000},
        'system_costs': []
    }
    
    generator.create_executive_summary(summary_data)
    generator.create_summary_sheet(summary_data)
    generator.save('test_report.xlsx')
