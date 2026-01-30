"""Summarize which license types have usage data in PKS sheet"""
from openpyxl import load_workbook
import pandas as pd

wb = load_workbook('data/output/Experion_License_Report_20260128_143237.xlsx')
ws = wb['PKS']

data = [[cell.value for cell in row] for row in ws.iter_rows()]
df = pd.DataFrame(data)

license_types_with_usage = []
license_types_without_usage = []

for i in range(len(df)):
    cell = df.iloc[i][0]
    if cell and isinstance(cell, str) and cell.strip() == '  - Used':
        if i > 0:
            prev_cell = df.iloc[i-1][0]
            if prev_cell and isinstance(prev_cell, str) and not prev_cell.startswith('  -'):
                license_type = prev_cell.strip()
                # Check if there are non-zero values
                values = [v for v in df.iloc[i][1:] if pd.notna(v) and v != '' and v != 0]
                if len(values) > 0:
                    license_types_with_usage.append(license_type)
                else:
                    license_types_without_usage.append(license_type)

print("="*60)
print(f"PKS Sheet Usage Data Summary")
print("="*60)
print(f"\n✅ License types WITH usage data ({len(license_types_with_usage)}):\n")
for lt in license_types_with_usage:
    print(f"  ✓ {lt}")

print(f"\n❌ License types WITHOUT usage data ({len(license_types_without_usage)}):\n")
for lt in license_types_without_usage:
    print(f"  ✗ {lt}")

print(f"\n{'='*60}")
print(f"Total license types checked: {len(license_types_with_usage) + len(license_types_without_usage)}")
print(f"Usage data displayed for: {len(license_types_with_usage)} types ({100*len(license_types_with_usage)/(len(license_types_with_usage)+len(license_types_without_usage)):.1f}%)")
print("="*60)
