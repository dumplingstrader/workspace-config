"""Add Executive Summary sheet to Enhanced Integrity Deployment Readiness Checklist"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load existing workbook
workbook_path = r'c:\Users\GF99\Documentation\Integrity\Audits\Enhanced_Integrity_Deployment_Readiness_Checklist.xlsx'
wb = load_workbook(workbook_path)

# Create Executive Summary sheet at position 0
ws = wb.create_sheet('Executive Summary', 0)

# Set column width
ws.column_dimensions['A'].width = 120

# Executive Summary content
summary_content = [
    ('HEXAGON INTEGRITY DEPLOYMENT - EXECUTIVE SUMMARY', 'header'),
    ('Critical Success Factors for Site Readiness', 'subheader'),
    ('', 'normal'),
    ('Based on lessons learned from Marathon Petroleum multi-site deployments (ROB, LAR, GBR, SPP, GVL, SLC)', 'normal'),
    ('', 'normal'),
    ('🎯 PROJECT OBJECTIVES', 'section'),
    ('• Centralized control system configuration management across all refinery assets', 'normal'),
    ('• Enable cross-system signal tracing (PLC → DCS → Safety → Monitoring)', 'normal'),
    ('• Establish single pane of glass enterprise dashboard (L5 visibility)', 'normal'),
    ('• Maintain asset integrity while implementing standardized hierarchies', 'normal'),
    ('', 'normal'),
    ('⚠️ TOP REASONS FOR DEPLOYMENT DELAYS & COST OVERRUNS', 'section'),
    ('', 'normal'),
    ('1. INCOMPLETE DATA COLLECTION (60% of delays)', 'subsection'),
    ('   • Missing L1 inventories (Relays, Vibration Monitoring, Analyzers not tracked by Process Control)', 'normal'),
    ('   • Stakeholders engaged AFTER OSW sign-off instead of during planning', 'normal'),
    ('   • Programs/backups not accessible on workstations when consultant arrives', 'normal'),
    ('   → IMPACT: 3-6 week delays, $50K-$100K rework per site', 'impact'),
    ('', 'normal'),
    ('2. SCOPE CHANGES AFTER OSW SIGN-OFF (30% of cost overruns)', 'subsection'),
    ('   • Example: GBR added 1,614 new assets requiring complete database rebuild', 'normal'),
    ('   • Asset hierarchy changes discovered during deployment (Area/Domain mismatches)', 'normal'),
    ('   • Quote: "We were not informed about this project" - GBR site personnel', 'normal'),
    ('   → IMPACT: $20K-$75K per scope change, 2-4 week schedule delays', 'impact'),
    ('', 'normal'),
    ('3. DATA MOVEMENT INFRASTRUCTURE NOT READY (20% of delays)', 'subsection'),
    ('   • L1→L2→L4 automated transfer not configured before deployment', 'normal'),
    ('   • Server permissions/accounts not established', 'normal'),
    ('   • SQL backup/restore not tested', 'normal'),
    ('   → IMPACT: Deployment cannot start, 1-3 week delays', 'impact'),
    ('', 'normal'),
    ('4. TECHNICAL PREREQUISITE GAPS (15% of issues)', 'subsection'),
    ('   • Carbon Black blocking Hexagon utilities', 'normal'),
    ('   • Missing vendor software/licenses on workstations', 'normal'),
    ('   • Network architecture undecided (Azure migration, server locations)', 'normal'),
    ('   → IMPACT: Consultant idle onsite, $10K-$30K wasted travel costs', 'impact'),
    ('', 'normal'),
    ('💰 FINANCIAL PERFORMANCE', 'section'),
    ('', 'normal'),
    ('Marathon Program (Dec 2025): Budget Spent 43.8% | Work Complete 27.4% | CPI 0.62', 'metric'),
    ('Translation: Spending $1.00 to get $0.62 worth of work', 'metric'),
    ('Primary Cost Drivers: Re-work ($900/hr) + Pre-work ($275/hr) + Deviations ($275-900/hr)', 'normal'),
    ('', 'normal'),
    ('Sites with BEST Performance: Complete OSW before sign-off + Remote deployment + Single contact', 'success'),
    ('Savings: $15K-$25K travel per site when remote deployment used', 'success'),
    ('', 'normal'),
    ('✅ CRITICAL SUCCESS FACTORS - DO THESE FIRST', 'section'),
    ('', 'normal'),
    ('BEFORE OSW (3-4 weeks lead time):', 'subsection'),
    ('   1. Designate Site Lead (20% FTE, 6-8 weeks)', 'normal'),
    ('   2. Engage ALL stakeholders: Process, Reliability, Electrical, IT/OT', 'normal'),
    ('   3. Complete L1 inventory (use template)', 'normal'),
    ('   4. Verify programs accessible', 'normal'),
    ('', 'normal'),
    ('DURING OSW (2-3 weeks):', 'subsection'),
    ('   5. Answer Hexagon questions within 48-72 hours', 'normal'),
    ('   6. Validate hierarchies match FDS standards', 'normal'),
    ('   7. Classify supported vs unsupported devices', 'normal'),
    ('   8. Executive sign-off: No additions after OSW', 'normal'),
    ('', 'normal'),
    ('BEFORE DEPLOYMENT (1-2 weeks):', 'subsection'),
    ('   9. Complete Onsite Readiness 100%', 'normal'),
    ('   10. Test L1→L4 data movement with fresh data', 'normal'),
    ('   11. Establish accounts and grant access', 'normal'),
    ('   12. Schedule SME availability', 'normal'),
    ('', 'normal'),
    ('📋 HOW TO USE THIS CHECKLIST', 'section'),
    ('', 'normal'),
    ('Phase 1 - Pre-Planning (Week -4 to -1)', 'normal'),
    ('Phase 2 - OSW Completion (Week 1-3)', 'normal'),
    ('Phase 3 - Deployment Readiness (Week 4-5)', 'normal'),
    ('Phase 4 - Data Collection (Week 6-8)', 'normal'),
    ('Phase 5 - Validation & Go-Live (Week 9-10)', 'normal'),
    ('', 'normal'),
    ('Each checklist tab includes:', 'normal'),
    ('• Task description with clear accountability', 'normal'),
    ('• Owner (Site/Hexagon/IT) color-coded', 'normal'),
    ('• Status tracking column', 'normal'),
    ('• Notes for issues/decisions', 'normal'),
    ('• Prerequisites showing dependencies', 'normal'),
    ('', 'normal'),
    ('🎓 LESSONS LEARNED QUOTES', 'section'),
    ('', 'normal'),
    ('"Ad-hoc model is not suited for such projects. GBR and SPP are examples of what NOT to do when planning. ROB is an example of what NOT to do when executing."', 'quote'),
    ('   - Hexagon Delivery Manager, November 2025', 'quote_attribution'),
    ('', 'normal'),
    ('"If after sign-off the site adds more scope, it will lead to Change Orders and cost Marathon more money."', 'quote'),
    ('   - Repeated in Multiple Status Meetings', 'quote_attribution'),
    ('', 'normal'),
    ('📞 SUPPORT', 'section'),
    ('', 'normal'),
    ('Marathon: lar-integrityrequests@marathonpetroleum.com', 'normal'),
    ('Hexagon PM: (assigned during kickoff)', 'normal'),
    ('Corporate IT: (Infrastructure team)', 'normal'),
    ('', 'normal'),
    ('Version 2.0 - Enhanced for Multi-Site Deployment (January 2026)', 'version'),
]

# Style definitions
styles = {
    'header': {
        'font': Font(size=14, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    },
    'subheader': {
        'font': Font(size=12, bold=True, color='1F4E78'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    },
    'section': {
        'font': Font(size=11, bold=True, color='1F4E78'),
        'fill': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    },
    'subsection': {
        'font': Font(size=10, bold=True),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    },
    'normal': {
        'font': Font(size=10),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    },
    'impact': {
        'font': Font(size=10, italic=True, color='C00000'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    },
    'metric': {
        'font': Font(size=10, bold=True, color='C00000'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    },
    'success': {
        'font': Font(size=10, bold=True, color='00B050'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    },
    'quote': {
        'font': Font(size=10, italic=True),
        'fill': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)
    },
    'quote_attribution': {
        'font': Font(size=9, italic=True, color='7F7F7F'),
        'alignment': Alignment(horizontal='right', vertical='center', wrap_text=True)
    },
    'version': {
        'font': Font(size=9, italic=True, color='7F7F7F'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
    }
}

# Add content with styling
for idx, (text, style_name) in enumerate(summary_content, start=1):
    cell = ws.cell(row=idx, column=1, value=text)
    
    # Apply style
    style = styles.get(style_name, styles['normal'])
    cell.font = style['font']
    cell.alignment = style['alignment']
    if 'fill' in style:
        cell.fill = style['fill']
    
    # Adjust row height for wrapped text
    if text and len(text) > 100:
        ws.row_dimensions[idx].height = 30
    elif text and len(text) > 80:
        ws.row_dimensions[idx].height = 25
    else:
        ws.row_dimensions[idx].height = 18

# Save workbook
wb.save(workbook_path)

print("✓ Added Executive Summary sheet at the beginning")
print("✓ Includes lessons learned from Marathon deployments")
print("✓ Highlights critical success factors and cost drivers")
print(f"✓ Total: {len(summary_content)} summary items formatted")
