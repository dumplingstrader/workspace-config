"""
Enhanced Hexagon Integrity Deployment Readiness Checklist
Based on Marathon Petroleum multi-site deployment lessons learned
Version 2.0 - January 2026
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create workbook
wb = Workbook()
wb.remove(wb.active)

# Define colors
TITLE_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PHASE_FILL = PatternFill(start_color="95B3D7", end_color="95B3D7", fill_type="solid")
SITE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
HEXAGON_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
IT_FILL = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_checklist_sheet(wb, sheet_name, phase_data):
    """Create a checklist sheet with proper formatting"""
    ws = wb.create_sheet(sheet_name)
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 35
    
    # Headers
    headers = ['#', 'Checklist Item', 'Owner', 'Status', 'Notes', 'Prerequisites']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    
    # Add data
    row_num = 2
    for section in phase_data:
        # Section header
        cell = ws.cell(row=row_num, column=1, value=section['section'])
        cell.font = Font(bold=True, size=12)
        cell.fill = PHASE_FILL
        ws.merge_cells(f'A{row_num}:F{row_num}')
        row_num += 1
        
        # Section items
        for idx, item in enumerate(section['items'], 1):
            ws.cell(row=row_num, column=1, value=idx).border = BORDER
            
            cell_b = ws.cell(row=row_num, column=2, value=item['task'])
            cell_b.alignment = Alignment(wrap_text=True, vertical='top')
            cell_b.border = BORDER
            
            cell_c = ws.cell(row=row_num, column=3, value=item['owner'])
            cell_c.alignment = Alignment(horizontal='center')
            cell_c.border = BORDER
            if item['owner'] == 'Site':
                cell_c.fill = SITE_FILL
            elif item['owner'] == 'Hexagon':
                cell_c.fill = HEXAGON_FILL
            elif 'IT' in item['owner']:
                cell_c.fill = IT_FILL
            
            ws.cell(row=row_num, column=4, value='Not Started').border = BORDER
            ws.cell(row=row_num, column=5, value='').border = BORDER
            
            cell_f = ws.cell(row=row_num, column=6, value=item.get('prereq', ''))
            cell_f.alignment = Alignment(wrap_text=True, vertical='top')
            cell_f.border = BORDER
            
            row_num += 1
        
        row_num += 1  # Blank row between sections
    
    return ws

# ===== PRE-PLANNING PHASE =====
preplanning_data = [
    {
        'section': '1. STAKEHOLDER IDENTIFICATION & ENGAGEMENT',
        'items': [
            {'task': 'Designate Site Lead with decision authority and 20% FTE allocation for 6-8 weeks', 'owner': 'Site', 'prereq': 'Executive approval'},
            {'task': 'Identify Process Control Subject Matter Expert (DCS/SCADA owner)', 'owner': 'Site', 'prereq': 'Site Lead assigned'},
            {'task': 'Identify Reliability Subject Matter Expert (vibration monitoring, analyzers)', 'owner': 'Site', 'prereq': 'Site Lead assigned'},
            {'task': 'Identify Electrical/Instrumentation Subject Matter Expert (relays, field devices)', 'owner': 'Site', 'prereq': 'Site Lead assigned'},
            {'task': 'Identify IT/OT Infrastructure contact (servers, network, accounts)', 'owner': 'Site', 'prereq': 'Site Lead assigned'},
            {'task': 'Schedule internal kickoff meeting with ALL stakeholders (inform about project scope, timeline, expectations)', 'owner': 'Site', 'prereq': 'All SMEs identified'},
            {'task': 'Confirm site commitment: No additions to scope after OSW sign-off without formal Change Order', 'owner': 'Site', 'prereq': 'Executive sign-off'},
        ]
    },
    {
        'section': '2. L1 DEVICE INVENTORY PREPARATION',
        'items': [
            {'task': 'Obtain complete list of Honeywell EPKS systems (all nodes, controllers, stations)', 'owner': 'Site', 'prereq': 'Process Control SME engaged'},
            {'task': 'Obtain complete list of Honeywell TPS systems (all highways, boxes, controllers)', 'owner': 'Site', 'prereq': 'Process Control SME engaged'},
            {'task': 'Obtain complete list of Allen-Bradley PLCs (ControlLogix, PLC5, CompactLogix, Micro800)', 'owner': 'Site', 'prereq': 'Process Control SME engaged'},
            {'task': 'Obtain complete list of Schneider/Triconex SIS systems (all racks, models, firmware versions)', 'owner': 'Site', 'prereq': 'Process Control SME engaged'},
            {'task': 'Obtain complete list of Honeywell Safety Manager systems', 'owner': 'Site', 'prereq': 'Process Control SME engaged'},
            {'task': 'Obtain complete list of Bently Nevada vibration monitoring systems (3500/3600 racks)', 'owner': 'Site', 'prereq': 'Reliability SME engaged'},
            {'task': 'Obtain complete list of Programmable Relays (Schweitzer, GE, ABB, Siemens)', 'owner': 'Site', 'prereq': 'Electrical SME engaged'},
            {'task': 'Obtain complete list of Motor Exciter PLCs (if applicable)', 'owner': 'Site', 'prereq': 'Electrical SME engaged'},
            {'task': 'Obtain complete list of Analyzers with PLC control (Rosemount, ABB, etc.)', 'owner': 'Site', 'prereq': 'Reliability SME engaged'},
            {'task': 'Obtain complete list of VFDs (Variable Frequency Drives - PowerFlex, etc.)', 'owner': 'Site', 'prereq': 'Electrical SME engaged'},
            {'task': 'Obtain complete list of OSI PI or PHD historians', 'owner': 'Site', 'prereq': 'IT/OT SME engaged'},
            {'task': 'Verify inventory lists include: Asset Name, Location/Area, Model/Type, Software/Firmware Version, IP Address', 'owner': 'Site', 'prereq': 'All device lists collected'},
        ]
    },
    {
        'section': '3. PROGRAM/BACKUP ACCESSIBILITY VERIFICATION',
        'items': [
            {'task': 'Verify DCS program backups are current (last 30 days) and accessible on engineering workstations', 'owner': 'Site', 'prereq': 'L1 inventory complete'},
            {'task': 'Verify PLC programs (.ACD, .RSS, .SLC files) are accessible', 'owner': 'Site', 'prereq': 'L1 inventory complete'},
            {'task': 'Verify Triconex programs + firmware exports (Diagnostic Expert) are accessible', 'owner': 'Site', 'prereq': 'L1 inventory complete'},
            {'task': 'Verify Safety Manager programs are accessible', 'owner': 'Site', 'prereq': 'L1 inventory complete'},
            {'task': 'Verify Bently Nevada rack configurations + AssetInfo.txt files are accessible', 'owner': 'Site', 'prereq': 'L1 inventory complete'},
            {'task': 'Verify Relay configuration files (.RDB, .SET) are accessible', 'owner': 'Site', 'prereq': 'L1 inventory complete'},
            {'task': 'Verify Honeywell CAB files can be generated from Control Builder', 'owner': 'Site', 'prereq': 'L1 inventory complete'},
            {'task': 'Test read access to all program storage locations (network drives, workstations)', 'owner': 'Site', 'prereq': 'IT/OT SME engaged'},
        ]
    },
    {
        'section': '4. DECISION ON NEW vs. UPGRADE INSTALLATION',
        'items': [
            {'task': 'Decide: Fresh installation (clean start, no historical data) OR Upgrade (preserve existing Integrity data)', 'owner': 'Site', 'prereq': 'Corporate IT input'},
            {'task': 'If UPGRADE: Confirm current Integrity database backup is available and restorable', 'owner': 'Site', 'prereq': 'Decision made'},
            {'task': 'If FRESH: Confirm acceptable to lose historical data and start new baseline', 'owner': 'Site', 'prereq': 'Decision made'},
            {'task': 'Confirm decision documented and communicated to Hexagon during kickoff', 'owner': 'Site', 'prereq': 'Decision finalized'},
        ]
    },
]

# ===== OSW COMPLETION PHASE =====
osw_data = [
    {
        'section': '1. OSW WORKBOOK COMPLETION (Hexagon 5B Site Lead ICS Form)',
        'items': [
            {'task': 'Receive OSW template from Hexagon Project Manager', 'owner': 'Hexagon', 'prereq': 'Kickoff meeting held'},
            {'task': 'Complete ALL fields in OSW workbook for every device in L1 inventory', 'owner': 'Site', 'prereq': 'Pre-planning complete'},
            {'task': 'Populate Asset Hierarchy (Complex, Unit, Area, Domain) per FDS naming conventions', 'owner': 'Site', 'prereq': 'Corporate standards confirmed'},
            {'task': 'Verify Asset Names follow 50-character limit and naming standards', 'owner': 'Site', 'prereq': 'Asset hierarchy defined'},
            {'task': 'Classify each asset as Supported (automatic import) or Unsupported (manual inventory later)', 'owner': 'Hexagon', 'prereq': 'Site provides device details'},
            {'task': 'Answer ALL Hexagon questions/clarifications within 48-72 hours', 'owner': 'Site', 'prereq': 'Site Lead monitors email'},
            {'task': 'Schedule weekly status calls between Site Lead and Hexagon PM', 'owner': 'Both', 'prereq': 'OSW work started'},
        ]
    },
    {
        'section': '2. EXTERNAL REFERENCES CONFIGURATION',
        'items': [
            {'task': 'Provide drawings showing how 3rd party devices connect to DCS (SCADA channels, EPLCG boxes, PCDI)', 'owner': 'Site', 'prereq': 'Process Control SME available'},
            {'task': 'Document PLC-to-DCS connections: Highway numbers, Box numbers, EPLCG configurations', 'owner': 'Site', 'prereq': 'Drawings available'},
            {'task': 'Document Triconex-to-EPKS connections: MODBUS addresses, PCDI Master names, IP addresses', 'owner': 'Site', 'prereq': 'Safety system SME available'},
            {'task': 'Document Bently Nevada-to-DCS connections: MODBUS addresses, SCADA controllers, Gateway slots', 'owner': 'Site', 'prereq': 'Reliability SME available'},
            {'task': 'Populate External Reference properties in OSW for all cross-system connections', 'owner': 'Site', 'prereq': 'Connection documentation complete'},
            {'task': 'Review External References with Hexagon to validate configurations', 'owner': 'Hexagon', 'prereq': 'Site provides connection data'},
        ]
    },
    {
        'section': '3. OSW VALIDATION & SIGN-OFF',
        'items': [
            {'task': 'Hexagon reviews completed OSW for completeness and consistency', 'owner': 'Hexagon', 'prereq': 'Site submits completed OSW'},
            {'task': 'Site addresses any Hexagon findings/questions from OSW review', 'owner': 'Site', 'prereq': 'Hexagon review complete'},
            {'task': 'Site validates OSW represents 100% complete scope (no known missing assets)', 'owner': 'Site', 'prereq': 'All SMEs review'},
            {'task': 'Site validates asset hierarchies match corporate FDS standards', 'owner': 'Site', 'prereq': 'Corporate IT review'},
            {'task': 'Executive sign-off: OSW scope is final, no additions without Change Order', 'owner': 'Site', 'prereq': 'All validation complete'},
            {'task': 'OSW formal sign-off by Site Lead and Hexagon PM', 'owner': 'Both', 'prereq': 'Executive approval obtained'},
            {'task': 'Hexagon creates deployment plan and resource schedule', 'owner': 'Hexagon', 'prereq': 'OSW signed off'},
        ]
    },
]

# ===== DEPLOYMENT READINESS PHASE =====
deployment_readiness_data = [
    {
        'section': '1. INFRASTRUCTURE SETUP (L1 → L2 → L4 Data Movement)',
        'items': [
            {'task': 'Confirm L4 Integrity Data Collector server is deployed and accessible', 'owner': 'Corporate IT', 'prereq': 'OSW signed off'},
            {'task': 'Confirm L5 Integrity Web/SQL server is deployed and accessible', 'owner': 'Corporate IT', 'prereq': 'OSW signed off'},
            {'task': 'Install Hexagon Integrity software on L4 Data Collector server', 'owner': 'Hexagon', 'prereq': 'Servers provisioned'},
            {'task': 'Install Hexagon Integrity software on L5 Web server', 'owner': 'Hexagon', 'prereq': 'Servers provisioned'},
            {'task': 'Restore Integrity seed database on L5 SQL Server (or existing database if upgrade)', 'owner': 'Corporate IT', 'prereq': 'Hexagon provides seed DB'},
            {'task': 'Configure automated data movement L1 → L2 (batch scripts, scheduled tasks)', 'owner': 'Site IT/OT', 'prereq': 'Engineering workstations ready'},
            {'task': 'Configure automated data movement L2 → L4 (FTP, file shares, SSH)', 'owner': 'Corporate IT', 'prereq': 'L4 server ready'},
            {'task': 'Test data movement end-to-end with sample files', 'owner': 'Corporate IT', 'prereq': 'Scripts configured'},
            {'task': 'Verify fresh data (last 48 hours) appears on L4 import folders', 'owner': 'Site IT/OT', 'prereq': 'Data movement tested'},
            {'task': 'Document data refresh frequency (hourly, daily, weekly) for each asset type', 'owner': 'Site IT/OT', 'prereq': 'Data movement working'},
        ]
    },
    {
        'section': '2. ACCOUNTS & PERMISSIONS',
        'items': [
            {'task': 'Create service account for Hexagon Data Collector service', 'owner': 'Corporate IT', 'prereq': 'Security requirements defined'},
            {'task': 'Grant service account read access to L4 import folders', 'owner': 'Corporate IT', 'prereq': 'Service account created'},
            {'task': 'Grant service account SQL db_owner role on Integrity database', 'owner': 'Corporate IT', 'prereq': 'Database restored'},
            {'task': 'Create user accounts for Hexagon consultants (L4, L5 access)', 'owner': 'Corporate IT', 'prereq': 'Consultant names provided'},
            {'task': 'Grant Hexagon consultants RDP access to L4/L5 servers', 'owner': 'Corporate IT', 'prereq': 'User accounts created'},
            {'task': 'Grant Hexagon consultants access to engineering workstations (if onsite data collection)', 'owner': 'Site IT/OT', 'prereq': 'Deployment plan confirmed'},
            {'task': 'Disable Carbon Black or whitelist Hexagon utilities (CDataload.exe, etc.)', 'owner': 'Site IT/OT', 'prereq': 'Security approval'},
            {'task': 'Test Hexagon consultant login and access to all required systems', 'owner': 'Hexagon', 'prereq': 'Accounts provisioned'},
        ]
    },
    {
        'section': '3. VENDOR SOFTWARE INSTALLATION (Engineering Workstations)',
        'items': [
            {'task': 'Install Honeywell Control Builder (EPKS) - version matching site systems', 'owner': 'Site IT/OT', 'prereq': 'License available'},
            {'task': 'Install Honeywell TPS software (if TPS systems present)', 'owner': 'Site IT/OT', 'prereq': 'License available'},
            {'task': 'Install Allen-Bradley Studio 5000 (all versions used at site)', 'owner': 'Site IT/OT', 'prereq': 'License available'},
            {'task': 'Install Allen-Bradley RSLogix 500/5000 (for older PLC5/SLC systems)', 'owner': 'Site IT/OT', 'prereq': 'License available'},
            {'task': 'Install Schneider Tristation + Diagnostic Expert (all versions)', 'owner': 'Site IT/OT', 'prereq': 'License available'},
            {'task': 'Install Honeywell Safety Manager software', 'owner': 'Site IT/OT', 'prereq': 'License available'},
            {'task': 'Install Bently Nevada System 1 software', 'owner': 'Site IT/OT', 'prereq': 'License available'},
            {'task': 'Install 7-Zip (for TPS file compression)', 'owner': 'Site IT/OT', 'prereq': 'No license required'},
            {'task': 'Test each software launches and can open sample program files', 'owner': 'Site IT/OT', 'prereq': 'All software installed'},
            {'task': 'Create list of installed software versions for Hexagon reference', 'owner': 'Site IT/OT', 'prereq': 'Testing complete'},
        ]
    },
    {
        'section': '4. ONSITE READINESS CHECKLIST COMPLETION',
        'items': [
            {'task': 'Complete Hexagon Onsite Readiness Checklist (all action items)', 'owner': 'Site', 'prereq': 'Checklist received from Hexagon'},
            {'task': 'Schedule Onsite Readiness Meeting with Hexagon PM', 'owner': 'Both', 'prereq': 'Checklist 90% complete'},
            {'task': 'Review checklist items and resolve any open issues/blockers', 'owner': 'Both', 'prereq': 'Meeting held'},
            {'task': 'Confirm site SME availability for Data Collection phase (dates, % FTE)', 'owner': 'Site', 'prereq': 'Deployment schedule known'},
            {'task': 'Decision: Remote deployment or Onsite consultant visit', 'owner': 'Both', 'prereq': 'Readiness confirmed'},
            {'task': 'If Remote: Test screen share and remote access to engineering workstations', 'owner': 'Site IT/OT', 'prereq': 'Remote deployment chosen'},
            {'task': 'If Onsite: Book consultant travel and coordinate site access badges', 'owner': 'Hexagon', 'prereq': 'Onsite deployment chosen'},
        ]
    },
]

# ===== DATA COLLECTION PHASE =====
data_collection_data = [
    {
        'section': '1. DCS DATA COLLECTION (EPKS, TPS)',
        'items': [
            {'task': 'Run Honeywell Control Builder database export utility for each EPKS node', 'owner': 'Site SME', 'prereq': 'Hexagon provides instructions'},
            {'task': 'Generate CAB files (inventory) for each EPKS system', 'owner': 'Site SME', 'prereq': 'UDC utility configured'},
            {'task': 'Run TPS extraction utility for each TPS system', 'owner': 'Site SME', 'prereq': 'Hexagon provides scripts'},
            {'task': 'Aggregate TPS files with 7-Zip (per Hexagon instructions)', 'owner': 'Site SME', 'prereq': '7-Zip installed'},
            {'task': 'Copy DCS files to L1 staging folder', 'owner': 'Site SME', 'prereq': 'Extraction complete'},
            {'task': 'Verify files automatically transfer to L2, then L4 import folder', 'owner': 'Site IT/OT', 'prereq': 'Files copied to L1'},
            {'task': 'Hexagon validates DCS data imports successfully', 'owner': 'Hexagon', 'prereq': 'Files on L4'},
        ]
    },
    {
        'section': '2. PLC DATA COLLECTION',
        'items': [
            {'task': 'Export ControlLogix programs (.ACD files)', 'owner': 'Site SME', 'prereq': 'Studio 5000 installed'},
            {'task': 'Export PLC5/SLC programs (.RSS, .SLC files)', 'owner': 'Site SME', 'prereq': 'RSLogix installed'},
            {'task': 'Copy PLC programs to L1 staging folder (organized by asset name)', 'owner': 'Site SME', 'prereq': 'Export complete'},
            {'task': 'Verify PLC files transfer to L4', 'owner': 'Site IT/OT', 'prereq': 'Files copied'},
            {'task': 'Hexagon validates PLC data imports successfully', 'owner': 'Hexagon', 'prereq': 'Files on L4'},
        ]
    },
    {
        'section': '3. SAFETY SYSTEM DATA COLLECTION (Triconex, Safety Manager)',
        'items': [
            {'task': 'Export Triconex programs + firmware using Diagnostic Expert', 'owner': 'Site SME', 'prereq': 'Tristation + Diag Expert installed'},
            {'task': 'Export Safety Manager programs', 'owner': 'Site SME', 'prereq': 'Safety Manager software installed'},
            {'task': 'Copy safety system files to L1 staging folder', 'owner': 'Site SME', 'prereq': 'Export complete'},
            {'task': 'Verify safety files transfer to L4', 'owner': 'Site IT/OT', 'prereq': 'Files copied'},
            {'task': 'Hexagon validates safety data imports successfully', 'owner': 'Hexagon', 'prereq': 'Files on L4'},
        ]
    },
    {
        'section': '4. RELIABILITY/ELECTRICAL DATA COLLECTION',
        'items': [
            {'task': 'Generate Bently Nevada AssetInfo.txt files from System 1', 'owner': 'Site SME', 'prereq': 'Hexagon provides instructions'},
            {'task': 'Export Bently Nevada rack configurations (.RAK files)', 'owner': 'Site SME', 'prereq': 'System 1 installed'},
            {'task': 'Export Relay configuration files (.RDB for Schweitzer, .SET for GE)', 'owner': 'Site SME', 'prereq': 'Relay software installed'},
            {'task': 'Export Analyzer PLC programs (if applicable)', 'owner': 'Site SME', 'prereq': 'Analyzer software installed'},
            {'task': 'Copy reliability/electrical files to L1 staging folder', 'owner': 'Site SME', 'prereq': 'Export complete'},
            {'task': 'Verify files transfer to L4', 'owner': 'Site IT/OT', 'prereq': 'Files copied'},
            {'task': 'Hexagon validates reliability/electrical data imports successfully', 'owner': 'Hexagon', 'prereq': 'Files on L4'},
        ]
    },
    {
        'section': '5. HISTORIAN/OTHER DATA COLLECTION',
        'items': [
            {'task': 'Provide OSI PI or PHD connection information (if in scope)', 'owner': 'Site SME', 'prereq': 'Scope confirmed'},
            {'task': 'Export historian configuration data', 'owner': 'Site SME', 'prereq': 'Hexagon provides instructions'},
            {'task': 'Copy historian files to L1 staging folder', 'owner': 'Site SME', 'prereq': 'Export complete'},
            {'task': 'Hexagon validates historian data imports successfully', 'owner': 'Hexagon', 'prereq': 'Files on L4'},
        ]
    },
    {
        'section': '6. DATA QUALITY CHECKS',
        'items': [
            {'task': 'Hexagon runs Data Quality report (missing files, import errors)', 'owner': 'Hexagon', 'prereq': 'All data collected'},
            {'task': 'Site addresses any missing files or data gaps within 48 hours', 'owner': 'Site SME', 'prereq': 'Report received'},
            {'task': 'Hexagon validates 100% of OSW scope is imported', 'owner': 'Hexagon', 'prereq': 'Missing items resolved'},
            {'task': 'Hexagon builds Integrity database with imported data', 'owner': 'Hexagon', 'prereq': 'Data quality validated'},
            {'task': 'Site reviews sample cross-references and loop diagrams for accuracy', 'owner': 'Site SME', 'prereq': 'Database built'},
        ]
    },
]

# ===== VALIDATION & GO-LIVE PHASE =====
validation_data = [
    {
        'section': '1. SYSTEM ACCEPTANCE TESTING (SAT)',
        'items': [
            {'task': 'Hexagon schedules SAT session with Site Lead and key SMEs', 'owner': 'Hexagon', 'prereq': 'Database complete'},
            {'task': 'SAT: Verify asset hierarchy is correct (Complex, Unit, Area, Domain)', 'owner': 'Both', 'prereq': 'SAT session scheduled'},
            {'task': 'SAT: Verify all assets from OSW are visible in Integrity', 'owner': 'Both', 'prereq': 'SAT in progress'},
            {'task': 'SAT: Test cross-references between systems (PLC → DCS, Triconex → EPKS, etc.)', 'owner': 'Both', 'prereq': 'SAT in progress'},
            {'task': 'SAT: Test search functionality (find tag, find loop, find references)', 'owner': 'Both', 'prereq': 'SAT in progress'},
            {'task': 'SAT: Test Spares Management (view spare I/O, make test reservation)', 'owner': 'Both', 'prereq': 'SAT in progress'},
            {'task': 'SAT: Test Change Tracking (verify configuration changes are logged)', 'owner': 'Both', 'prereq': 'SAT in progress'},
            {'task': 'SAT: Test reporting (run sample queries, generate Excel exports)', 'owner': 'Both', 'prereq': 'SAT in progress'},
            {'task': 'SAT: Verify data refresh is working (check last import timestamp)', 'owner': 'Both', 'prereq': 'SAT in progress'},
            {'task': 'Site documents SAT findings and creates punch list for any issues', 'owner': 'Site', 'prereq': 'SAT complete'},
            {'task': 'Hexagon addresses SAT punch list items', 'owner': 'Hexagon', 'prereq': 'Punch list provided'},
            {'task': 'Site approves SAT completion', 'owner': 'Site', 'prereq': 'Punch list resolved'},
        ]
    },
    {
        'section': '2. USER TRAINING',
        'items': [
            {'task': 'Hexagon delivers INT-101 training (basic navigation, search, viewers)', 'owner': 'Hexagon', 'prereq': 'SAT approved'},
            {'task': 'Hexagon delivers Spares Management training (reserve/manage spare I/O)', 'owner': 'Hexagon', 'prereq': 'INT-101 complete'},
            {'task': 'Hexagon delivers Change Tracking training (view modifications, audit trail)', 'owner': 'Hexagon', 'prereq': 'INT-101 complete'},
            {'task': 'Hexagon delivers Query/Report training (create custom queries, export data)', 'owner': 'Hexagon', 'prereq': 'INT-101 complete'},
            {'task': 'Site identifies ongoing Integrity administrator(s) for INT-500 training', 'owner': 'Site', 'prereq': 'User training complete'},
            {'task': 'Hexagon schedules INT-500 administrator training (future)', 'owner': 'Hexagon', 'prereq': 'Administrator identified'},
        ]
    },
    {
        'section': '3. GO-LIVE PREPARATION',
        'items': [
            {'task': 'Configure Integrity user accounts for site personnel (Active Directory sync)', 'owner': 'Corporate IT', 'prereq': 'User list provided'},
            {'task': 'Assign security roles to users (Viewer, Power User, Administrator)', 'owner': 'Corporate IT', 'prereq': 'Accounts created'},
            {'task': 'Publish Integrity URL to site personnel (L5 web server)', 'owner': 'Site', 'prereq': 'User accounts ready'},
            {'task': 'Enable L5 Enterprise Dashboard (if multi-site fleet visibility required)', 'owner': 'Hexagon', 'prereq': 'Corporate IT approval'},
            {'task': 'Document site-specific customizations (custom queries, reports, hierarchies)', 'owner': 'Hexagon', 'prereq': 'Customizations approved'},
            {'task': 'Deliver "How to keep Integrity evergreen" documentation (add/delete assets)', 'owner': 'Hexagon', 'prereq': 'Go-live approved'},
        ]
    },
    {
        'section': '4. POST-GO-LIVE SUPPORT',
        'items': [
            {'task': 'Establish site support contact for Integrity issues (email: lar-integrityrequests@marathonpetroleum.com)', 'owner': 'Site', 'prereq': 'Go-live complete'},
            {'task': 'Schedule 30-day post-go-live check-in (usage, issues, questions)', 'owner': 'Hexagon', 'prereq': 'Go-live complete'},
            {'task': 'Add unsupported assets to separate tracking file for future manual inventory', 'owner': 'Hexagon', 'prereq': 'Go-live complete'},
            {'task': 'Plan for adding new assets as site evolves (turnarounds, projects)', 'owner': 'Site', 'prereq': 'Documentation received'},
            {'task': 'Close project and transition to sustainment phase', 'owner': 'Both', 'prereq': '30-day check-in complete'},
        ]
    },
]

# Create sheets
create_checklist_sheet(wb, "1-Pre-Planning", preplanning_data)
create_checklist_sheet(wb, "2-OSW Completion", osw_data)
create_checklist_sheet(wb, "3-Deployment Readiness", deployment_readiness_data)
create_checklist_sheet(wb, "4-Data Collection", data_collection_data)
create_checklist_sheet(wb, "5-Validation & Go-Live", validation_data)

# Save workbook
output_path = "c:/Users/GF99/Documentation/Integrity/Audits/Enhanced_Integrity_Deployment_Readiness_Checklist.xlsx"
wb.save(output_path)
print(f"✓ Created enhanced checklist: {output_path}")
print(f"✓ Includes 5 phase-based checklists with {sum(len(section['items']) for phase in [preplanning_data, osw_data, deployment_readiness_data, data_collection_data, validation_data] for section in phase)} total checklist items")
print("✓ Color-coded by Owner: Site (Orange) | Hexagon (Green) | IT (Blue)")
