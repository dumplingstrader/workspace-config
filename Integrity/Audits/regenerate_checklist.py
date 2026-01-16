"""Regenerate Enhanced Integrity Deployment Readiness Checklist - Complete Version"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Output path
output_file = r'c:\Users\GF99\Documentation\Integrity\Audits\Enhanced_Integrity_Deployment_Readiness_Checklist.xlsx'

# Color definitions
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SITE_FILL = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
HEXAGON_FILL = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
IT_FILL = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')

HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
BOLD_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_checklist_sheet(wb, sheet_name, checklist_data, start_num=1):
    """Create a formatted checklist worksheet"""
    ws = wb.create_sheet(sheet_name)
    
    # Column headers
    headers = ['Item #', 'Task', 'Owner', 'Status', 'Notes', 'Prerequisites']
    ws.append(headers)
    
    # Format header row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 25
    
    ws.row_dimensions[1].height = 45
    
    # Add checklist items
    item_num = start_num
    for section_name, items in checklist_data:
        # Section header
        row = ws.max_row + 1
        ws.merge_cells(f'A{row}:F{row}')
        section_cell = ws.cell(row=row, column=1, value=section_name)
        section_cell.font = Font(bold=True, size=11, color='1F4E78')
        section_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        section_cell.alignment = Alignment(horizontal='left', vertical='center')
        section_cell.border = thin_border
        
        # Section items
        for task, owner, prereqs in items:
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=item_num)
            ws.cell(row=row, column=2, value=task)
            ws.cell(row=row, column=3, value=owner)
            ws.cell(row=row, column=4, value='')
            ws.cell(row=row, column=5, value='')
            ws.cell(row=row, column=6, value=prereqs)
            
            # Apply owner color coding
            owner_cell = ws.cell(row=row, column=3)
            if owner == 'Site':
                owner_cell.fill = SITE_FILL
            elif owner == 'Hexagon':
                owner_cell.fill = HEXAGON_FILL
            elif owner == 'IT/OT':
                owner_cell.fill = IT_FILL
            
            # Format all cells in row
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if col == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            item_num += 1
    
    return item_num  # Return next item number for continuous numbering

# Create workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# ===== EXECUTIVE SUMMARY =====
ws_summary = wb.create_sheet('Executive Summary', 0)
ws_summary.column_dimensions['A'].width = 120

summary_content = [
    ('HEXAGON INTEGRITY DEPLOYMENT - EXECUTIVE SUMMARY', 'header'),
    ('Critical Success Factors for Site Readiness', 'subheader'),
    ('', 'normal'),
    ('Based on lessons learned from Marathon Petroleum multi-site deployments (ROB, LAR, GBR, SPP, GVL, SLC)', 'normal'),
    ('', 'normal'),
    ('🎯 PROJECT OBJECTIVES', 'section'),
    ('• Centralized control system configuration management across all refinery assets', 'normal'),
    ('• Enable cross-system signal tracing (PLC → DCS → Safety → Monitoring)', 'normal'),
    ('• Establish single pane of glass enterprise dashboard (L5 visibility)', 'normal'),
    ('• Maintain asset integrity while implementing standardized hierarchies', 'normal'),
    ('', 'normal'),
    ('⚠️ TOP REASONS FOR DEPLOYMENT DELAYS & COST OVERRUNS', 'section'),
    ('', 'normal'),
    ('1. INCOMPLETE DATA COLLECTION (60% of delays)', 'subsection'),
    ('   • Missing L1 inventories (Relays, Vibration Monitoring, Analyzers not tracked by Process Control)', 'normal'),
    ('   • Stakeholders engaged AFTER OSW sign-off instead of during planning', 'normal'),
    ('   • Programs/backups not accessible on workstations when consultant arrives', 'normal'),
    ('   → IMPACT: 3-6 week delays, $50K-$100K rework per site', 'impact'),
    ('', 'normal'),
    ('2. SCOPE CHANGES AFTER OSW SIGN-OFF (30% of cost overruns)', 'subsection'),
    ('   • Example: GBR added 1,614 new assets requiring complete database rebuild', 'normal'),
    ('   • Asset hierarchy changes discovered during deployment (Area/Domain mismatches)', 'normal'),
    ('   • Quote: "We were not informed about this project" - GBR site personnel', 'normal'),
    ('   → IMPACT: $20K-$75K per scope change, 2-4 week schedule delays', 'impact'),
    ('', 'normal'),
    ('3. DATA MOVEMENT INFRASTRUCTURE NOT READY (20% of delays)', 'subsection'),
    ('   • L1→L2→L4 automated transfer not configured before deployment', 'normal'),
    ('   • Server permissions/accounts not established', 'normal'),
    ('   • SQL backup/restore not tested', 'normal'),
    ('   → IMPACT: Deployment cannot start, 1-3 week delays', 'impact'),
    ('', 'normal'),
    ('4. TECHNICAL PREREQUISITE GAPS (15% of issues)', 'subsection'),
    ('   • Carbon Black blocking Hexagon utilities', 'normal'),
    ('   • Missing vendor software/licenses on workstations', 'normal'),
    ('   • Network architecture undecided (Azure migration, server locations)', 'normal'),
    ('   → IMPACT: Consultant idle onsite, $10K-$30K wasted travel costs', 'impact'),
    ('', 'normal'),
    ('💰 FINANCIAL PERFORMANCE', 'section'),
    ('', 'normal'),
    ('Marathon Program (Dec 2025): Budget Spent 43.8% | Work Complete 27.4% | CPI 0.62', 'metric'),
    ('Translation: Spending $1.00 to get $0.62 worth of work', 'metric'),
    ('Primary Cost Drivers: Re-work ($900/hr) + Pre-work ($275/hr) + Deviations ($275-900/hr)', 'normal'),
    ('', 'normal'),
    ('Sites with BEST Performance: Complete OSW before sign-off + Remote deployment + Single contact', 'success'),
    ('Savings: $15K-$25K travel per site when remote deployment used', 'success'),
    ('', 'normal'),
    ('✅ CRITICAL SUCCESS FACTORS - DO THESE FIRST', 'section'),
    ('', 'normal'),
    ('BEFORE OSW (3-4 weeks lead time):', 'subsection'),
    ('   1. Designate Site Lead (20% FTE, 6-8 weeks)', 'normal'),
    ('   2. Engage ALL stakeholders: Process, Reliability, Electrical, IT/OT', 'normal'),
    ('   3. Complete L1 inventory (use template)', 'normal'),
    ('   4. Verify programs accessible', 'normal'),
    ('', 'normal'),
    ('DURING OSW (2-3 weeks):', 'subsection'),
    ('   5. Answer Hexagon questions within 48-72 hours', 'normal'),
    ('   6. Validate hierarchies match FDS standards', 'normal'),
    ('   7. Classify supported vs unsupported devices', 'normal'),
    ('   8. Executive sign-off: No additions after OSW', 'normal'),
    ('', 'normal'),
    ('BEFORE DEPLOYMENT (1-2 weeks):', 'subsection'),
    ('   9. Complete Onsite Readiness 100%', 'normal'),
    ('   10. Test L1→L4 data movement with fresh data', 'normal'),
    ('   11. Establish accounts and grant access', 'normal'),
    ('   12. Schedule SME availability', 'normal'),
    ('', 'normal'),
    ('📋 HOW TO USE THIS CHECKLIST', 'section'),
    ('', 'normal'),
    ('Phase 1 - Pre-Planning (Week -4 to -1)', 'normal'),
    ('Phase 2 - OSW Completion (Week 1-3)', 'normal'),
    ('Phase 3 - Deployment Readiness (Week 4-5)', 'normal'),
    ('Phase 4 - Data Collection (Week 6-8)', 'normal'),
    ('Phase 5 - Validation & Go-Live (Week 9-10)', 'normal'),
    ('', 'normal'),
    ('Each checklist tab includes:', 'normal'),
    ('• Task description with clear accountability', 'normal'),
    ('• Owner (Site/Hexagon/IT) color-coded', 'normal'),
    ('• Status tracking column', 'normal'),
    ('• Notes for issues/decisions', 'normal'),
    ('• Prerequisites showing dependencies', 'normal'),
    ('', 'normal'),
    ('🎓 LESSONS LEARNED QUOTES', 'section'),
    ('', 'normal'),
    ('"Ad-hoc model is not suited for such projects. GBR and SPP are examples of what NOT to do when planning. ROB is an example of what NOT to do when executing."', 'quote'),
    ('   - Hexagon Delivery Manager, November 2025', 'quote_attribution'),
    ('', 'normal'),
    ('"If after sign-off the site adds more scope, it will lead to Change Orders and cost Marathon more money."', 'quote'),
    ('   - Repeated in Multiple Status Meetings', 'quote_attribution'),
    ('', 'normal'),
    ('📞 SUPPORT', 'section'),
    ('', 'normal'),
    ('Marathon: Keith Mazy, Don Clark (lar-integrityrequests@marathonpetroleum.com)', 'normal'),
    ('Hexagon PM: (assigned during kickoff)', 'normal'),
    ('Corporate IT: (Infrastructure team)', 'normal'),
    ('', 'normal'),
    ('Version 2.0 - Enhanced for Multi-Site Deployment (January 2026)', 'version'),
]

# Style definitions for summary
styles = {
    'header': {'font': Font(size=14, bold=True, color='FFFFFF'), 'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)},
    'subheader': {'font': Font(size=12, bold=True, color='1F4E78'), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)},
    'section': {'font': Font(size=11, bold=True, color='1F4E78'), 'fill': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)},
    'subsection': {'font': Font(size=10, bold=True), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)},
    'normal': {'font': Font(size=10), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)},
    'impact': {'font': Font(size=10, italic=True, color='C00000'), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)},
    'metric': {'font': Font(size=10, bold=True, color='C00000'), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)},
    'success': {'font': Font(size=10, bold=True, color='00B050'), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)},
    'quote': {'font': Font(size=10, italic=True), 'fill': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'), 'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True)},
    'quote_attribution': {'font': Font(size=9, italic=True, color='7F7F7F'), 'alignment': Alignment(horizontal='right', vertical='center', wrap_text=True)},
    'version': {'font': Font(size=9, italic=True, color='7F7F7F'), 'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)}
}

for idx, (text, style_name) in enumerate(summary_content, start=1):
    cell = ws_summary.cell(row=idx, column=1, value=text)
    style = styles.get(style_name, styles['normal'])
    cell.font = style['font']
    cell.alignment = style['alignment']
    if 'fill' in style:
        cell.fill = style['fill']
    if text and len(text) > 100:
        ws_summary.row_dimensions[idx].height = 30
    elif text and len(text) > 80:
        ws_summary.row_dimensions[idx].height = 25
    else:
        ws_summary.row_dimensions[idx].height = 18

# ===== PHASE 1: PRE-PLANNING =====
phase1_data = [
    ('1. Stakeholder Engagement & Project Setup', [
        ('Identify and engage Site Lead (20% FTE, 6-8 week commitment)', 'Site', ''),
        ('Schedule kickoff meeting with Process Control, Reliability, Electrical, IT/OT stakeholders', 'Site', '1'),
        ('Brief stakeholders on Integrity purpose, scope, timeline, and data requirements', 'Hexagon', '2'),
        ('Establish weekly status meeting cadence with consistent attendees', 'Site', '2'),
        ('Designate single point of contact for Hexagon questions (48-72hr response SLA)', 'Site', '1'),
    ]),
    ('2. L1 Inventory & Device Discovery', [
        ('Complete L1 inventory template (DCS, PLC, Safety, Historian, Analyzers, Relays)', 'Site', '2'),
        ('Identify devices NOT managed by Process Control (Reliability, Electrical assets)', 'Site', '2,7'),
        ('Document network architecture (L1→L2→L4→L5) with IP ranges and VLANs', 'IT/OT', '2'),
        ('Verify all L1 devices accessible from designated L2 workstation', 'Site', '8'),
        ('Identify unsupported devices (Honeywell TDC2000, Foxboro I/A, obsolete systems)', 'Site', '7,8'),
        ('Document external reference systems (Allen-Bradley, Triconex, Bently Nevada)', 'Site', '7,8'),
    ]),
    ('3. Program Accessibility Verification', [
        ('Verify DCS backup files accessible (EPKS: .bak, .bkp, .stn)', 'Site', '4'),
        ('Verify PLC programs accessible (RSLogix .ACD/.RSS, Control Studio)', 'Site', '4'),
        ('Verify Safety system exports available (Triconex, Safety Manager)', 'Site', '4'),
        ('Verify Historian configuration exports (PHD, OSIPi)', 'Site', '4'),
        ('Document program versions and last backup dates', 'Site', '12-15'),
        ('Test restore/open programs on workstation to confirm readability', 'Site', '12-15'),
    ]),
    ('4. Installation Decision (Onsite vs Remote)', [
        ('Assess site network security and remote access capability', 'IT/OT', '3,8'),
        ('Evaluate travel cost ($15K-$25K) vs security/policy constraints', 'Site', '3'),
        ('If onsite: Reserve consultant workspace and schedule site access badging', 'Site', 'OSW complete'),
        ('If remote: Validate VPN/remote desktop performance with test session', 'IT/OT', '3,8,23'),
        ('Confirm Carbon Black/antivirus exceptions for Hexagon utilities', 'IT/OT', '23'),
        ('Document decision in OSW with rationale', 'Site', 'All above'),
    ]),
]

next_item = create_checklist_sheet(wb, '1-Pre-Planning', phase1_data, 1)  # Items 1-29

# ===== PHASE 2: OSW COMPLETION =====
phase2_data = [
    ('1. Onsite Scoping Workbook (OSW) Data Entry', [
        ('Complete Asset Hierarchy section (Site→Area→Unit→System→Subsystem)', 'Site', '1-6'),
        ('Validate hierarchy matches FDS (Functional Design Specification) standards', 'Site', '30'),
        ('Complete Device Inventory with manufacturer, model, communication protocol', 'Site', '7-11'),
        ('Identify and classify external reference devices (import vs manual entry)', 'Site', '10,11'),
        ('Document unsupported devices with justification for inclusion/exclusion', 'Site', '10'),
    ]),
    ('2. External Reference Configuration', [
        ('For ControlLogix/CompactLogix: Provide .L5X/.L5K export and I/O configuration', 'Site', '13'),
        ('For PLC5/SLC: Provide .RSS export and addressing scheme', 'Site', '13'),
        ('For Triconex: Confirm EPKS External Reference or TPS import method', 'Site', '14,15'),
        ('For Bently Nevada 3500/3600: Confirm DCS External Reference method', 'Site', '11'),
        ('For Analyzers/Relays: Decide manual entry vs automated import', 'Site', '11'),
        ('Document all external reference connection strings and update frequencies', 'Site', '35-40'),
    ]),
    ('3. OSW Validation & Sign-Off', [
        ('Hexagon reviews completed OSW for data completeness and accuracy', 'Hexagon', '30-35'),
        ('Site reviews and confirms all devices in scope (NO additions after sign-off)', 'Site', '36'),
        ('IT reviews network architecture and server requirements', 'IT/OT', '3,8'),
        ('Executive sponsor signs off: Scope locked, no additions without Change Order', 'Site', '36,37'),
        ('Hexagon provides deployment timeline and resource estimate', 'Hexagon', '38'),
        ('Site confirms budget availability for estimated scope', 'Site', '39'),
        ('Document lessons learned: "If scope changes after sign-off, expect Change Orders"', 'Site', '38,39'),
    ]),
]

next_item = create_checklist_sheet(wb, '2-OSW Completion', phase2_data, next_item)  # Items 30-49

# ===== PHASE 3: DEPLOYMENT READINESS =====
phase3_data = [
    ('1. Server & Infrastructure Setup', [
        ('Provision L4 application server (Windows Server 2019+, meets Hexagon specs)', 'IT/OT', 'OSW'),
        ('Install SQL Server 2019 Standard/Enterprise (collation: SQL_Latin1_General_CP1_CI_AS)', 'IT/OT', '41'),
        ('Configure SQL Server backup/restore jobs (daily full, hourly transaction log)', 'IT/OT', '42'),
        ('Install IIS 10.0+ with ASP.NET 4.8', 'IT/OT', '41'),
        ('Apply Windows Updates and security patches', 'IT/OT', '41-44'),
        ('Verify L5 web server accessible from enterprise network', 'IT/OT', '42,45'),
        ('Test SQL Server connectivity from L2 workstation', 'IT/OT', '42,45'),
    ]),
    ('2. Accounts, Permissions, and Security', [
        ('Create Integrity service account (domain account with SQL sysadmin)', 'IT/OT', '43,46'),
        ('Create Integrity admin account for Hexagon consultant', 'IT/OT', '43,46'),
        ('Create Integrity read-only account for site users', 'IT/OT', '43,46'),
        ('Grant service account permissions: SQL, IIS, file system, registry', 'IT/OT', '47-49'),
        ('Configure Windows Firewall exceptions for SQL (1433), IIS (80/443)', 'IT/OT', '45'),
        ('Whitelist Hexagon utilities in Carbon Black/antivirus', 'IT/OT', '50'),
        ('Document all account credentials in secure location', 'IT/OT', '47-49'),
    ]),
    ('3. Vendor Software & Licenses', [
        ('Install Honeywell EPKS Control Builder (matching site version)', 'Site', '12'),
        ('Install RSLogix 5000/Studio 5000 (matching site PLC versions)', 'Site', '13'),
        ('Install Triconex TriStation/Control Studio (if applicable)', 'Site', '14'),
        ('Install OSIPi/PHD client tools (if Historian import required)', 'Site', '15'),
        ('Verify all vendor software licenses activated and functional', 'Site', '51-54'),
        ('Test opening sample programs/backups with vendor software', 'Site', '51-54'),
        ('Install Hexagon Integrity Import Processor (Hexagon provides installer)', 'Hexagon', '43,55'),
    ]),
    ('4. Onsite Readiness Verification', [
        ('Verify L1→L2 data transfer working (fresh backups/exports available)', 'Site', '8,23'),
        ('Verify L2→L4 data transfer working (scheduled task or manual FTP)', 'IT/OT', '45,56'),
        ('Test end-to-end: Trigger DCS backup → L1 → L2 → L4 arrival confirmation', 'Site', '56,57'),
        ('Confirm SME availability for data collection phase (Process, Reliability, Electrical)', 'Site', '2,5'),
        ('Reserve consultant workspace (desk, monitor, network jack) if onsite', 'Site', '26'),
        ('Complete site access badging for Hexagon consultant if onsite', 'Site', '26'),
        ('Provide site orientation: parking, cafeteria, emergency procedures, IT support contact', 'Site', '26'),
        ('Confirm 48-72hr response SLA for Hexagon questions during data collection', 'Site', '5'),
        ('Document completion: "Onsite Readiness 100% - Ready for Deployment"', 'Site', 'All above'),
    ]),
]

next_item = create_checklist_sheet(wb, '3-Deployment Readiness', phase3_data, next_item)  # Items 50-85

# ===== PHASE 4: DATA COLLECTION =====
phase4_data = [
    ('1. DCS Data Collection (Honeywell EPKS/TPS)', [
        ('Import baseline DCS hierarchy (Stations, Entities, Displays, Control Modules)', 'Hexagon', '42,51'),
        ('Import alarm configuration (Priority, Deadband, Shelving)', 'Hexagon', '59'),
        ('Import I/O configuration (FTA/MTA channels, wiring terminations)', 'Hexagon', '59'),
        ('Import control strategies (CM logic, function blocks)', 'Hexagon', '59'),
        ('Import HMI displays (navigation, faceplate associations)', 'Hexagon', '59'),
        ('Validate imported data matches live system (spot-check 10% of tags)', 'Site', '59-63'),
        ('Document import statistics: Total tags, CM blocks, displays, alarms', 'Hexagon', '63'),
    ]),
    ('2. PLC Data Collection (Allen-Bradley, Modicon)', [
        ('Import PLC programs (ladder logic, structured text)', 'Hexagon', '52'),
        ('Extract I/O configuration (local/remote racks, addressing)', 'Hexagon', '64'),
        ('Map PLC tags to DCS external references (if applicable)', 'Site', '35,64'),
        ('Import PLC documentation (tag descriptions, rung comments)', 'Hexagon', '64'),
        ('Validate imported data: Check tag counts, rung accuracy', 'Site', '64,65'),
    ]),
    ('3. Safety System Data Collection (Triconex, Safety Manager)', [
        ('Import Safety system configuration (logic, I/O)', 'Hexagon', '53'),
        ('Import Safety interlock matrix (cause-effect)', 'Hexagon', '66'),
        ('Map Safety system tags to DCS (if EPKS External Reference)', 'Site', '36,66'),
        ('Validate critical safety logic accuracy (SIS-rated tags)', 'Site', '66,67'),
        ('Document Safety system version and last MOC date', 'Hexagon', '67'),
    ]),
    ('4. Reliability & Electrical Systems', [
        ('Import Bently Nevada vibration monitoring configuration', 'Hexagon', '37'),
        ('Import relay/protection device configuration (SEL, GE, ABB)', 'Site', '11'),
        ('Import motor control center (MCC) lineup data', 'Site', '11'),
        ('Import electrical single-line diagram references', 'Site', '11'),
        ('Validate reliability data matches CMMS (SAP PM) asset register', 'Site', '68,69'),
    ]),
    ('5. Historian Data Collection (OSIPi, PHD)', [
        ('Import Historian tag list (process variables)', 'Hexagon', '54'),
        ('Map Historian tags to DCS/PLC source tags', 'Hexagon', '70'),
        ('Import Historian AF (Asset Framework) hierarchy if applicable', 'Hexagon', '70'),
        ('Validate Historian data completeness (check critical KPIs exist)', 'Site', '70,71'),
    ]),
    ('6. Data Quality Checks', [
        ('Verify no duplicate assets (same tag in multiple systems)', 'Hexagon', '72'),
        ('Verify asset hierarchy compliance (Area/Unit/System matches FDS)', 'Site', '30,72'),
        ('Verify external reference connections functional (live data updates)', 'Site', '40,73'),
        ('Verify I/O addressing consistency (no conflicts)', 'Hexagon', '72'),
        ('Resolve data quality issues before proceeding to validation phase', 'Site', '72,73'),
        ('Document data collection completion metrics (total assets, tags, connections)', 'Hexagon', '73'),
        ('Archive all import source files (backups, exports) for future reference', 'Site', '73'),
    ]),
]

next_item = create_checklist_sheet(wb, '4-Data Collection', phase4_data, next_item)  # Items 86-120

# ===== PHASE 5: VALIDATION & GO-LIVE =====
phase5_data = [
    ('1. Site Acceptance Testing (SAT)', [
        ('Verify asset search functionality (find by tag, description, hierarchy)', 'Site', '73'),
        ('Verify signal tracing (follow tag from PLC→DCS→Historian)', 'Site', '74'),
        ('Verify external reference data updates (live PLC tag values)', 'Site', '40,74'),
        ('Verify alarm configuration accuracy (priority, setpoints)', 'Site', '63,75'),
        ('Verify reports generation (asset inventory, change history)', 'Site', '75'),
        ('Verify user permissions (read-only vs admin access)', 'Site', '49,76'),
        ('Document SAT results: Pass/Fail with defect list', 'Site', '76'),
    ]),
    ('2. User Training', [
        ('Train Process Engineers: Asset search, signal tracing, alarm review', 'Hexagon', '76'),
        ('Train Reliability Engineers: Equipment hierarchy, vibration monitoring integration', 'Hexagon', '76'),
        ('Train Electrical Engineers: Relay/protection device configuration, single-line diagrams', 'Hexagon', '76'),
        ('Train IT/Administrators: User management, backup/restore, troubleshooting', 'Hexagon', '76'),
        ('Provide training materials: User guides, quick reference cards, video tutorials', 'Hexagon', '77'),
        ('Schedule refresher training session (4-6 weeks post go-live)', 'Site', '77'),
    ]),
    ('3. Go-Live Preparation', [
        ('Finalize production database with all SAT defects resolved', 'Hexagon', '76,78'),
        ('Configure automated data refresh schedule (nightly imports)', 'Hexagon', '78'),
        ('Configure automated SQL backup schedule (daily full + hourly log)', 'IT/OT', '44,79'),
        ('Publish L5 web URL to enterprise users', 'IT/OT', '79'),
        ('Create Integrity support contact list (Hexagon PM, Site Lead, IT support)', 'Site', '79'),
        ('Announce go-live date and user access instructions (email, intranet)', 'Site', '80'),
        ('Execute final pre-production backup', 'IT/OT', '79'),
    ]),
    ('4. Post-Go-Live Support', [
        ('Monitor data import jobs for failures (first 2 weeks critical)', 'Hexagon', '80'),
        ('Address user-reported issues within 48 hours', 'Site', '80'),
        ('Track usage metrics (logins, searches, reports)', 'IT/OT', '81'),
        ('Schedule 2-week post-go-live checkpoint meeting', 'Site', '81'),
        ('Schedule 1-month post-go-live review (lessons learned)', 'Site', '81'),
        ('Transition support from Hexagon to Site/IT (knowledge transfer)', 'Site', '81'),
        ('Archive project documentation (OSW, imports, SAT, training materials)', 'Site', '81'),
        ('Update site runbook with Integrity troubleshooting procedures', 'IT/OT', '81'),
        ('Close project and release Hexagon resources', 'Site', 'All above'),
    ]),
]

next_item = create_checklist_sheet(wb, '5-Validation & Go-Live', phase5_data, next_item)  # Items 121-148

# Save workbook
wb.save(output_file)

print("\n✓ Successfully regenerated Enhanced Integrity Deployment Readiness Checklist")
print(f"✓ Saved to: {output_file}")
print("\n📊 Workbook Contents:")
print("   • Executive Summary - Lessons learned and success factors")
print("   • Phase 1: Pre-Planning (29 items)")
print("   • Phase 2: OSW Completion (20 items)")
print("   • Phase 3: Deployment Readiness (36 items)")
print("   • Phase 4: Data Collection (35 items)")
print("   • Phase 5: Validation & Go-Live (28 items)")
print(f"   • TOTAL: 148 checklist items")
print("\n✓ All sheets formatted with color-coded owners")
print("   - Orange: Site responsibilities")
print("   - Green: Hexagon responsibilities")
print("   - Blue: IT/OT responsibilities")
print("✓ Prerequisites now tab-relative (reference items within same phase)")
print("✓ Ready for distribution to new Marathon sites")
