"""
Enhanced Process Controls Skill Matrix Generator
Creates comprehensive skill matrix with MPC 1-5 proficiency scale and individual assessment table
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Define the skill categories and sub-skills - reorganized for clarity
SKILL_TAXONOMY = {
    "Control Platforms": [
        "Honeywell Experion",
        "Honeywell TDC 3000",
        "Triconex SIS",
        "Allen-Bradley ControlLogix",
        "Allen-Bradley Legacy PLC (PLC-5, SLC500, MicroLogix)",
        "GE Mark VIe",
        "BN3500",
        "DCS HMI (Honeywell)",
        "AB HMI (PanelView/FTView)"
    ],
    "Process Knowledge": [
        "Fired heaters",
        "Compressors",
        "Rotating equipment fundamentals",
        "Instrumentation fundamentals",
        "Electrical fundamentals",
        "Advanced regulatory control"
    ],
    "Programming & Control Logic": [
        "TDC Control Language (CL)",
        "Experion SCM",
        "Interlocks & permissives",
        "Startup/shutdown sequences",
        "Cause-and-effect logic"
    ],
    "Advanced Process Control": [
        "Aspen APC",
        "Imubit APC"
    ],
    "Enterprise Systems": [
        "SAP",
        "Dynamo M&R",
        "ACM (Alarm Management)",
        "Integrity (SIS/LOPA)",
        "PI System",
        "OPC/PCDI/OPCI"
    ],
    "Infrastructure & Networking": [
        "Historian connections",
        "SCADA/network routing",
        "Controller redundancy & health",
        "Network diagnostics",
        "Virtual machines"
    ],
    "Implementation & Commissioning": [
        "Commissioning & loop checks",
        "Logic migration & upgrades",
        "Patching & updates",
        "Troubleshooting guides",
        "Peer reviews"
    ],
    "Engineering Tools & Methods": [
        "Python scripting",
        "App development",
        "Data analytics (Excel, Power BI)",
        "Version control"
    ],
    "Professional Development": [
        "Create training content",
        "Deliver training/workshops",
        "Mentoring",
        "Write procedures/standards",
        "Knowledge management"
    ],
    "Business & Leadership": [
        "Leadership",
        "Project management",
        "Coordination",
        "Budget planning",
        "Cost estimation",
        "Vendor evaluation",
        "Lifecycle cost analysis",
        "Industry presentation",
        "Documentation quality",
        "Communication",
        "Stakeholder management",
        "Time management",
        "Safety/MOC proficiency"
    ]
}

# MPC Proficiency Scale (1-5)
PROFICIENCY_SCALE = {
    1: "Forming - Awareness level, basic understanding",
    2: "Developing - Can perform with guidance",
    3: "Applying - Independent practitioner",
    4: "Leading - Expert, can guide others",
    5: "Shaping - Subject matter expert, strategic influence"
}

# Role definitions
ROLES = [
    "Process Controls Engineer I",
    "Process Controls Engineer II",
    "Process Controls Engineer III / Senior",
    "Lead Process Controls Engineer",
    "APC Engineer II",
    "APC Engineer III / Senior",
    "Lead APC Engineer",
    "Senior Technologist (P3)",
    "Lead Technologist (P4)",
    "Principal Technologist (P5)"
]

def create_header_style():
    """Create header cell style"""
    return {
        'font': Font(bold=True, size=11, color='FFFFFF'),
        'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_category_style():
    """Create category header style"""
    return {
        'font': Font(bold=True, size=10, color='FFFFFF'),
        'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_data_style():
    """Create data cell style"""
    return {
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def apply_cell_style(cell, style_dict):
    """Apply style dictionary to a cell"""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def create_skill_dictionary_sheet(wb):
    """Create the Skill Dictionary sheet"""
    ws = wb.create_sheet("Skill_Dictionary")
    
    # Headers
    headers = ["Category", "Sub-Skill", "Definition"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        apply_cell_style(cell, create_header_style())
    
    # Data
    row = 2
    for category, skills in SKILL_TAXONOMY.items():
        for skill in skills:
            ws.cell(row, 1, category)
            ws.cell(row, 2, skill)
            ws.cell(row, 3, f"Competency in {skill.lower()}")
            
            # Apply data style
            for col in range(1, 4):
                apply_cell_style(ws.cell(row, col), create_data_style())
            
            row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 50
    
    return ws

def create_proficiency_scale_sheet(wb):
    """Create the Proficiency Scale reference sheet"""
    ws = wb.create_sheet("Proficiency_Scale")
    
    # Title
    title_cell = ws.cell(1, 1, "MPC Proficiency Scale Reference")
    title_cell.font = Font(bold=True, size=14, color='366092')
    ws.merge_cells('A1:C1')
    
    # Headers
    headers = ["Level", "Name", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        apply_cell_style(cell, create_header_style())
    
    # Scale data
    scale_details = [
        (1, "Forming", "Awareness level, basic understanding. Can identify concepts and recognize when expertise is needed."),
        (2, "Developing", "Can perform with guidance. Developing proficiency, requires supervision or assistance."),
        (3, "Applying", "Independent practitioner. Can execute tasks independently with consistent quality."),
        (4, "Leading", "Expert, can guide others. Deep expertise, mentors others, drives improvements."),
        (5, "Shaping", "Subject matter expert. Strategic influence, sets standards, recognized authority.")
    ]
    
    for i, (level, name, desc) in enumerate(scale_details, 4):
        ws.cell(i, 1, level)
        ws.cell(i, 2, name)
        ws.cell(i, 3, desc)
        
        # Apply data style
        for col in range(1, 4):
            cell = ws.cell(i, col)
            apply_cell_style(cell, create_data_style())
            if col == 1:
                cell.font = Font(bold=True, size=11)
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 80
    
    return ws

def create_individual_assessment_sheet(wb):
    """Create the Individual Assessment entry table - PEOPLE AS ROWS"""
    ws = wb.create_sheet("Individual_Assessments")
    
    # Proficiency scale in Row 1
    scale_text = ws.cell(1, 1, "Scale: 1=Forming (awareness)  |  2=Developing (guided)  |  3=Applying (independent)  |  4=Leading (expert)  |  5=Shaping (SME)")
    scale_text.font = Font(italic=True, size=9, color='366092')
    ws.merge_cells('A1:E1')
    
    # Instructions in Row 2
    instr_cell = ws.cell(2, 1, "Enter proficiency level (1-5) for each person. Leave blank if not applicable.")
    instr_cell.font = Font(italic=True, size=9)
    ws.merge_cells('A2:E2')
    
    # Build column structure: First Name | Last Name | Role | Assessment Date | Category:Skill columns
    col = 1
    
    # Fixed columns
    ws.cell(4, col, "First Name")
    ws.column_dimensions[get_column_letter(col)].width = 18
    col += 1
    
    ws.cell(4, col, "Last Name")
    ws.column_dimensions[get_column_letter(col)].width = 18
    col += 1
    
    ws.cell(4, col, "Role")
    ws.column_dimensions[get_column_letter(col)].width = 30
    col += 1
    
    ws.cell(4, col, "Assessment Date")
    ws.column_dimensions[get_column_letter(col)].width = 15
    col += 1
    
    # Skill columns with category grouping
    start_col = col
    for category, skills in SKILL_TAXONOMY.items():
        category_start_col = col
        
        # Add skills
        for skill in skills:
            ws.cell(4, col, skill)
            ws.cell(3, col, category)  # Category header above skill
            ws.column_dimensions[get_column_letter(col)].width = 12
            col += 1
        
        # Merge category header
        if col > category_start_col + 1:
            ws.merge_cells(start_row=3, start_column=category_start_col, 
                          end_row=3, end_column=col-1)
    
    # Apply header styles
    for c in range(1, col):
        # Category row (row 3)
        if ws.cell(3, c).value:
            apply_cell_style(ws.cell(3, c), create_category_style())
        
        # Skill row (row 4)
        apply_cell_style(ws.cell(4, c), create_header_style())
    
    # Freeze panes (freeze first 4 rows and first 4 columns)
    ws.freeze_panes = 'E5'
    
    # Add sample rows for demonstration
    sample_employees = [
        ("John", "Smith", "Process Controls Engineer II", "2026-01-19"),
        ("Jane", "Doe", "APC Engineer III / Senior", "2026-01-19"),
        ("Bob", "Johnson", "Lead Process Controls Engineer", "2026-01-19"),
    ]
    
    for i, (first_name, last_name, role, date) in enumerate(sample_employees, 5):
        ws.cell(i, 1, first_name)
        ws.cell(i, 2, last_name)
        ws.cell(i, 3, role)
        ws.cell(i, 4, date)
        
        # Apply data style to fixed columns
        for c in range(1, 5):
            apply_cell_style(ws.cell(i, c), create_data_style())
        
        # Apply data style to skill columns (leave empty for data entry)
        for c in range(5, col):
            apply_cell_style(ws.cell(i, c), create_data_style())
    
    return ws

def create_role_requirements_sheet(wb):
    """Create Role Requirements sheet with target proficiency levels"""
    ws = wb.create_sheet("Role_Requirements")
    
    # Proficiency scale in Row 1
    scale_text = ws.cell(1, 1, "Scale: 1=Forming (awareness)  |  2=Developing (guided)  |  3=Applying (independent)  |  4=Leading (expert)  |  5=Shaping (SME)")
    scale_text.font = Font(italic=True, size=9, color='366092')
    ws.merge_cells('A1:C1')
    
    # Instructions in Row 2
    instr_cell = ws.cell(2, 1, "Target proficiency levels (1-5) for each role")
    instr_cell.font = Font(italic=True, size=9)
    ws.merge_cells('A2:C2')
    
    # Headers
    row = 4
    col = 1
    
    ws.cell(row, col, "Category")
    ws.column_dimensions[get_column_letter(col)].width = 20
    col += 1
    
    ws.cell(row, col, "Sub-Skill")
    ws.column_dimensions[get_column_letter(col)].width = 40
    col += 1
    
    # Role columns
    for role in ROLES:
        ws.cell(row, col, role)
        ws.column_dimensions[get_column_letter(col)].width = 12
        col += 1
    
    # Apply header style
    for c in range(1, col):
        apply_cell_style(ws.cell(row, c), create_header_style())
    
    # Add skills
    row += 1
    for category, skills in SKILL_TAXONOMY.items():
        for skill in skills:
            ws.cell(row, 1, category)
            ws.cell(row, 2, skill)
            
            # Apply data style
            for c in range(1, col):
                apply_cell_style(ws.cell(row, c), create_data_style())
            
            row += 1
    
    # Freeze panes
    ws.freeze_panes = 'C5'
    
    return ws

def create_summary_sheet(wb):
    """Create summary dashboard sheet"""
    ws = wb.create_sheet("Summary")
    
    # Title
    title_cell = ws.cell(1, 1, "Process Controls Skill Matrix - Summary Dashboard")
    title_cell.font = Font(bold=True, size=16, color='366092')
    ws.merge_cells('A1:D1')
    
    # Metadata
    ws.cell(3, 1, "Generated:").font = Font(bold=True)
    ws.cell(3, 2, datetime.now().strftime("%Y-%m-%d %H:%M"))
    
    ws.cell(4, 1, "Proficiency Scale:").font = Font(bold=True)
    ws.cell(4, 2, "MPC 1-5 Scale")
    
    ws.cell(5, 1, "Total Skills:").font = Font(bold=True)
    total_skills = sum(len(skills) for skills in SKILL_TAXONOMY.values())
    ws.cell(5, 2, total_skills)
    
    ws.cell(6, 1, "Total Roles:").font = Font(bold=True)
    ws.cell(6, 2, len(ROLES))
    
    # Category breakdown
    ws.cell(8, 1, "Skills by Category").font = Font(bold=True, size=12)
    
    headers = ["Category", "Skill Count"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(9, col, header)
        apply_cell_style(cell, create_header_style())
    
    row = 10
    for category, skills in SKILL_TAXONOMY.items():
        ws.cell(row, 1, category)
        ws.cell(row, 2, len(skills))
        
        for col in range(1, 3):
            apply_cell_style(ws.cell(row, col), create_data_style())
        
        row += 1
    
    # Instructions
    ws.cell(row + 2, 1, "Usage Instructions:").font = Font(bold=True, size=12)
    instructions = [
        "1. Review the Proficiency_Scale sheet to understand the 1-5 rating system",
        "2. Use Individual_Assessments sheet to enter employee skill ratings",
        "3. Compare against Role_Requirements for target proficiency levels",
        "4. Use Skill_Dictionary for detailed skill definitions",
        "5. Track progress and identify training needs"
    ]
    
    for i, instr in enumerate(instructions, row + 3):
        ws.cell(i, 1, instr)
    
    # Column widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    
    return ws

def main():
    """Main function to create the enhanced skill matrix workbook"""
    print("Creating Enhanced Process Controls Skill Matrix...")
    
    import os
    from pathlib import Path
    
    output_file = "Process_Controls_Skill_Matrix_Enhanced.xlsx"
    file_exists = Path(output_file).exists()
    
    if file_exists:
        # Workbook exists - preserve Individual_Assessments data
        print(f"⚠️  {output_file} already exists!")
        print("⚠️  Individual_Assessments tab contains manual data.")
        print("⚠️  Run populate_role_requirements.py instead to update role targets only.")
        print("\nTo regenerate from scratch (will lose Individual_Assessments data):")
        print(f"  1. Backup {output_file}")
        print(f"  2. Delete {output_file}")
        print("  3. Run this script again")
        return
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets in order
    print("Creating Summary sheet...")
    create_summary_sheet(wb)
    
    print("Creating Individual Assessments sheet...")
    create_individual_assessment_sheet(wb)
    
    print("Creating Role Requirements sheet...")
    create_role_requirements_sheet(wb)
    
    print("Creating Proficiency Scale sheet...")
    create_proficiency_scale_sheet(wb)
    
    print("Creating Skill Dictionary sheet...")
    create_skill_dictionary_sheet(wb)
    
    # Save workbook
    wb.save(output_file)
    print(f"\n✓ Enhanced workbook created: {output_file}")
    
    # Print summary
    total_skills = sum(len(skills) for skills in SKILL_TAXONOMY.values())
    print(f"\nSummary:")
    print(f"  - Total skills: {total_skills}")
    print(f"  - Categories: {len(SKILL_TAXONOMY)}")
    print(f"  - Roles: {len(ROLES)}")
    print(f"  - Proficiency scale: MPC 1-5")
    print(f"\nSheet structure:")
    print(f"  1. Summary - Overview and instructions")
    print(f"  2. Individual_Assessments - Data entry table (ROWS = People, COLUMNS = Skills)")
    print(f"  3. Role_Requirements - Target proficiency by role")
    print(f"  4. Proficiency_Scale - MPC 1-5 scale reference")
    print(f"  5. Skill_Dictionary - Skill definitions")

if __name__ == "__main__":
    main()
