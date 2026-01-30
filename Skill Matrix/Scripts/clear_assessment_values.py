"""
Clear individual assessment proficiency values while keeping employee names
"""

import openpyxl
from openpyxl.styles import Alignment, Border, Side

def create_data_style():
    """Create data cell style"""
    return {
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def apply_cell_style(cell, style_dict):
    """Apply style dictionary to a cell"""
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

# Load workbook
wb = openpyxl.load_workbook('Process_Controls_Skill_Matrix_Enhanced.xlsx')
ws = wb['Individual_Assessments']

print("Clearing proficiency values from Individual_Assessments...")
print("Keeping employee names, job duties, and assessment dates")

# Find the data rows (start at row 5)
cleared_count = 0
data_style = create_data_style()

for row in range(5, ws.max_row + 1):
    employee_name = ws.cell(row, 1).value
    
    # Skip if no employee name
    if not employee_name:
        continue
    
    # Clear all skill columns (column 4 onwards)
    skills_cleared = 0
    for col in range(4, ws.max_column + 1):
        if ws.cell(row, col).value is not None:
            ws.cell(row, col).value = None
            skills_cleared += 1
        # Maintain formatting
        apply_cell_style(ws.cell(row, col), data_style)
    
    if skills_cleared > 0:
        cleared_count += 1
        print(f"  {employee_name}: cleared {skills_cleared} values")

# Save workbook
wb.save('Process_Controls_Skill_Matrix_Enhanced.xlsx')

print(f"\n✓ Cleared proficiency values for {cleared_count} employees")
print("✓ Employee names, job duties, and dates preserved")
print("\nIndividual_Assessments is now ready for manual data entry")
