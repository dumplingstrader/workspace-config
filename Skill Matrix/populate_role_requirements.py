"""
Populate Role Requirements based on Career Guide and populate from Training Tracker
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import PyPDF2
import re

# Read the enhanced skill matrix
wb = openpyxl.load_workbook('Process_Controls_Skill_Matrix_Enhanced.xlsx')
ws_role_req = wb['Role_Requirements']

# Role-based proficiency targets (based on typical career progression)
# These are baseline targets - can be refined based on Career Guide
ROLE_PROFICIENCY_TARGETS = {
    "Process Controls Engineer I": {
        # Junior level - Forming to Developing
        "Control Platforms": 2,  # Developing on main platforms
        "Process Knowledge": 2,
        "Programming & Control Logic": 1,  # Forming
        "Advanced Process Control": 1,
        "Enterprise Systems": 2,
        "Infrastructure & Networking": 1,
        "Implementation & Commissioning": 2,
        "Engineering Tools & Methods": 1,
        "Professional Development": 1,
        "Business & Leadership": 1
    },
    "Process Controls Engineer II": {
        # Mid-level - Developing to Applying
        "Control Platforms": 3,  # Applying - independent
        "Process Knowledge": 3,
        "Programming & Control Logic": 2,
        "Advanced Process Control": 2,
        "Enterprise Systems": 3,
        "Infrastructure & Networking": 2,
        "Implementation & Commissioning": 3,
        "Engineering Tools & Methods": 2,
        "Professional Development": 2,
        "Business & Leadership": 2
    },
    "Process Controls Engineer III / Senior": {
        # Senior level - Applying to Leading
        "Control Platforms": 4,  # Leading - can guide others
        "Process Knowledge": 4,
        "Programming & Control Logic": 3,
        "Advanced Process Control": 2,
        "Enterprise Systems": 4,
        "Infrastructure & Networking": 3,
        "Implementation & Commissioning": 4,
        "Engineering Tools & Methods": 3,
        "Professional Development": 3,
        "Business & Leadership": 3
    },
    "Lead Process Controls Engineer": {
        # Lead level - Leading to Shaping
        "Control Platforms": 5,  # Shaping - SME
        "Process Knowledge": 4,
        "Programming & Control Logic": 4,
        "Advanced Process Control": 3,
        "Enterprise Systems": 4,
        "Infrastructure & Networking": 4,
        "Implementation & Commissioning": 5,
        "Engineering Tools & Methods": 4,
        "Professional Development": 4,
        "Business & Leadership": 4
    },
    "APC Engineer II": {
        # APC focus - mid level
        "Control Platforms": 3,
        "Process Knowledge": 4,  # Strong process knowledge needed
        "Programming & Control Logic": 3,
        "Advanced Process Control": 4,  # Leading in APC
        "Enterprise Systems": 3,
        "Infrastructure & Networking": 2,
        "Implementation & Commissioning": 3,
        "Engineering Tools & Methods": 3,
        "Professional Development": 2,
        "Business & Leadership": 2
    },
    "APC Engineer III / Senior": {
        # Senior APC
        "Control Platforms": 3,
        "Process Knowledge": 5,  # Expert process knowledge
        "Programming & Control Logic": 4,
        "Advanced Process Control": 5,  # Shaping - APC expert
        "Enterprise Systems": 4,
        "Infrastructure & Networking": 3,
        "Implementation & Commissioning": 4,
        "Engineering Tools & Methods": 4,
        "Professional Development": 3,
        "Business & Leadership": 3
    },
    "Lead APC Engineer": {
        # Lead APC
        "Control Platforms": 4,
        "Process Knowledge": 5,
        "Programming & Control Logic": 4,
        "Advanced Process Control": 5,  # Shaping
        "Enterprise Systems": 4,
        "Infrastructure & Networking": 3,
        "Implementation & Commissioning": 4,
        "Engineering Tools & Methods": 4,
        "Professional Development": 4,
        "Business & Leadership": 4
    },
    "Senior Technologist (P3)": {
        # Technologist - broader technical depth
        "Control Platforms": 4,
        "Process Knowledge": 4,
        "Programming & Control Logic": 4,
        "Advanced Process Control": 4,
        "Enterprise Systems": 4,
        "Infrastructure & Networking": 4,
        "Implementation & Commissioning": 4,
        "Engineering Tools & Methods": 4,
        "Professional Development": 4,
        "Business & Leadership": 3
    },
    "Lead Technologist (P4)": {
        # Lead Technologist
        "Control Platforms": 5,
        "Process Knowledge": 5,
        "Programming & Control Logic": 4,
        "Advanced Process Control": 4,
        "Enterprise Systems": 4,
        "Infrastructure & Networking": 4,
        "Implementation & Commissioning": 5,
        "Engineering Tools & Methods": 5,
        "Professional Development": 4,
        "Business & Leadership": 4
    },
    "Principal Technologist (P5)": {
        # Principal - Strategic leader
        "Control Platforms": 5,
        "Process Knowledge": 5,
        "Programming & Control Logic": 5,
        "Advanced Process Control": 5,
        "Enterprise Systems": 5,
        "Infrastructure & Networking": 4,
        "Implementation & Commissioning": 5,
        "Engineering Tools & Methods": 5,
        "Professional Development": 5,
        "Business & Leadership": 5
    }
}

# Skill categories in order
CATEGORIES = [
    "Control Platforms",
    "Process Knowledge",
    "Programming & Control Logic",
    "Advanced Process Control",
    "Enterprise Systems",
    "Infrastructure & Networking",
    "Implementation & Commissioning",
    "Engineering Tools & Methods",
    "Professional Development",
    "Business & Leadership"
]

# Read Skill Dictionary to get skill list
df_skills = pd.read_excel('Process_Controls_Skill_Matrix_Enhanced.xlsx', sheet_name='Skill_Dictionary')

# Build skill to category mapping
skill_to_category = dict(zip(df_skills['Sub-Skill'], df_skills['Category']))

# Populate Role Requirements
print("Populating Role Requirements...")

# Find where data starts (after headers in row 4)
data_start_row = 5
current_row = data_start_row

for _, skill_row in df_skills.iterrows():
    category = skill_row['Category']
    skill = skill_row['Sub-Skill']
    
    # Write category and skill
    ws_role_req.cell(current_row, 1, category)
    ws_role_req.cell(current_row, 2, skill)
    
    # For each role, populate proficiency target
    for col_idx, role in enumerate(ROLE_PROFICIENCY_TARGETS.keys(), 3):
        # Get category-level target
        category_target = ROLE_PROFICIENCY_TARGETS[role].get(category, 2)
        
        # Adjust for specific skills
        # Platform-specific adjustments
        if category == "Control Platforms":
            if "Honeywell" in skill and "APC" not in role:
                ws_role_req.cell(current_row, col_idx, category_target)
            elif "Allen-Bradley" in skill or "ControlLogix" in skill:
                ws_role_req.cell(current_row, col_idx, category_target)
            elif "Triconex" in skill or "SIS" in skill:
                ws_role_req.cell(current_row, col_idx, max(category_target - 1, 1))
            else:
                ws_role_req.cell(current_row, col_idx, max(category_target - 1, 1))
        
        # APC specific
        elif category == "Advanced Process Control":
            if "APC" in role:
                ws_role_req.cell(current_row, col_idx, category_target)
            else:
                # PC Engineers need less APC depth
                ws_role_req.cell(current_row, col_idx, max(category_target - 1, 1))
        
        # Programming - varies by experience
        elif category == "Programming & Control Logic":
            ws_role_req.cell(current_row, col_idx, category_target)
        
        # Default: use category target
        else:
            ws_role_req.cell(current_row, col_idx, category_target)
    
    current_row += 1

print(f"Populated {current_row - data_start_row} skills across {len(ROLE_PROFICIENCY_TARGETS)} roles")

# Save workbook
wb.save('Process_Controls_Skill_Matrix_Enhanced.xlsx')
print("\n✓ Role Requirements populated successfully!")

# Now read training data and create summary
print("\n" + "="*60)
print("Training Attendance Summary")
print("="*60)

df_training = pd.read_excel('Training_Attendance_Tracker.xlsx', sheet_name='Employee-Course Matrix')
print(f"\nTotal employees tracked: {len(df_training)}")

# Count completions per course
training_columns = [col for col in df_training.columns if 'LAR' in col or 'DCS' in col]
print(f"\nTotal courses tracked: {len(training_columns)}")

print("\nCourse completion summary:")
for course in training_columns:
    completed = df_training[course].notna().sum()
    if completed > 0:
        print(f"  {course}: {completed} completions")

# Map training to skills
training_to_skill_map = {
    "LAR 8901C Tricon System Maintenance": "Triconex SIS",
    "LAR 8921C Tricon CX Custom Maintenance": "Triconex SIS",
    "LAR ControlLogix Fundamentals and Troubleshooting": "Allen-Bradley ControlLogix",
    "LAR Custom PLC-5 & SLC-500": "Allen-Bradley Legacy PLC (PLC-5, SLC500, MicroLogix)",
    "LAR Mark VIe Maintenance Training (L2)": "GE Mark VIe",
    "LAR Python (Session 1)": "Python scripting",
    "LAR Python (Session 2)": "Python scripting",
    "LAR Python (Session 3)": "Python scripting",
    "DCS Training": "Honeywell Experion",
    "LAR SM-4550 Safety Manager: Fundamentals - Maintenance": "Triconex SIS"
}

print("\n" + "="*60)
print("Training mapped to skill categories:")
print("="*60)
for training, skill in training_to_skill_map.items():
    if training in df_training.columns:
        count = df_training[training].notna().sum()
        print(f"  {skill}: {count} trained (via {training})")

print("\n✓ Training analysis complete!")
