"""
Import training attendance data to populate Individual Assessments
Provides starting point for manual refinement
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Map training courses to skills and proficiency levels
# Format: "Training Course Name": ("Skill Name", proficiency_level)
TRAINING_TO_SKILL_MAPPING = {
    # DCS/HMI Training
    "DCS Training": [
        ("Honeywell Experion", 3),  # Applying level from formal training
        ("DCS HMI (Honeywell)", 3)
    ],
    
    # Triconex/SIS Training
    "LAR 8901C Tricon System Maintenance": [
        ("Triconex SIS", 3)  # Applying level
    ],
    "LAR 8921C Tricon CX Custom Maintenance": [
        ("Triconex SIS", 3)
    ],
    "LAR SM-4550 Safety Manager: Fundamentals - Maintenance": [
        ("Triconex SIS", 3),
        ("Integrity (SIS/LOPA)", 2)  # Developing level for related skill
    ],
    
    # Allen-Bradley Training
    "LAR ControlLogix Fundamentals and Troubleshooting": [
        ("Allen-Bradley ControlLogix", 3),
        ("AB HMI (PanelView/FTView)", 2)
    ],
    "LAR Custom PLC-5 & SLC-500": [
        ("Allen-Bradley Legacy PLC (PLC-5, SLC500, MicroLogix)", 3)
    ],
    
    # GE Training
    "LAR Mark VIe Maintenance Training (L2)": [
        ("GE Mark VIe", 3)
    ],
    
    # Python Training
    "LAR Python (Session 1)": [
        ("Python scripting", 2)  # Developing
    ],
    "LAR Python (Session 2)": [
        ("Python scripting", 3)  # Applying after multiple sessions
    ],
    "LAR Python (Session 3)": [
        ("Python scripting", 3)  # Applying
    ],
    
    # Trilogger
    "LAR 8965C Trilogger Custom with SOE Course and PSA Trip Analyzer": [
        ("Triconex SIS", 3),
        ("PI System", 2)  # Related data analysis
    ]
}

def create_data_style():
    """Create data cell style"""
    return {
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def apply_cell_style(cell, style_dict):
    """Apply style dictionary to a cell"""
    if 'font' in style_dict:
        cell.font = style_dict['font']
    if 'fill' in style_dict:
        cell.fill = style_dict['fill']
    if 'alignment' in style_dict:
        cell.alignment = style_dict['alignment']
    if 'border' in style_dict:
        cell.border = style_dict['border']

def main():
    print("Importing training data to Individual Assessments...")
    print("="*70)
    
    # Read training tracker
    df_training = pd.read_excel('Training_Attendance_Tracker.xlsx', sheet_name='Employee-Course Matrix')
    print(f"Found {len(df_training)} employees in training tracker")
    
    # Read enhanced skill matrix
    wb = openpyxl.load_workbook('Process_Controls_Skill_Matrix_Enhanced.xlsx')
    ws_assessments = wb['Individual_Assessments']
    
    # Read skill dictionary to get skill order
    df_skills = pd.read_excel('Process_Controls_Skill_Matrix_Enhanced.xlsx', sheet_name='Skill_Dictionary')
    skill_list = df_skills['Sub-Skill'].tolist()
    
    # Build column mapping for skills (starting at column 4)
    skill_to_col = {}
    col = 4
    for skill in skill_list:
        skill_to_col[skill] = col
        col += 1
    
    print(f"Mapped {len(skill_to_col)} skills to columns")
    
    # Clear existing sample data (rows 5-7)
    for row in range(5, 8):
        for col in range(1, ws_assessments.max_column + 1):
            ws_assessments.cell(row, col).value = None
    
    # Import employees and their training
    current_row = 5
    imported_count = 0
    data_style = create_data_style()
    
    for _, employee_row in df_training.iterrows():
        employee_name = employee_row['Employee']
        job_duty = employee_row.get('Job Duty', '')
        
        # Skip if no name
        if pd.isna(employee_name) or str(employee_name).strip() == '':
            continue
        
        # Write employee info
        ws_assessments.cell(current_row, 1, employee_name)
        ws_assessments.cell(current_row, 2, job_duty)
        ws_assessments.cell(current_row, 3, datetime.now().strftime("%Y-%m-%d"))
        
        # Apply style to fixed columns
        for col in range(1, 4):
            apply_cell_style(ws_assessments.cell(current_row, col), data_style)
        
        # Initialize skill proficiency dictionary for this employee
        employee_skills = {}
        
        # Process each training course
        for training_course, skill_mappings in TRAINING_TO_SKILL_MAPPING.items():
            # Check if employee completed this training
            if training_course in employee_row.index:
                completion_status = employee_row[training_course]
                
                # If completed (has a value like "✓ (Mon YYYY)")
                if pd.notna(completion_status) and '✓' in str(completion_status):
                    # Apply all skill proficiencies for this training
                    for skill_name, proficiency in skill_mappings:
                        # Use highest proficiency if multiple trainings map to same skill
                        if skill_name in employee_skills:
                            employee_skills[skill_name] = max(employee_skills[skill_name], proficiency)
                        else:
                            employee_skills[skill_name] = proficiency
        
        # Write skill proficiencies to worksheet
        for skill_name, proficiency in employee_skills.items():
            if skill_name in skill_to_col:
                col_idx = skill_to_col[skill_name]
                ws_assessments.cell(current_row, col_idx, proficiency)
                apply_cell_style(ws_assessments.cell(current_row, col_idx), data_style)
        
        # Apply data style to all skill columns (including empty ones)
        for col in range(4, 4 + len(skill_list)):
            if ws_assessments.cell(current_row, col).value is None:
                apply_cell_style(ws_assessments.cell(current_row, col), data_style)
        
        imported_count += 1
        current_row += 1
        
        # Print summary for employee
        if len(employee_skills) > 0:
            print(f"  {employee_name}: {len(employee_skills)} skills populated from training")
    
    print("="*70)
    print(f"\n✓ Imported {imported_count} employees")
    
    # Save workbook
    wb.save('Process_Controls_Skill_Matrix_Enhanced.xlsx')
    print("✓ Individual Assessments populated successfully!")
    
    # Print summary statistics
    print("\n" + "="*70)
    print("Summary Statistics")
    print("="*70)
    
    # Count skills populated per category
    category_counts = {}
    for skill_name in employee_skills.keys():
        category = df_skills[df_skills['Sub-Skill'] == skill_name]['Category'].values
        if len(category) > 0:
            cat = category[0]
            category_counts[cat] = category_counts.get(cat, 0) + 1
    
    if category_counts:
        print("\nSkills populated by category (across all employees):")
        for category, count in sorted(category_counts.items()):
            print(f"  {category}: {count} skill entries")
    
    print("\n" + "="*70)
    print("Next Steps:")
    print("="*70)
    print("1. Open Process_Controls_Skill_Matrix_Enhanced.xlsx")
    print("2. Review Individual_Assessments sheet")
    print("3. Manually adjust proficiency levels based on:")
    print("   - Actual on-the-job experience")
    print("   - Project complexity handled")
    print("   - Leadership/mentoring roles")
    print("   - Self-assessments or manager input")
    print("4. Fill in skills not covered by formal training")
    print("5. Update Job Roles (column B) as needed")
    print("\nNote: Training completion = Proficiency 2-3 (baseline)")
    print("      Adjust higher (4-5) for experts/leaders")
    print("      Adjust lower (1) for limited exposure")

if __name__ == "__main__":
    main()
