"""
Create Monthly Report Excel Template
=====================================
Generates a pre-formatted Excel template for monthly value tracking reports.
Template includes:
- Summary metrics (placeholders)
- Issue detail table
- Charts for system breakdown
- Notes section for narrative

Usage:
    python scripts/create_monthly_report_template.py

Outputs: templates/Monthly_Report_Template.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from pathlib import Path


def create_template():
    """Create Excel template with pre-formatted sections."""
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # =========================================
    # SHEET 1: Executive Summary
    # =========================================
    ws_summary = wb.create_sheet('Executive Summary')
    
    # Title
    ws_summary['A1'] = 'MONTHLY TECHNICAL ASSISTANCE REPORT'
    ws_summary['A1'].font = Font(size=16, bold=True, color='1C2833')
    
    # Report info
    ws_summary['A3'] = 'Period:'
    ws_summary['B3'] = '[Month YYYY]'
    ws_summary['B3'].font = Font(color='0000FF')  # Blue = user input
    
    ws_summary['A4'] = 'Prepared By:'
    ws_summary['B4'] = '[Your Name]'
    ws_summary['B4'].font = Font(color='0000FF')
    
    ws_summary['A5'] = 'Date:'
    ws_summary['B5'] = datetime.now().strftime('%Y-%m-%d')
    ws_summary['B5'].font = Font(color='0000FF')
    
    # Key Metrics
    ws_summary['A7'] = 'KEY METRICS'
    ws_summary['A7'].font = Font(size=14, bold=True)
    
    metrics = [
        ('Total Issues Handled', '[#]'),
        ('High Complexity Issues', '[#]'),
        ('Cross-Site Assistance', '[#]'),
        ('Training Requests', '[#]'),
        ('Systems Supported', '[#]'),
        ('Average Response Time', '[hours]')
    ]
    
    row = 8
    for label, placeholder in metrics:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = placeholder
        ws_summary[f'B{row}'].font = Font(color='0000FF')
        row += 1
    
    # Narrative Section
    ws_summary[f'A{row+1}'] = 'EXECUTIVE SUMMARY'
    ws_summary[f'A{row+1}'].font = Font(size=14, bold=True)
    ws_summary[f'A{row+2}'] = '[Write 2-3 paragraphs highlighting key achievements, trends, and challenges]'
    ws_summary[f'A{row+2}'].font = Font(color='0000FF', italic=True)
    ws_summary.merge_cells(f'A{row+2}:D{row+5}')
    ws_summary[f'A{row+2}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # =========================================
    # SHEET 2: Issue Detail
    # =========================================
    ws_detail = wb.create_sheet('Issue Detail')
    
    # Headers
    headers = ['Date', 'System', 'Issue Summary', 'Complexity', 'My Role', 'Impact', 'Outcome']
    for col, header in enumerate(headers, 1):
        cell = ws_detail.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='1C2833', end_color='1C2833', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Sample row
    sample_data = ['2026-01-15', 'DCS System 1', 'Controller communication loss', 'Major', 
                   'Troubleshooting', 'Production impact avoided', 'Restored in 2 hours']
    for col, value in enumerate(sample_data, 1):
        cell = ws_detail.cell(2, col, value)
        cell.font = Font(color='808080', italic=True)  # Gray italic = example
    
    # Instructions row
    ws_detail.cell(3, 1, '[Copy data from generated reports or manual entry]')
    ws_detail.cell(3, 1).font = Font(color='0000FF', italic=True)
    ws_detail.merge_cells('A3:G3')
    
    # Column widths
    widths = [12, 20, 40, 12, 20, 30, 30]
    for col, width in enumerate(widths, 1):
        ws_detail.column_dimensions[chr(64 + col)].width = width
    
    # =========================================
    # SHEET 3: System Breakdown
    # =========================================
    ws_systems = wb.create_sheet('System Breakdown')
    
    ws_systems['A1'] = 'ISSUES BY SYSTEM'
    ws_systems['A1'].font = Font(size=14, bold=True)
    
    # Table headers
    ws_systems['A3'] = 'System'
    ws_systems['B3'] = 'Count'
    ws_systems['C3'] = 'Percentage'
    for col in ['A3', 'B3', 'C3']:
        ws_systems[col].font = Font(bold=True)
        ws_systems[col].fill = PatternFill(start_color='E8F8F5', end_color='E8F8F5', fill_type='solid')
    
    # Sample data
    sample_systems = [
        ('DCS System 1', 10, '35%'),
        ('DCS System 2', 8, '28%'),
        ('PLC Line 1', 6, '21%'),
        ('SIS Safety', 4, '14%'),
        ('[Add more rows as needed]', '', '')
    ]
    
    for row, (system, count, pct) in enumerate(sample_systems, 4):
        ws_systems[f'A{row}'] = system
        ws_systems[f'B{row}'] = count
        ws_systems[f'C{row}'] = pct
        if row == 8:
            ws_systems[f'A{row}'].font = Font(color='0000FF', italic=True)
    
    # Chart instructions
    ws_systems['A10'] = 'INSERT CHART HERE'
    ws_systems['A10'].font = Font(color='0000FF', italic=True, size=12)
    ws_systems.merge_cells('A10:C10')
    ws_systems['A10'].alignment = Alignment(horizontal='center')
    
    ws_systems['A11'] = '[Recommended: Pie chart or bar chart from System Breakdown data]'
    ws_systems['A11'].font = Font(color='808080', italic=True)
    ws_systems.merge_cells('A11:C11')
    ws_systems['A11'].alignment = Alignment(horizontal='center')
    
    ws_systems.column_dimensions['A'].width = 30
    ws_systems.column_dimensions['B'].width = 15
    ws_systems.column_dimensions['C'].width = 15
    
    # =========================================
    # SHEET 4: Action Items / Recommendations
    # =========================================
    ws_actions = wb.create_sheet('Action Items')
    
    ws_actions['A1'] = 'RECOMMENDATIONS & ACTION ITEMS'
    ws_actions['A1'].font = Font(size=14, bold=True)
    
    # Table headers
    headers = ['Priority', 'Description', 'Rationale', 'Owner', 'Due Date', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws_actions.cell(3, col, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
    
    # Sample row
    sample_actions = [
        ('High', 'Schedule DCS training for operations', 'Multiple DCS questions this month', 
         '[Manager Name]', '[Date]', 'Open'),
        ('Medium', 'Review PLC program for Line 1', 'Recurring communication errors', 
         '[Engineer Name]', '[Date]', 'Open')
    ]
    
    for row, action in enumerate(sample_actions, 4):
        for col, value in enumerate(action, 1):
            cell = ws_actions.cell(row, col, value)
            cell.font = Font(color='808080', italic=True)
    
    # Add more rows instruction
    ws_actions.cell(6, 1, '[Add more action items as needed]')
    ws_actions.cell(6, 1).font = Font(color='0000FF', italic=True)
    ws_actions.merge_cells('A6:F6')
    
    # Column widths
    widths = [10, 35, 35, 20, 12, 12]
    for col, width in enumerate(widths, 1):
        ws_actions.column_dimensions[chr(64 + col)].width = width
    
    # =========================================
    # SHEET 5: Notes
    # =========================================
    ws_notes = wb.create_sheet('Notes')
    
    ws_notes['A1'] = 'ADDITIONAL NOTES'
    ws_notes['A1'].font = Font(size=14, bold=True)
    
    notes_sections = [
        'TRENDS OBSERVED',
        'CHALLENGES ENCOUNTERED',
        'SUCCESSES AND WINS',
        'IMPROVEMENT OPPORTUNITIES'
    ]
    
    row = 3
    for section in notes_sections:
        ws_notes[f'A{row}'] = section
        ws_notes[f'A{row}'].font = Font(bold=True, color='2E4053')
        ws_notes[f'A{row+1}'] = '[Write your observations here]'
        ws_notes[f'A{row+1}'].font = Font(color='0000FF', italic=True)
        ws_notes.merge_cells(f'A{row+1}:D{row+3}')
        ws_notes[f'A{row+1}'].alignment = Alignment(wrap_text=True, vertical='top')
        row += 5
    
    ws_notes.column_dimensions['A'].width = 80
    
    # Save template
    output_path = Path('templates/Monthly_Report_Template.xlsx')
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb.save(output_path)
    
    print("\n" + "="*60)
    print("Monthly Report Template Created")
    print("="*60)
    print(f"\n✓ Template saved: {output_path}")
    print("\nTemplate includes:")
    print("  • Executive Summary (metrics + narrative)")
    print("  • Issue Detail (table for all issues)")
    print("  • System Breakdown (chart-ready data)")
    print("  • Action Items (recommendations)")
    print("  • Notes (observations and trends)")
    print("\n📘 Blue text = User inputs (fill these in)")
    print("📊 Gray italic text = Example data (replace with actual)")
    print("="*60)


if __name__ == '__main__':
    create_template()
