#!/usr/bin/env python3
"""
Generate Monthly Summary Report

Reads master.json and generates a comprehensive Excel report for the specified month.
Output includes summary statistics, system breakdown, stream analysis, and detailed issue list.
"""

import pandas as pd
import json
from pathlib import Path
from datetime import datetime
import argparse
from collections import Counter


def load_master_data(master_file):
    """Load and return master.json data."""
    with open(master_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def filter_by_month(records, year_month):
    """Filter records for specified month (format: YYYY-MM)."""
    return [r for r in records if r['date'] and r['date'].startswith(year_month)]


def generate_summary_stats(records):
    """Generate summary statistics."""
    total = len(records)
    
    # By stream
    streams = Counter(r['stream'] for r in records if r['stream'])
    
    # By system (group DCS variants, PLC variants, etc.)
    systems = Counter()
    for r in records:
        if r['system']:
            # Group by main category
            sys = r['system']
            if 'DCS' in sys or 'Experion' in sys:
                systems['DCS'] += 1
            elif 'PLC' in sys or 'ControlLogix' in sys or 'Studio' in sys:
                systems['PLC'] += 1
            elif 'SIS' in sys or 'Triconex' in sys:
                systems['SIS'] += 1
            elif 'Alarm' in sys:
                systems['Alarm'] += 1
            elif 'Network' in sys:
                systems['Network'] += 1
            elif 'HMI' in sys:
                systems['HMI'] += 1
            else:
                systems['Other'] += 1
    
    # By complexity
    complexity = Counter(r['complexity'] for r in records if r['complexity'])
    
    # By business impact
    impact = Counter(r['business_impact'] for r in records if r['business_impact'])
    
    # By resolution
    resolution = Counter(r['resolution'] for r in records if r['resolution'])
    
    return {
        'total': total,
        'streams': dict(streams),
        'systems': dict(systems),
        'complexity': dict(complexity),
        'impact': dict(impact),
        'resolution': dict(resolution)
    }


def create_summary_sheet(stats, month_name):
    """Create summary statistics sheet."""
    rows = []
    
    rows.append(['PC VALUE TRACKER — MONTHLY SUMMARY', ''])
    rows.append(['Month:', month_name])
    rows.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')])
    rows.append(['', ''])
    
    rows.append(['TOTAL ISSUES', stats['total']])
    rows.append(['', ''])
    
    # By Stream
    rows.append(['BY STREAM', 'Count'])
    for stream, count in sorted(stats['streams'].items(), key=lambda x: -x[1]):
        rows.append([stream, count])
    rows.append(['', ''])
    
    # By System
    rows.append(['BY SYSTEM', 'Count'])
    for system, count in sorted(stats['systems'].items(), key=lambda x: -x[1]):
        rows.append([system, count])
    rows.append(['', ''])
    
    # By Complexity
    rows.append(['BY COMPLEXITY', 'Count'])
    for comp, count in sorted(stats['complexity'].items(), key=lambda x: -x[1]):
        rows.append([comp, count])
    rows.append(['', ''])
    
    # By Business Impact
    rows.append(['BY BUSINESS IMPACT', 'Count'])
    for imp, count in sorted(stats['impact'].items(), key=lambda x: -x[1]):
        rows.append([imp, count])
    rows.append(['', ''])
    
    # By Resolution
    if stats['resolution']:
        rows.append(['BY RESOLUTION', 'Count'])
        for res, count in sorted(stats['resolution'].items(), key=lambda x: -x[1]):
            rows.append([res, count])
    
    return pd.DataFrame(rows, columns=['Metric', 'Value'])


def create_detail_sheet(records):
    """Create detailed issue list."""
    df_data = []
    
    for r in records:
        df_data.append({
            'Date': r['date'],
            'System': r['system'],
            'Summary': r['summary'],
            'Stream': r['stream'],
            'Complexity': r['complexity'],
            'Resolution': r['resolution'],
            'Business Impact': r['business_impact']
        })
    
    return pd.DataFrame(df_data)


def create_stream_sheet(records):
    """Create stream-focused analysis."""
    # Group by stream
    by_stream = {}
    for r in records:
        stream = r['stream'] or 'Unclassified'
        if stream not in by_stream:
            by_stream[stream] = []
        by_stream[stream].append(r)
    
    rows = []
    
    for stream in sorted(by_stream.keys()):
        stream_records = by_stream[stream]
        rows.append([f' === {stream.upper()} ===', '', '', '', '', ''])  # Space prefix prevents Excel formula error
        rows.append(['Date', 'System', 'Summary', 'Complexity', 'Resolution', 'Impact'])
        
        for r in sorted(stream_records, key=lambda x: x['date'] or ''):
            rows.append([
                r['date'],
                r['system'],
                r['summary'][:100],  # Truncate long summaries
                r['complexity'],
                r['resolution'],
                r['business_impact']
            ])
        
        rows.append(['', '', '', '', '', ''])  # Blank row between streams
    
    return pd.DataFrame(rows, columns=['Date', 'System', 'Summary', 'Complexity', 'Resolution', 'Impact'])


def create_system_sheet(records):
    """Create system-focused analysis."""
    # Group by system category
    by_system = {}
    for r in records:
        system = r['system'] or 'Unspecified'
        # Group into categories
        if 'DCS' in system or 'Experion' in system:
            cat = 'DCS'
        elif 'PLC' in system or 'ControlLogix' in system:
            cat = 'PLC'
        elif 'SIS' in system or 'Triconex' in system:
            cat = 'SIS'
        elif 'Alarm' in system:
            cat = 'Alarm'
        elif 'Network' in system:
            cat = 'Network'
        else:
            cat = 'Other'
        
        if cat not in by_system:
            by_system[cat] = []
        by_system[cat].append(r)
    
    rows = []
    
    for cat in sorted(by_system.keys()):
        cat_records = by_system[cat]
        rows.append([f' === {cat.upper()} ===', '', '', '', ''])  # Space prefix prevents Excel formula error
        rows.append(['Date', 'System', 'Summary', 'Stream', 'Complexity'])
        
        for r in sorted(cat_records, key=lambda x: x['date'] or ''):
            rows.append([
                r['date'],
                r['system'],
                r['summary'][:100],
                r['stream'],
                r['complexity']
            ])
        
        rows.append(['', '', '', '', ''])
    
    return pd.DataFrame(rows, columns=['Date', 'System', 'Summary', 'Stream', 'Complexity'])


def generate_report(master_file, month, output_dir, verbose=False):
    """Generate the full monthly report."""
    
    # Load data
    if verbose:
        print(f"Loading data from {master_file}...")
    records = load_master_data(master_file)
    
    # Filter by month
    month_records = filter_by_month(records, month)
    
    if not month_records:
        print(f"⚠️  No records found for {month}")
        return False
    
    if verbose:
        print(f"Found {len(month_records)} records for {month}")
    
    # Generate statistics
    stats = generate_summary_stats(month_records)
    
    # Create output filename
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    month_name = datetime.strptime(month, '%Y-%m').strftime('%B %Y')
    output_file = output_path / f"Monthly_Report_{month}.xlsx"
    
    # Create Excel file with multiple sheets
    if verbose:
        print(f"Creating Excel report...")
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Summary sheet
        df_summary = create_summary_sheet(stats, month_name)
        df_summary.to_excel(writer, sheet_name='Summary', index=False, header=False)
        
        # Detail sheet
        df_detail = create_detail_sheet(month_records)
        df_detail.to_excel(writer, sheet_name='All Issues', index=False)
        
        # Stream analysis
        df_streams = create_stream_sheet(month_records)
        df_streams.to_excel(writer, sheet_name='By Stream', index=False, header=False)
        
        # System analysis
        df_systems = create_system_sheet(month_records)
        df_systems.to_excel(writer, sheet_name='By System', index=False, header=False)
        
        # Format Summary sheet
        workbook = writer.book
        summary_sheet = writer.sheets['Summary']
        
        # Set column widths
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 15
        
        # Bold headers (title, section headers, not data rows)
        from openpyxl.styles import Font, Alignment
        for row_idx in range(1, summary_sheet.max_row + 1):
            cell_value = summary_sheet.cell(row=row_idx, column=1).value
            if cell_value and isinstance(cell_value, str):
                # Bold only these specific headers
                if cell_value in ['PC VALUE TRACKER — MONTHLY SUMMARY', 'TOTAL ISSUES', 
                                  'BY STREAM', 'BY SYSTEM', 'BY COMPLEXITY', 
                                  'BY BUSINESS IMPACT', 'BY RESOLUTION']:
                    summary_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
    
    print(f"\n✅ Report generated!")
    print(f"   Month: {month_name}")
    print(f"   Issues: {len(month_records)}")
    print(f"   Output: {output_file}")
    
    return True


def main():
    parser = argparse.ArgumentParser(description='Generate monthly summary report')
    parser.add_argument('--master', default='data/master.json', help='Master data file')
    parser.add_argument('--month', required=True, help='Month to report (YYYY-MM)')
    parser.add_argument('--output-dir', default='output/monthly', help='Output directory')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')
    
    args = parser.parse_args()
    
    # Validate month format
    try:
        datetime.strptime(args.month, '%Y-%m')
    except ValueError:
        print(f"Error: Month must be in YYYY-MM format (e.g., 2026-01)")
        exit(1)
    
    success = generate_report(args.master, args.month, args.output_dir, args.verbose)
    
    if not success:
        exit(1)


if __name__ == '__main__':
    main()
