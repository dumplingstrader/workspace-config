#!/usr/bin/env python3
"""
Generate Stream-Specific Scorecards

Creates targeted reports for specific audiences based on the six streams.
Each scorecard is designed with a specific ask for a specific audience.

Scorecards:
1. Project Handoff Scorecard → Projects team
2. Technical Debt Register → Capital Planning
3. Diagnostic Value Report → Operations/Maintenance leadership
4. After-Hours Burden Report → HR/Leadership
5. Applications Dashboard → IT/OT coordination
6. Day-to-Day Dashboard → PC Supervisors

Usage:
    # Generate all scorecards for a quarter
    python generate_stream_scorecards.py --quarter 2025-Q4

    # Generate specific scorecard
    python generate_stream_scorecards.py --quarter 2025-Q4 --scorecard project
    python generate_stream_scorecards.py --quarter 2025-Q4 --scorecard legacy
    python generate_stream_scorecards.py --quarter 2025-Q4 --scorecard diagnostic
    python generate_stream_scorecards.py --quarter 2025-Q4 --scorecard after-hours
    python generate_stream_scorecards.py --quarter 2025-Q4 --scorecard applications
    python generate_stream_scorecards.py --quarter 2025-Q4 --scorecard day-to-day
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from collections import Counter
import argparse

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Styling constants
HEADER_FILL = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="40B4B4", end_color="40B4B4", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="006666")
METRIC_FONT = Font(bold=True, size=24, color="008080")
LABEL_FONT = Font(size=10, color="666666")
HIGHLIGHT_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def get_quarter_data(data, quarter):
    """Filter data for a specific quarter."""
    year, q = quarter.split('-Q')
    year = int(year)
    q = int(q)

    quarter_months = {
        1: ['01', '02', '03'],
        2: ['04', '05', '06'],
        3: ['07', '08', '09'],
        4: ['10', '11', '12'],
    }
    months = quarter_months[q]

    return [r for r in data if r.get('date') and
            r['date'][:4] == str(year) and r['date'][5:7] in months]


def style_header_row(ws, row, cols):
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def add_metric_box(ws, row, col, value, label):
    """Add a metric box with large value and small label."""
    ws.cell(row=row, column=col, value=value).font = METRIC_FONT
    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    ws.cell(row=row + 1, column=col, value=label).font = LABEL_FONT
    ws.cell(row=row + 1, column=col).alignment = Alignment(horizontal='center')


def generate_project_scorecard(data, quarter, output_path):
    """
    Generate Project Handoff Scorecard.

    Audience: AMP Project Manager, Project Engineers, Capital Projects leadership
    The Ask: PC sign-off required before project phases close
    """
    # Filter for Project stream
    project_data = [r for r in data if r.get('stream') == 'Project']

    wb = Workbook()
    ws = wb.active
    ws.title = "Project Handoff Scorecard"

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"PROJECT HANDOFF SCORECARD — {quarter}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    # Key metrics row
    row = 3
    ws.cell(row=row, column=1, value="Issues Found Post-Cutover:")
    ws.cell(row=row, column=2, value=len(project_data))
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="008080")

    # Calculate estimated hours (based on complexity)
    complexity_hours = {'Major': 8, 'High': 4, 'Moderate': 2, 'Medium': 2, 'Low': 1, 'Quick': 0.5}
    total_hours = sum(complexity_hours.get(r.get('complexity', 'Low'), 1) for r in project_data)

    row = 4
    ws.cell(row=row, column=1, value="Estimated PC Hours to Remediate:")
    ws.cell(row=row, column=2, value=f"{total_hours:.0f}")
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="008080")

    # Resolution breakdown
    row = 6
    ws.cell(row=row, column=1, value="Resolution Status:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1
    resolutions = Counter(r.get('resolution', 'Unknown') for r in project_data)
    for res, count in resolutions.most_common():
        ws.cell(row=row, column=1, value=f"  {res}:")
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Business impact breakdown
    row += 1
    ws.cell(row=row, column=1, value="Business Impact:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1
    impacts = Counter(r.get('business_impact', 'Unknown') for r in project_data)
    for impact, count in impacts.most_common():
        ws.cell(row=row, column=1, value=f"  {impact}:")
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Issue list
    row += 2
    ws.cell(row=row, column=1, value="Issues Detail:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1

    headers = ['Date', 'System', 'Summary', 'Complexity', 'Resolution', 'Impact']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    style_header_row(ws, row, len(headers))
    row += 1

    for record in sorted(project_data, key=lambda x: x.get('date', '')):
        ws.cell(row=row, column=1, value=record.get('date', ''))
        ws.cell(row=row, column=2, value=record.get('system', ''))
        ws.cell(row=row, column=3, value=record.get('summary', '')[:80])
        ws.cell(row=row, column=4, value=record.get('complexity', ''))
        ws.cell(row=row, column=5, value=record.get('resolution', ''))
        ws.cell(row=row, column=6, value=record.get('business_impact', ''))
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Recommendation
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="RECOMMENDATION: Require PC acceptance testing before project phase close.")
    ws['A' + str(row)].font = Font(bold=True, color="CC0000")

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

    wb.save(output_path)
    return len(project_data)


def generate_legacy_scorecard(data, quarter, output_path):
    """
    Generate Technical Debt Register.

    Audience: Capital Planning, Reliability Engineering, Leadership
    The Ask: Fund the obsolescence backlog before catastrophic failure
    """
    # Filter for Legacy Modernization stream
    legacy_data = [r for r in data if r.get('stream') == 'Legacy Modernization']

    wb = Workbook()
    ws = wb.active
    ws.title = "Technical Debt Register"

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"TECHNICAL DEBT REGISTER — {quarter}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    # Summary metrics
    row = 3
    ws.cell(row=row, column=1, value="Legacy System Issues:")
    ws.cell(row=row, column=2, value=len(legacy_data))
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="CC0000")

    # System breakdown
    row = 5
    ws.cell(row=row, column=1, value="Issues by System:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1

    systems = Counter()
    for r in legacy_data:
        sys_name = r.get('system', 'Unknown')
        # Extract key identifier
        if 'TDC' in sys_name.upper():
            systems['TDC'] += 1
        elif 'PLC-5' in sys_name.upper() or 'PLC5' in sys_name.upper():
            systems['PLC-5'] += 1
        elif 'SLC' in sys_name.upper():
            systems['SLC-500'] += 1
        else:
            systems[sys_name.split(':')[0].strip()] += 1

    for sys, count in systems.most_common():
        ws.cell(row=row, column=1, value=f"  {sys}:")
        ws.cell(row=row, column=2, value=count)
        if 'PLC-5' in sys or 'TDC' in sys:
            ws.cell(row=row, column=1).fill = CRITICAL_FILL
            ws.cell(row=row, column=2).fill = CRITICAL_FILL
        row += 1

    # Risk assessment
    row += 2
    ws.cell(row=row, column=1, value="Risk Assessment:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1

    headers = ['System', 'Issue Count', 'Risk Level', 'Notes']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    style_header_row(ws, row, len(headers))
    row += 1

    for sys, count in systems.most_common():
        risk = 'CRITICAL' if count >= 3 or 'PLC-5' in sys else ('HIGH' if count >= 2 else 'MEDIUM')
        notes = 'End of life, no vendor support' if 'PLC-5' in sys or 'TDC' in sys else 'Monitor closely'

        ws.cell(row=row, column=1, value=sys)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=risk)
        ws.cell(row=row, column=4, value=notes)

        if risk == 'CRITICAL':
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = CRITICAL_FILL

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Issue detail
    row += 2
    ws.cell(row=row, column=1, value="Issue Detail:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1

    headers = ['Date', 'System', 'Summary', 'Resolution', 'Impact']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    style_header_row(ws, row, len(headers))
    row += 1

    for record in sorted(legacy_data, key=lambda x: x.get('date', '')):
        ws.cell(row=row, column=1, value=record.get('date', ''))
        ws.cell(row=row, column=2, value=record.get('system', ''))
        ws.cell(row=row, column=3, value=record.get('summary', '')[:80])
        ws.cell(row=row, column=4, value=record.get('resolution', ''))
        ws.cell(row=row, column=5, value=record.get('business_impact', ''))
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Recommendation
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="RECOMMENDATION: Fund obsolescence replacements before catastrophic failure.")
    ws['A' + str(row)].font = Font(bold=True, color="CC0000")

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12

    wb.save(output_path)
    return len(legacy_data)


def generate_diagnostic_scorecard(data, quarter, output_path):
    """
    Generate Diagnostic Value Report.

    Audience: Leadership, Operations management, Maintenance management
    The Ask: Recognize diagnostic services or restore Maintenance troubleshooting
    """
    # Filter for Diagnostic stream
    diag_data = [r for r in data if r.get('stream') == 'Diagnostic']

    wb = Workbook()
    ws = wb.active
    ws.title = "Diagnostic Value"

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = f"DIAGNOSTIC VALUE REPORT — {quarter}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    # Metrics
    row = 3
    ws.cell(row=row, column=1, value="Issues Investigated & Handed Off:")
    ws.cell(row=row, column=2, value=len(diag_data))
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="008080")

    # Estimate hours
    complexity_hours = {'Major': 8, 'High': 4, 'Moderate': 2, 'Medium': 2, 'Low': 1, 'Quick': 0.5}
    total_hours = sum(complexity_hours.get(r.get('complexity', 'Low'), 1) for r in diag_data)

    row = 4
    ws.cell(row=row, column=1, value="Estimated PC Hours on Non-PC Issues:")
    ws.cell(row=row, column=2, value=f"{total_hours:.0f}")
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="008080")

    # Value statement
    row = 6
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="VALUE STATEMENT: We provide diagnostic services for issues that aren't our responsibility.")
    ws['A' + str(row)].font = Font(italic=True)
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="This saves other groups time but is invisible in our workload metrics.")
    ws['A' + str(row)].font = Font(italic=True)

    # Issue detail
    row += 3
    ws.cell(row=row, column=1, value="Diagnostic Issues:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1

    headers = ['Date', 'System', 'Summary', 'Complexity', 'Resolution']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    style_header_row(ws, row, len(headers))
    row += 1

    for record in sorted(diag_data, key=lambda x: x.get('date', '')):
        ws.cell(row=row, column=1, value=record.get('date', ''))
        ws.cell(row=row, column=2, value=record.get('system', ''))
        ws.cell(row=row, column=3, value=record.get('summary', '')[:80])
        ws.cell(row=row, column=4, value=record.get('complexity', ''))
        ws.cell(row=row, column=5, value=record.get('resolution', ''))
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Recommendation
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="RECOMMENDATION: Recognize diagnostic services or restore Maintenance first-line troubleshooting.")
    ws['A' + str(row)].font = Font(bold=True, color="CC0000")

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15

    wb.save(output_path)
    return len(diag_data)


def generate_afterhours_scorecard(data, quarter, output_path):
    """
    Generate After-Hours Burden Report.

    Audience: Leadership, HR, Supervisors
    The Ask: Fair on-call compensation for responsiveness
    """
    # Filter for After-Hours stream
    ah_data = [r for r in data if r.get('stream') == 'After-Hours']

    wb = Workbook()
    ws = wb.active
    ws.title = "After-Hours Burden"

    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = f"AFTER-HOURS BURDEN REPORT — {quarter}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    # Metrics
    row = 3
    ws.cell(row=row, column=1, value="Total Call-Outs:")
    ws.cell(row=row, column=2, value=len(ah_data))
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="CC0000")

    # Estimate hours
    complexity_hours = {'Major': 8, 'High': 4, 'Moderate': 2, 'Medium': 2, 'Low': 1, 'Quick': 0.5}
    total_hours = sum(complexity_hours.get(r.get('complexity', 'Low'), 1) for r in ah_data)

    row = 4
    ws.cell(row=row, column=1, value="Estimated Off-Hours Worked:")
    ws.cell(row=row, column=2, value=f"{total_hours:.0f} hours")
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="CC0000")

    # Resolution breakdown
    row = 6
    ws.cell(row=row, column=1, value="Resolution:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1
    resolutions = Counter(r.get('resolution', 'Unknown') for r in ah_data)
    for res, count in resolutions.most_common():
        ws.cell(row=row, column=1, value=f"  {res}:")
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Impact breakdown
    row += 1
    ws.cell(row=row, column=1, value="Business Impact:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1
    impacts = Counter(r.get('business_impact', 'Unknown') for r in ah_data)
    for impact, count in impacts.most_common():
        ws.cell(row=row, column=1, value=f"  {impact}:")
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Burden statement
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value=f"BURDEN STATEMENT: Team members responded to {len(ah_data)} after-hours calls this quarter,")
    ws['A' + str(row)].font = Font(italic=True)
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value=f"working approximately {total_hours:.0f} hours outside normal schedule.")
    ws['A' + str(row)].font = Font(italic=True)

    # Issue detail
    row += 3
    ws.cell(row=row, column=1, value="After-Hours Issues:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1

    headers = ['Date', 'System', 'Summary', 'Complexity', 'Resolution']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    style_header_row(ws, row, len(headers))
    row += 1

    for record in sorted(ah_data, key=lambda x: x.get('date', '')):
        ws.cell(row=row, column=1, value=record.get('date', ''))
        ws.cell(row=row, column=2, value=record.get('system', ''))
        ws.cell(row=row, column=3, value=record.get('summary', '')[:80])
        ws.cell(row=row, column=4, value=record.get('complexity', ''))
        ws.cell(row=row, column=5, value=record.get('resolution', ''))
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Recommendation
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1, value="RECOMMENDATION: Implement fair on-call compensation reflecting actual after-hours burden.")
    ws['A' + str(row)].font = Font(bold=True, color="CC0000")

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15

    wb.save(output_path)
    return len(ah_data)


def generate_applications_scorecard(data, quarter, output_path):
    """
    Generate Applications Dashboard.

    Audience: Operations leadership, Reliability Engineering, IT/OT coordination
    The Ask: Recognize application support as a distinct skill set
    """
    # Filter for Applications stream
    app_data = [r for r in data if r.get('stream') == 'Applications']

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications Dashboard"

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"APPLICATIONS DASHBOARD — {quarter}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    # Metrics
    row = 3
    ws.cell(row=row, column=1, value="Total Application Issues:")
    ws.cell(row=row, column=2, value=len(app_data))
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="008080")

    # Estimate hours
    complexity_hours = {'Major': 8, 'High': 4, 'Moderate': 2, 'Medium': 2, 'Low': 1, 'Quick': 0.5}
    total_hours = sum(complexity_hours.get(r.get('complexity', 'Low'), 1) for r in app_data)

    row = 4
    ws.cell(row=row, column=1, value="Estimated Hours Invested:")
    ws.cell(row=row, column=2, value=f"{total_hours:.0f}")
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="008080")

    # Application breakdown
    row = 6
    ws.cell(row=row, column=1, value="Issues by Application:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1

    apps = Counter()
    for r in app_data:
        sys_name = r.get('system', '').upper()
        summary = r.get('summary', '').upper()
        combined = sys_name + ' ' + summary

        if 'DYNAMO' in combined or 'ACM' in combined or 'APO' in combined:
            apps['Alarm Management (DynAMo/ACM/APO)'] += 1
        elif 'INTEGRITY' in combined:
            apps['Asset Integrity'] += 1
        elif 'PHD' in combined or 'HISTORIAN' in combined or 'PI ' in combined:
            apps['Historian (PHD/PI)'] += 1
        else:
            apps['Other Applications'] += 1

    for app, count in apps.most_common():
        pct = round(count / len(app_data) * 100) if app_data else 0
        ws.cell(row=row, column=1, value=f"  {app}:")
        ws.cell(row=row, column=2, value=f"{count} ({pct}%)")
        row += 1

    # Business impact
    row += 1
    ws.cell(row=row, column=1, value="Business Impact:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1
    impacts = Counter(r.get('business_impact', 'Unknown') for r in app_data)
    for impact, count in impacts.most_common():
        ws.cell(row=row, column=1, value=f"  {impact}:")
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Issue detail
    row += 2
    ws.cell(row=row, column=1, value="Application Issues:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1

    headers = ['Date', 'System', 'Summary', 'Complexity', 'Resolution', 'Impact']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    style_header_row(ws, row, len(headers))
    row += 1

    for record in sorted(app_data, key=lambda x: x.get('date', '')):
        ws.cell(row=row, column=1, value=record.get('date', ''))
        ws.cell(row=row, column=2, value=record.get('system', ''))
        ws.cell(row=row, column=3, value=record.get('summary', '')[:60])
        ws.cell(row=row, column=4, value=record.get('complexity', ''))
        ws.cell(row=row, column=5, value=record.get('resolution', ''))
        ws.cell(row=row, column=6, value=record.get('business_impact', ''))
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Recommendation
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="RECOMMENDATION: Recognize application support as a distinct skill set requiring dedicated expertise.")
    ws['A' + str(row)].font = Font(bold=True, color="006666")

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

    wb.save(output_path)
    return len(app_data)


def generate_daytoday_scorecard(data, quarter, output_path):
    """
    Generate Day-to-Day Dashboard.

    Audience: PC Supervisors, Department Manager, Leadership
    The Ask: Resource planning based on workload data
    """
    # Filter for Day-to-Day stream
    dd_data = [r for r in data if r.get('stream') == 'Day-to-Day']

    wb = Workbook()
    ws = wb.active
    ws.title = "Day-to-Day Dashboard"

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"DAY-TO-DAY SUPPORT DASHBOARD — {quarter}"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')

    # Metrics
    row = 3
    ws.cell(row=row, column=1, value="Total Issues:")
    ws.cell(row=row, column=2, value=len(dd_data))
    ws.cell(row=row, column=2).font = Font(bold=True, size=14, color="008080")

    # System breakdown
    row = 5
    ws.cell(row=row, column=1, value="By System:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1

    systems = Counter()
    for r in dd_data:
        sys_name = r.get('system', 'Unknown')
        sys_simple = sys_name.split(':')[0].strip() if ':' in sys_name else sys_name
        systems[sys_simple] += 1

    for sys, count in systems.most_common(8):
        pct = round(count / len(dd_data) * 100) if dd_data else 0
        ws.cell(row=row, column=1, value=f"  {sys}:")
        ws.cell(row=row, column=2, value=f"{count} ({pct}%)")
        row += 1

    # Complexity breakdown
    row += 1
    ws.cell(row=row, column=1, value="By Complexity:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1
    complexity = Counter(r.get('complexity', 'Unknown') for r in dd_data)
    for comp, count in complexity.most_common():
        pct = round(count / len(dd_data) * 100) if dd_data else 0
        ws.cell(row=row, column=1, value=f"  {comp}:")
        ws.cell(row=row, column=2, value=f"{count} ({pct}%)")
        row += 1

    # Resolution breakdown
    row += 1
    ws.cell(row=row, column=1, value="By Resolution:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1
    resolutions = Counter(r.get('resolution', 'Unknown') for r in dd_data)
    for res, count in resolutions.most_common():
        pct = round(count / len(dd_data) * 100) if dd_data else 0
        ws.cell(row=row, column=1, value=f"  {res}:")
        ws.cell(row=row, column=2, value=f"{count} ({pct}%)")
        row += 1

    # Issue detail
    row += 2
    ws.cell(row=row, column=1, value="Issue Detail:")
    ws['A' + str(row)].font = Font(bold=True)
    row += 1

    headers = ['Date', 'System', 'Summary', 'Complexity', 'Resolution', 'Impact']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=header)
    style_header_row(ws, row, len(headers))
    row += 1

    for record in sorted(dd_data, key=lambda x: x.get('date', '')):
        ws.cell(row=row, column=1, value=record.get('date', ''))
        ws.cell(row=row, column=2, value=record.get('system', ''))
        ws.cell(row=row, column=3, value=record.get('summary', '')[:60])
        ws.cell(row=row, column=4, value=record.get('complexity', ''))
        ws.cell(row=row, column=5, value=record.get('resolution', ''))
        ws.cell(row=row, column=6, value=record.get('business_impact', ''))
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Summary
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value=f"BASELINE: {len(dd_data)} routine support issues this quarter. Use for resource planning.")
    ws['A' + str(row)].font = Font(bold=True, color="006666")

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12

    wb.save(output_path)
    return len(dd_data)


def main():
    parser = argparse.ArgumentParser(
        description='Generate Stream-Specific Scorecards',
        epilog='Example: python generate_stream_scorecards.py --quarter 2025-Q4'
    )
    parser.add_argument('--quarter', required=True, help='Quarter (e.g., 2025-Q4)')
    parser.add_argument('--scorecard', choices=['project', 'legacy', 'diagnostic', 'after-hours', 'applications', 'day-to-day', 'all'],
                       default='all', help='Which scorecard to generate (default: all)')
    parser.add_argument('--input', default='data/master.json', help='Input JSON file')
    parser.add_argument('--output-dir', default='output/scorecards', help='Output directory')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose output')

    args = parser.parse_args()

    # Load data
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"[ERROR] Input file not found: {args.input}")
        exit(1)

    if args.verbose:
        print(f"Loading data from {args.input}...")

    with open(input_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Filter for quarter
    quarter_data = get_quarter_data(data, args.quarter)

    if not quarter_data:
        print(f"[ERROR] No data found for {args.quarter}")
        exit(1)

    if args.verbose:
        print(f"Found {len(quarter_data)} records for {args.quarter}")

    # Create output directory
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # Generate scorecards
    generators = {
        'project': ('Project_Handoff_Scorecard', generate_project_scorecard),
        'legacy': ('Technical_Debt_Register', generate_legacy_scorecard),
        'diagnostic': ('Diagnostic_Value_Report', generate_diagnostic_scorecard),
        'after-hours': ('AfterHours_Burden_Report', generate_afterhours_scorecard),
        'applications': ('Applications_Dashboard', generate_applications_scorecard),
        'day-to-day': ('DayToDay_Dashboard', generate_daytoday_scorecard),
    }

    if args.scorecard == 'all':
        scorecards_to_gen = list(generators.keys())
    else:
        scorecards_to_gen = [args.scorecard]

    results = []
    for sc_name in scorecards_to_gen:
        filename, gen_func = generators[sc_name]
        output_path = output_dir / f"{filename}_{args.quarter}.xlsx"

        if args.verbose:
            print(f"Generating {filename}...")

        count = gen_func(quarter_data, args.quarter, output_path)
        results.append((sc_name, count, output_path))

        if args.verbose:
            print(f"  -> {count} records")

    print(f"\n[SUCCESS] Scorecards generated for {args.quarter}!")
    print(f"   Output directory: {output_dir}")
    print()
    for sc_name, count, path in results:
        print(f"   {sc_name}: {count} records -> {path.name}")


if __name__ == '__main__':
    main()
